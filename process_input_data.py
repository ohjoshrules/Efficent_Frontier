"""
Input Data Processor
====================

This script processes input files from data/input/ and generates:
1. TODO.md from PDF files (archives old TODOs)
2. Processed Excel workbook with:
   - All original sheets copied
   - New "LN_Returns" sheet with log returns, stats, and covariance matrix

Usage:
    python process_input_data.py
"""

import os
import sys
import shutil
import logging
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Tuple
from scipy.optimize import minimize

# PDF reading support
try:
    import fitz  # PyMuPDF
    PDF_SUPPORT = "pymupdf"
except ImportError:
    try:
        from PyPDF2 import PdfReader
        PDF_SUPPORT = "pypdf2"
    except ImportError:
        PDF_SUPPORT = None

# Excel support
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# =============================================================================
# CONFIGURATION
# =============================================================================

PROJECT_ROOT = Path(__file__).parent
DATA_DIR = PROJECT_ROOT / "data"
INPUT_DIR = DATA_DIR / "input"
EXCEL_INPUT_DIR = INPUT_DIR / "excel"
PDF_INPUT_DIR = INPUT_DIR / "pdf"
PROCESSED_DIR = DATA_DIR / "processed"
ARCHIVE_DIR = DATA_DIR / "archived"
LOGS_DIR = PROJECT_ROOT / "logs"

# Ensure directories exist
PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
LOGS_DIR.mkdir(parents=True, exist_ok=True)

# Global variables to store processed data
PROCESSED_FILE_PATH = None
STORED_COV_MATRIX = None
STORED_MEANS = None


# =============================================================================
# LOGGING SETUP
# =============================================================================

def setup_logging() -> logging.Logger:
    """Set up logging to both console and file."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = LOGS_DIR / f"process_input_{timestamp}.log"

    # Create logger
    logger = logging.getLogger("InputProcessor")
    logger.setLevel(logging.INFO)

    # Clear any existing handlers
    logger.handlers = []

    # File handler
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_format = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(file_format)

    # Console handler
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_format = logging.Formatter('%(message)s')
    console_handler.setFormatter(console_format)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


# =============================================================================
# PDF READING
# =============================================================================

def read_pdf(file_path: Path) -> str:
    """
    Read text content from a PDF file.

    Args:
        file_path: Path to PDF file

    Returns:
        Extracted text content
    """
    if PDF_SUPPORT is None:
        return f"[PDF reading not available - install PyMuPDF: pip install PyMuPDF]\n\nFile: {file_path.name}"

    try:
        if PDF_SUPPORT == "pymupdf":
            import fitz
            doc = fitz.open(str(file_path))
            text = ""
            for page_num, page in enumerate(doc):
                text += f"\n--- Page {page_num + 1} ---\n"
                text += page.get_text()
            doc.close()
            return text
        else:
            from PyPDF2 import PdfReader
            reader = PdfReader(str(file_path))
            text = ""
            for page_num, page in enumerate(reader.pages):
                text += f"\n--- Page {page_num + 1} ---\n"
                text += page.extract_text() or ""
            return text
    except Exception as e:
        return f"[Error reading PDF: {e}]\n\nFile: {file_path.name}"


def process_pdfs_to_todo() -> str:
    """
    Process all PDFs in the input folder and create TODO content.

    Returns:
        Markdown content for TODO.md
    """
    pdf_files = list(PDF_INPUT_DIR.glob("*.pdf"))

    if not pdf_files:
        return "# TODO\n\nNo PDF files found in input folder.\n"

    content = f"# TODO\n\n"
    content += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
    content += f"---\n\n"

    for pdf_file in sorted(pdf_files):
        content += f"## {pdf_file.stem}\n\n"
        content += f"**Source:** `{pdf_file.name}`\n\n"

        # Read PDF content
        pdf_text = read_pdf(pdf_file)

        # Add content in a collapsible section or as-is
        content += "### Content\n\n"
        content += "```\n"
        content += pdf_text.strip()
        content += "\n```\n\n"

        # Add TODO checklist section
        content += "### Tasks\n\n"
        content += "- [ ] Review content\n"
        content += "- [ ] Extract key requirements\n"
        content += "- [ ] Complete analysis\n"
        content += "\n---\n\n"

    return content


def archive_old_todo():
    """Archive existing TODO.md if it exists."""
    todo_path = PROJECT_ROOT / "TODO.md"

    if todo_path.exists():
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        archive_name = f"TODO_{timestamp}.md"
        archive_path = ARCHIVE_DIR / archive_name

        shutil.move(str(todo_path), str(archive_path))
        print(f"Archived old TODO.md to: {archive_path}")


# =============================================================================
# EXCEL PROCESSING
# =============================================================================

def find_prices_sheet(workbook) -> Optional[str]:
    """
    Find the sheet called 'Prices' (or contains 'price' in name).

    Args:
        workbook: openpyxl Workbook object

    Returns:
        Sheet name or None
    """
    # First look for exact match "Prices"
    for sheet_name in workbook.sheetnames:
        if sheet_name.lower() == 'prices':
            return sheet_name

    # Then look for sheets containing 'price'
    for sheet_name in workbook.sheetnames:
        if 'price' in sheet_name.lower():
            return sheet_name

    # If no match, return None
    return None


def load_price_data(excel_path: Path) -> Tuple[Optional[pd.DataFrame], Optional[str]]:
    """
    Load price data from Excel file, looking for 'Prices' sheet.

    Args:
        excel_path: Path to Excel file

    Returns:
        Tuple of (DataFrame, sheet_name) or (None, None)
    """
    try:
        # Get all sheet names
        xl = pd.ExcelFile(excel_path)
        target_sheet = None

        # Look for exact "Prices" sheet first
        for sheet in xl.sheet_names:
            if sheet.lower() == 'prices':
                target_sheet = sheet
                break

        # Then look for sheets containing 'price'
        if target_sheet is None:
            for sheet in xl.sheet_names:
                if 'price' in sheet.lower():
                    target_sheet = sheet
                    break

        # Default to first sheet if no Prices sheet found
        if target_sheet is None and xl.sheet_names:
            print(f"  Warning: No 'Prices' sheet found, using first sheet")
            target_sheet = xl.sheet_names[0]

        if target_sheet is None:
            return None, None

        # Try reading with header on different rows to find the right one
        # First try header=1 (common case where row 0 is empty or title)
        df = pd.read_excel(excel_path, sheet_name=target_sheet, header=1)

        # Check if we got valid column names (not "Unnamed")
        unnamed_count = sum(1 for col in df.columns if 'unnamed' in str(col).lower())

        if unnamed_count > len(df.columns) / 2:
            # Too many unnamed columns, try header=0
            df = pd.read_excel(excel_path, sheet_name=target_sheet, header=0)

            # Still bad? Try header=2
            unnamed_count = sum(1 for col in df.columns if 'unnamed' in str(col).lower())
            if unnamed_count > len(df.columns) / 2:
                df = pd.read_excel(excel_path, sheet_name=target_sheet, header=2)

        return df, target_sheet

    except Exception as e:
        print(f"Error loading Excel: {e}")
        return None, None


def detect_date_column(df: pd.DataFrame) -> Optional[str]:
    """Find the date column in the DataFrame."""
    for col in df.columns:
        col_str = str(col).lower()
        if 'date' in col_str:
            return col
        # Check if column contains datetime-like values
        if df[col].dtype == 'datetime64[ns]':
            return col

    # Check first column if it looks like dates
    first_col = df.columns[0]
    try:
        pd.to_datetime(df[first_col].iloc[:5])
        return first_col
    except:
        pass

    return None


def detect_date_order(dates: pd.Series) -> str:
    """
    Detect if dates are sorted newest first or oldest first.

    Args:
        dates: Series of dates

    Returns:
        'newest_first' or 'oldest_first'

    Raises:
        ValueError: If dates are not sorted or cannot be determined
    """
    # Convert to datetime if not already
    try:
        dates = pd.to_datetime(dates)
    except Exception as e:
        raise ValueError(f"Could not parse dates: {e}")

    # Check first and last valid dates
    first_date = dates.iloc[0]
    last_date = dates.iloc[-1]

    if pd.isna(first_date) or pd.isna(last_date):
        raise ValueError("First or last date is NaN - cannot determine order")

    if first_date > last_date:
        return 'newest_first'
    elif first_date < last_date:
        return 'oldest_first'
    else:
        raise ValueError("First and last dates are equal - cannot determine order")


def ensure_newest_first(prices_df: pd.DataFrame, date_col: str, logger=None) -> pd.DataFrame:
    """
    Ensure data is sorted with newest dates first.

    Args:
        prices_df: DataFrame with price data
        date_col: Name of the date column
        logger: Optional logger for messages

    Returns:
        DataFrame sorted newest first
    """
    dates = prices_df[date_col]

    try:
        order = detect_date_order(dates)

        if order == 'oldest_first':
            if logger:
                logger.warning("  WARNING:  Data is sorted OLDEST FIRST - reversing to NEWEST FIRST")
            # Reverse the dataframe to make it newest first
            prices_df = prices_df.iloc[::-1].reset_index(drop=True)
            if logger:
                logger.info(f"  Data reordered: {prices_df[date_col].iloc[0]} (newest) to {prices_df[date_col].iloc[-1]} (oldest)")
        else:
            if logger:
                logger.info(f"  [OK] Data is correctly sorted NEWEST FIRST")
                logger.info(f"    First date: {dates.iloc[0]} (newest)")
                logger.info(f"    Last date:  {dates.iloc[-1]} (oldest)")

    except ValueError as e:
        if logger:
            logger.error(f"  WARNING:  Could not determine date order: {e}")
            logger.warning("  Assuming data is sorted correctly (newest first)")

    return prices_df


def compute_ln_returns(prices_df: pd.DataFrame, date_col: str, logger=None) -> pd.DataFrame:
    """
    Compute log returns from price data.

    LN Return = ln(P_t / P_{t-1}) * 100 (as percentage)

    This function automatically detects and handles both date orderings:
    - NEWEST FIRST: Most recent date at row 0 (e.g., June 2024 at top)
    - OLDEST FIRST: Oldest date at row 0 (e.g., June 2018 at top)

    Data is always converted to NEWEST FIRST before calculation.

    Example: For June 2024 return with June at row 0, May at row 1:
        June return = ln(June price / May price) = ln(row 0 / row 1)

    Args:
        prices_df: DataFrame with price data (any date order - will be auto-detected)
        date_col: Name of the date column
        logger: Optional logger for status messages

    Returns:
        DataFrame with log returns (sorted newest first)
    """
    # Make a copy to avoid modifying original
    prices_df = prices_df.copy()

    # Ensure data is sorted newest first
    prices_df = ensure_newest_first(prices_df, date_col, logger)

    # Separate date and price columns
    dates = prices_df[date_col].copy()

    # Get numeric columns (tickers)
    price_cols = [col for col in prices_df.columns if col != date_col]
    prices = prices_df[price_cols].copy()

    # Convert to numeric, handling any errors
    for col in prices.columns:
        prices[col] = pd.to_numeric(prices[col], errors='coerce')

    # Data is now guaranteed to be sorted NEWEST FIRST
    # For the return at date t, we need: ln(Price_t / Price_{t-1})
    # With newest first: Price_t is in current row, Price_{t-1} is in next row
    # So we use: prices / prices.shift(-1)
    #   Row 0 (June): June / May = correct June return
    #   Row 1 (May): May / April = correct May return
    #   Last row: oldest / NaN = NaN (drop this)
    ln_returns = np.log(prices / prices.shift(-1))

    # Convert to percentage
    ln_returns = ln_returns * 100

    # Drop the last row (NaN from shift(-1)) - this is the oldest date
    ln_returns = ln_returns.iloc[:-1].reset_index(drop=True)
    dates = dates.iloc[:-1].reset_index(drop=True)

    # Combine date with returns
    result = pd.DataFrame({date_col: dates})
    for col in price_cols:
        result[col] = ln_returns[col].values

    return result


def create_ln_returns_sheet(wb: Workbook, ln_returns: pd.DataFrame, date_col: str,
                            means: pd.Series = None, stds: pd.Series = None,
                            cov_matrix: pd.DataFrame = None):
    """
    Create the LN_Returns sheet with returns, stats, and covariance matrix.

    Args:
        wb: openpyxl Workbook
        ln_returns: DataFrame with log returns
        date_col: Name of date column
        means: Series of mean returns
        stds: Series of standard deviations
        cov_matrix: Covariance matrix DataFrame
    """
    # Create new sheet
    if "LN_Returns" in wb.sheetnames:
        del wb["LN_Returns"]
    ws = wb.create_sheet("LN_Returns")

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    section_font = Font(bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Get ticker columns (all except date)
    tickers = [col for col in ln_returns.columns if col != date_col]
    n_tickers = len(tickers)

    # ==========================================================================
    # SECTION 1: LN Returns Table
    # ==========================================================================
    ws['A1'] = "LOG RETURNS (LN)"
    ws['A1'].font = section_font

    ws['A2'] = "Formula: LN(P_t / P_{t-1}) * 100 (%)"

    # Headers (row 4)
    header_row = 4
    ws.cell(row=header_row, column=1, value=date_col).font = header_font_white
    ws.cell(row=header_row, column=1).fill = header_fill
    ws.cell(row=header_row, column=1).border = thin_border

    for j, ticker in enumerate(tickers):
        cell = ws.cell(row=header_row, column=j+2, value=ticker)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # Data rows
    data_start_row = header_row + 1
    for i, (_, row) in enumerate(ln_returns.iterrows()):
        current_row = data_start_row + i

        # Date
        date_val = row[date_col]
        if pd.notna(date_val):
            ws.cell(row=current_row, column=1, value=date_val).border = thin_border

        # Returns
        for j, ticker in enumerate(tickers):
            val = row[ticker]
            if pd.notna(val):
                cell = ws.cell(row=current_row, column=j+2, value=val)
                cell.number_format = '0.0000'
                cell.border = thin_border

    data_end_row = data_start_row + len(ln_returns) - 1

    # ==========================================================================
    # SECTION 2: Statistics (Mean and Std Dev)
    # ==========================================================================
    stats_row = data_end_row + 3

    ws.cell(row=stats_row, column=1, value="STATISTICS").font = section_font

    # Mean row
    mean_row = stats_row + 1
    ws.cell(row=mean_row, column=1, value="Mean").font = header_font
    ws.cell(row=mean_row, column=1).border = thin_border

    # Std Dev row (percentage form)
    std_row = mean_row + 1
    ws.cell(row=std_row, column=1, value="Std Dev (%)").font = header_font
    ws.cell(row=std_row, column=1).border = thin_border

    # Variance row (decimal form)
    var_row = std_row + 1
    ws.cell(row=var_row, column=1, value="Variance").font = header_font
    ws.cell(row=var_row, column=1).border = thin_border

    # Compute and write stats
    returns_only = ln_returns[tickers]
    means = returns_only.mean()
    stds = returns_only.std(ddof=0)  # Population std dev
    # Variance in decimal form: (std_pct / 100)^2
    variances = (stds / 100) ** 2

    for j, ticker in enumerate(tickers):
        # Mean (percentage)
        mean_cell = ws.cell(row=mean_row, column=j+2, value=means[ticker])
        mean_cell.number_format = '0.0000'
        mean_cell.border = thin_border

        # Std Dev (percentage)
        std_cell = ws.cell(row=std_row, column=j+2, value=stds[ticker])
        std_cell.number_format = '0.0000'
        std_cell.border = thin_border

        # Variance (decimal form)
        var_cell = ws.cell(row=var_row, column=j+2, value=variances[ticker])
        var_cell.number_format = '0.000000'
        var_cell.border = thin_border

    # ==========================================================================
    # SECTION 3: Covariance Matrix
    # ==========================================================================
    cov_header_row = var_row + 3

    ws.cell(row=cov_header_row, column=1, value="COVARIANCE MATRIX").font = section_font

    # Compute covariance matrix (population) in DECIMAL form
    # Returns are in percentage form (e.g., 3.47 = 3.47%), so divide by 100 first
    # Covariance of decimals = Covariance of percentages / 10000 (since cov scales with square of factor)
    returns_decimal = returns_only / 100  # Convert to decimal form (0.0347 = 3.47%)
    cov_matrix = returns_decimal.cov() * (len(returns_decimal) - 1) / len(returns_decimal)

    # Column headers for covariance matrix
    cov_col_header_row = cov_header_row + 1
    ws.cell(row=cov_col_header_row, column=1, value="").border = thin_border

    for j, ticker in enumerate(tickers):
        cell = ws.cell(row=cov_col_header_row, column=j+2, value=ticker)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # Covariance values
    cov_data_start = cov_col_header_row + 1
    for i, ticker_i in enumerate(tickers):
        current_row = cov_data_start + i

        # Row label
        cell = ws.cell(row=current_row, column=1, value=ticker_i)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

        # Covariance values
        for j, ticker_j in enumerate(tickers):
            cov_val = cov_matrix.loc[ticker_i, ticker_j]
            cell = ws.cell(row=current_row, column=j+2, value=cov_val)
            cell.number_format = '0.000000'
            cell.border = thin_border

    # ==========================================================================
    # SECTION 4: Correlation Matrix (bonus)
    # ==========================================================================
    corr_header_row = cov_data_start + n_tickers + 2

    ws.cell(row=corr_header_row, column=1, value="CORRELATION MATRIX").font = section_font

    # Compute correlation matrix
    corr_matrix = returns_only.corr()

    # Column headers
    corr_col_header_row = corr_header_row + 1
    ws.cell(row=corr_col_header_row, column=1, value="").border = thin_border

    for j, ticker in enumerate(tickers):
        cell = ws.cell(row=corr_col_header_row, column=j+2, value=ticker)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # Correlation values
    corr_data_start = corr_col_header_row + 1
    for i, ticker_i in enumerate(tickers):
        current_row = corr_data_start + i

        # Row label
        cell = ws.cell(row=current_row, column=1, value=ticker_i)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

        # Correlation values
        for j, ticker_j in enumerate(tickers):
            corr_val = corr_matrix.loc[ticker_i, ticker_j]
            cell = ws.cell(row=current_row, column=j+2, value=corr_val)
            cell.number_format = '0.0000'
            cell.border = thin_border

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    for j in range(n_tickers):
        ws.column_dimensions[get_column_letter(j+2)].width = 12


def create_summary_sheet(wb: Workbook, prices_df: pd.DataFrame, ln_returns: pd.DataFrame,
                         date_col: str, tickers: List[str], means: pd.Series,
                         stds: pd.Series, cov_matrix: pd.DataFrame):
    """
    Create a comprehensive Summary sheet with all data.

    Args:
        wb: openpyxl Workbook
        prices_df: DataFrame with price data
        ln_returns: DataFrame with log returns
        date_col: Name of date column
        tickers: List of ticker names
        means: Series of mean returns
        stds: Series of standard deviations
        cov_matrix: Covariance matrix DataFrame
    """
    # Create new sheet
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary", 0)  # Insert at beginning

    # Styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    section_font = Font(bold=True, size=12)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    n_tickers = len(tickers)
    current_row = 1

    # =========================================================================
    # TITLE
    # =========================================================================
    ws['A1'] = "PORTFOLIO ANALYSIS SUMMARY"
    ws['A1'].font = title_font
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    current_row = 4

    # =========================================================================
    # SECTION 1: PRICE DATA
    # =========================================================================
    ws.cell(row=current_row, column=1, value="SECTION 1: PRICE DATA").font = section_font
    current_row += 1
    ws.cell(row=current_row, column=1, value=f"Source: Prices sheet | {len(prices_df)} observations")
    current_row += 2

    # Headers
    ws.cell(row=current_row, column=1, value=date_col).font = header_font_white
    ws.cell(row=current_row, column=1).fill = header_fill
    ws.cell(row=current_row, column=1).border = thin_border

    for j, ticker in enumerate(tickers):
        cell = ws.cell(row=current_row, column=j+2, value=ticker)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    current_row += 1

    # Price data
    for _, row in prices_df.iterrows():
        ws.cell(row=current_row, column=1, value=row[date_col]).border = thin_border
        for j, ticker in enumerate(tickers):
            val = row[ticker]
            if pd.notna(val):
                cell = ws.cell(row=current_row, column=j+2, value=val)
                cell.number_format = '0.0000000'
                cell.border = thin_border
        current_row += 1

    current_row += 2

    # =========================================================================
    # SECTION 2: LN RETURNS
    # =========================================================================
    ws.cell(row=current_row, column=1, value="SECTION 2: LOG RETURNS (LN)").font = section_font
    current_row += 1
    ws.cell(row=current_row, column=1, value="Formula: LN(P_t / P_{t-1}) × 100 (as %)")
    current_row += 2

    # Headers
    ws.cell(row=current_row, column=1, value=date_col).font = header_font_white
    ws.cell(row=current_row, column=1).fill = header_fill
    ws.cell(row=current_row, column=1).border = thin_border

    for j, ticker in enumerate(tickers):
        cell = ws.cell(row=current_row, column=j+2, value=ticker)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    current_row += 1

    # LN Returns data
    for _, row in ln_returns.iterrows():
        ws.cell(row=current_row, column=1, value=row[date_col]).border = thin_border
        for j, ticker in enumerate(tickers):
            val = row[ticker]
            if pd.notna(val):
                cell = ws.cell(row=current_row, column=j+2, value=val)
                cell.number_format = '0.0000000'
                cell.border = thin_border
        current_row += 1

    current_row += 2

    # =========================================================================
    # SECTION 3: STATISTICS
    # =========================================================================
    ws.cell(row=current_row, column=1, value="SECTION 3: STATISTICS").font = section_font
    current_row += 2

    # Header row
    ws.cell(row=current_row, column=1, value="Statistic").font = header_font_white
    ws.cell(row=current_row, column=1).fill = header_fill
    ws.cell(row=current_row, column=1).border = thin_border

    for j, ticker in enumerate(tickers):
        cell = ws.cell(row=current_row, column=j+2, value=ticker)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    current_row += 1

    # Mean row
    ws.cell(row=current_row, column=1, value="Mean (%)").font = header_font
    ws.cell(row=current_row, column=1).border = thin_border
    for j, ticker in enumerate(tickers):
        cell = ws.cell(row=current_row, column=j+2, value=means[ticker])
        cell.number_format = '0.0000000'
        cell.border = thin_border
    current_row += 1

    # Std Dev row
    ws.cell(row=current_row, column=1, value="Std Dev (%)").font = header_font
    ws.cell(row=current_row, column=1).border = thin_border
    for j, ticker in enumerate(tickers):
        cell = ws.cell(row=current_row, column=j+2, value=stds[ticker])
        cell.number_format = '0.0000000'
        cell.border = thin_border
    current_row += 1

    # Variance row (decimal form: (std_pct / 100)^2)
    ws.cell(row=current_row, column=1, value="Variance").font = header_font
    ws.cell(row=current_row, column=1).border = thin_border
    for j, ticker in enumerate(tickers):
        # Convert std from percentage to decimal, then square for variance
        variance_decimal = (stds[ticker] / 100) ** 2
        cell = ws.cell(row=current_row, column=j+2, value=variance_decimal)
        cell.number_format = '0.000000'
        cell.border = thin_border

    current_row += 3

    # =========================================================================
    # SECTION 4: COVARIANCE MATRIX
    # =========================================================================
    ws.cell(row=current_row, column=1, value="SECTION 4: COVARIANCE MATRIX").font = section_font
    current_row += 1
    ws.cell(row=current_row, column=1, value="Population covariance (divides by N)")
    current_row += 2

    # Header row
    ws.cell(row=current_row, column=1, value="").border = thin_border
    for j, ticker in enumerate(tickers):
        cell = ws.cell(row=current_row, column=j+2, value=ticker)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    current_row += 1

    # Covariance matrix data
    for i, ticker_i in enumerate(tickers):
        cell = ws.cell(row=current_row, column=1, value=ticker_i)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

        for j, ticker_j in enumerate(tickers):
            cov_val = cov_matrix.loc[ticker_i, ticker_j]
            cell = ws.cell(row=current_row, column=j+2, value=cov_val)
            cell.number_format = '0.0000000'
            cell.border = thin_border
        current_row += 1

    current_row += 2

    # =========================================================================
    # SECTION 5: CORRELATION MATRIX
    # =========================================================================
    ws.cell(row=current_row, column=1, value="SECTION 5: CORRELATION MATRIX").font = section_font
    current_row += 2

    # Compute correlation
    corr_matrix = ln_returns[tickers].corr()

    # Header row
    ws.cell(row=current_row, column=1, value="").border = thin_border
    for j, ticker in enumerate(tickers):
        cell = ws.cell(row=current_row, column=j+2, value=ticker)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
    current_row += 1

    # Correlation matrix data
    for i, ticker_i in enumerate(tickers):
        cell = ws.cell(row=current_row, column=1, value=ticker_i)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

        for j, ticker_j in enumerate(tickers):
            corr_val = corr_matrix.loc[ticker_i, ticker_j]
            cell = ws.cell(row=current_row, column=j+2, value=corr_val)
            cell.number_format = '0.0000000'
            cell.border = thin_border
        current_row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    for j in range(n_tickers):
        ws.column_dimensions[get_column_letter(j+2)].width = 14


def get_excel_col_letter(col_idx: int) -> str:
    """Convert 0-based column index to Excel column letter (A, B, ..., Z, AA, AB, ...)."""
    result = ""
    while col_idx >= 0:
        result = chr(col_idx % 26 + ord('A')) + result
        col_idx = col_idx // 26 - 1
    return result


def log_dataframe(logger: logging.Logger, df: pd.DataFrame, title: str, decimals: int = 7,
                  start_row: int = 3, start_col: int = 1, show_excel_refs: bool = True):
    """
    Log a DataFrame with all rows and columns visible, with Excel-style row/column references.

    Args:
        logger: Logger instance
        df: DataFrame to log
        title: Title for the section
        decimals: Number of decimal places
        start_row: Excel row number where data starts (default 3 for typical layout)
        start_col: Excel column index where data starts, 0=A, 1=B (default 1 for B)
        show_excel_refs: Whether to show Excel row/column references
    """
    logger.info("")
    logger.info("=" * 100)
    logger.info(f"  {title}")
    logger.info("=" * 100)

    if show_excel_refs:
        # Show column mapping
        col_letters = []
        logger.info("")
        logger.info("  Excel Column Reference:")
        col_info = "    "
        for i, col in enumerate(df.columns):
            col_letter = get_excel_col_letter(start_col + i)
            col_letters.append(col_letter)
            col_info += f"{col_letter}={col}  "
            if (i + 1) % 8 == 0:  # Line break every 8 columns
                logger.info(col_info)
                col_info = "    "
        if col_info.strip():
            logger.info(col_info)
        logger.info("")
        logger.info(f"  Data starts at row {start_row} in Excel (Row 1 = headers, Row 2 = column names)")
        logger.info("")

    # Set pandas display options for full output
    with pd.option_context('display.max_rows', None,
                           'display.max_columns', None,
                           'display.width', None,
                           'display.float_format', f'{{:.{decimals}f}}'.format):

        if show_excel_refs:
            # Add Excel row numbers to the output
            df_copy = df.copy()
            excel_rows = [f"Row {start_row + i}" for i in range(len(df))]
            df_copy.insert(0, 'Excel Row', excel_rows)
            df_str = df_copy.to_string(index=True)
        else:
            df_str = df.to_string()

        for line in df_str.split('\n'):
            logger.info(line)

    logger.info("")


def process_single_excel(excel_file: Path, logger: logging.Logger):
    """
    Process a single Excel file.
    Creates new processed workbook with LN_Returns sheet.

    Args:
        excel_file: Path to the Excel file
        logger: Logger instance
    """
    try:
        # Load the workbook
        wb = load_workbook(excel_file)
        logger.info(f"  Sheets found: {wb.sheetnames}")

        # Load price/return data using pandas
        df, source_sheet = load_price_data(excel_file)

        if df is None:
            logger.error(f"  Could not load data from {excel_file.name}")
            return

        logger.info(f"  Using sheet: {source_sheet}")
        logger.info(f"  Data shape: {df.shape}")

        # Find date column
        date_col = detect_date_column(df)
        if date_col is None:
            logger.warning(f"  Could not find date column")
            # Use first column as date
            date_col = df.columns[0]
            logger.info(f"  Using first column as date: {date_col}")

        # Get ticker columns (exclude unnamed columns)
        tickers = [col for col in df.columns if col != date_col and 'unnamed' not in str(col).lower()]
        logger.info(f"  Tickers found: {len(tickers)}")
        logger.info(f"  Tickers: {tickers}")

        if not tickers:
            logger.error(f"  No ticker columns found")
            return

        # Check if data is prices or returns
        numeric_cols = df[tickers].select_dtypes(include=[np.number])
        if numeric_cols.empty:
            logger.error(f"  No numeric columns found")
            return

        sample_mean = numeric_cols.mean().mean()
        logger.info(f"  Sample mean: {sample_mean:.7f}")

        # Assume data from "Prices" sheet is price data
        # Compute LN returns from prices
        logger.info(f"  Processing PRICE data - computing LN returns")

        # Clean the data
        prices_df = df[[date_col] + tickers].copy()
        prices_df = prices_df.dropna()
        logger.info(f"  Clean data rows: {len(prices_df)}")

        # =====================================================================
        # LOG ALL PRICE DATA
        # =====================================================================
        log_dataframe(logger, prices_df, f"PRICE DATA FROM '{source_sheet}' SHEET - ALL {len(prices_df)} ROWS",
                      start_row=3, start_col=0)  # A=Date, B=first ticker

        # Compute LN returns (auto-detects and ensures newest-first sorting)
        ln_returns = compute_ln_returns(prices_df, date_col, logger)
        logger.info(f"  Computed {len(ln_returns)} LN return observations")

        # Filter to 5 years of data through June 2024 (as per exam instructions)
        # Returns are dated by END month, so:
        # - Uses PRICES from June 2019 to June 2024 (61 prices)
        # - Gives RETURNS from July 2019 to June 2024 (60 returns)
        # Return dated July 2019 = ln(July price / June price)
        ln_returns[date_col] = pd.to_datetime(ln_returns[date_col])
        start_date = pd.Timestamp('2019-07-01')  # First return: July 2019 (uses June & July prices)
        end_date = pd.Timestamp('2024-06-30')    # Last return: June 2024 (uses May & June prices)
        ln_returns = ln_returns[(ln_returns[date_col] >= start_date) & (ln_returns[date_col] <= end_date)]
        ln_returns = ln_returns.reset_index(drop=True)
        logger.info(f"  Filtered to 5 years: {len(ln_returns)} returns (Jul 2019 - Jun 2024, using prices Jun 2019 - Jun 2024)")

        # =====================================================================
        # LOG ALL LN RETURNS
        # =====================================================================
        log_dataframe(logger, ln_returns, f"LN RETURNS (%) - ALL {len(ln_returns)} ROWS",
                      start_row=3, start_col=0)  # A=Date, B=first ticker

        # Compute statistics
        returns_only = ln_returns[tickers]
        means = returns_only.mean()
        stds = returns_only.std(ddof=0)  # Population std dev
        # Variance in decimal form: (std_pct / 100)^2
        variances = (stds / 100) ** 2
        n_obs = len(returns_only)

        # Create stats DataFrame
        stats_df = pd.DataFrame({
            'Ticker': tickers,
            'Mean (%)': means.values,
            'Std Dev (%)': stds.values,
            'Variance': variances.values
        }).set_index('Ticker')

        # =====================================================================
        # LOG MEAN AND STD DEV WITH EXCEL INSTRUCTIONS
        # =====================================================================
        logger.info("")
        logger.info("=" * 100)
        logger.info("  STATISTICS: MEAN AND STANDARD DEVIATION")
        logger.info("=" * 100)
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  MATHEMATICAL FORMULAS:                                                         │")
        logger.info("  │                                                                                 │")
        logger.info("  │  MEAN (μ):     μ = (1/N) × Σᵢ rᵢ  =  Σᵢ rᵢ / N                                 │")
        logger.info("  │                                                                                 │")
        logger.info("  │  STD DEV (σ):  σ = √[ (1/N) × Σᵢ (rᵢ - μ)² ]                                   │")
        logger.info("  │                                                                                 │")
        logger.info("  │  VARIABLE DEFINITIONS:                                                          │")
        logger.info(f"  │    N  = Number of observations = {n_obs}                                         │")
        logger.info("  │    rᵢ = Return for observation i (each monthly LN return)                       │")
        logger.info("  │    μ  = Mean (average) return = Greek letter 'mu'                               │")
        logger.info("  │    σ  = Standard deviation (volatility) = Greek letter 'sigma'                  │")
        logger.info("  │    Σᵢ = Summation over all i from 1 to N = Greek letter 'Sigma'                 │")
        logger.info("  │    √  = Square root                                                             │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")
        # Get sample cell values for examples
        sample_ticker = tickers[0]  # First ticker (e.g., SPY)
        first_return = returns_only[sample_ticker].iloc[0]  # B3 value
        last_return = returns_only[sample_ticker].iloc[-1]  # Last row value

        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  EXCEL FORMULAS (with cell values):                                             │")
        logger.info("  │                                                                                 │")
        logger.info(f"  │  Cell B3 = {first_return:.7f} (first {sample_ticker} return, June 2024)                   │")
        logger.info(f"  │  Cell B{n_obs + 2} = {last_return:.7f} (last {sample_ticker} return)                            │")
        logger.info("  │                                                                                 │")
        logger.info("  │  MEAN (Average):                                                                │")
        logger.info(f"  │    =AVERAGE(B3:B{n_obs + 2})                                                     │")
        logger.info(f"  │    Result for {sample_ticker}: {means[sample_ticker]:.7f}%                                     │")
        logger.info("  │    Drag across all ticker columns                                               │")
        logger.info("  │                                                                                 │")
        logger.info("  │  STANDARD DEVIATION (Population - divides by N):                                │")
        logger.info(f"  │    =STDEV.P(B3:B{n_obs + 2})                                                     │")
        logger.info(f"  │    Result for {sample_ticker}: {stds[sample_ticker]:.7f}%                                      │")
        logger.info("  │    Drag across all ticker columns                                               │")
        logger.info("  │                                                                                 │")
        logger.info("  │  NOTE: Use STDEV.P (population), NOT STDEV.S (sample)                           │")
        logger.info("  │        STDEV.P divides by N, STDEV.S divides by (N-1)                           │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")
        logger.info("  ╔═════════════════════════════════════════════════════════════════════════════════╗")
        logger.info("  ║           EXCEL METHOD 2: Data Analysis ToolPak - Descriptive Statistics        ║")
        logger.info("  ╚═════════════════════════════════════════════════════════════════════════════════╝")
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  HOW TO USE:                                                                    │")
        logger.info("  │    1. Data tab → Data Analysis → Descriptive Statistics                         │")
        logger.info(f"  │    2. Input Range: B3:AF{n_obs + 2} (all returns data)                          │")
        logger.info("  │    3. [x] Labels in first row (if you include row 2 with tickers)                 │")
        logger.info("  │    4. [x] Summary statistics                                                      │")
        logger.info("  │    5. Output Range: Select where you want results                               │")
        logger.info("  │    6. Click OK                                                                  │")
        logger.info("  │                                                                                 │")
        logger.info("  │  OUTPUT INCLUDES:                                                               │")
        logger.info("  │    • Mean (same as AVERAGE formula)                                             │")
        logger.info("  │    • Standard Deviation (WARNING: uses N-1, sample std)                               │")
        logger.info("  │    • Variance, Min, Max, Sum, Count, etc.                                       │")
        logger.info("  │                                                                                 │")
        logger.info("  │  WARNING:  IMPORTANT: Data Analysis gives SAMPLE std dev (divides by N-1)             │")
        logger.info(f"  │     To convert to POPULATION std dev: multiply by √((N-1)/N) = √({n_obs-1}/{n_obs})      │")
        logger.info(f"  │     Or: =StdDev_Sample × SQRT({n_obs-1}/{n_obs})                                        │")
        logger.info("  │                                                                                 │")
        logger.info("  │  CELL VALUES IN INPUT RANGE:                                                    │")
        # Get more tickers for display
        t1, t2, t3, t4 = tickers[0], tickers[1], tickers[2], tickers[3]
        r1_0 = returns_only[t1].iloc[0]
        r2_0 = returns_only[t2].iloc[0]
        r3_0 = returns_only[tickers[2]].iloc[0]
        r4_0 = returns_only[tickers[3]].iloc[0]
        logger.info(f"  │    B3={r1_0:>8.4f} ({t1})  C3={r2_0:>8.4f} ({t2})  D3={r3_0:>8.4f} ({tickers[2]})        │")
        logger.info(f"  │    B4={returns_only[t1].iloc[1]:>8.4f}       C4={returns_only[t2].iloc[1]:>8.4f}       D4={returns_only[tickers[2]].iloc[1]:>8.4f}             │")
        logger.info(f"  │    ...                                                                         │")
        logger.info(f"  │    B{n_obs+2}={returns_only[t1].iloc[-1]:>8.4f}      C{n_obs+2}={returns_only[t2].iloc[-1]:>8.4f}      D{n_obs+2}={returns_only[tickers[2]].iloc[-1]:>8.4f}            │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")

        # Verify a sample calculation with actual values
        manual_mean = returns_only[sample_ticker].sum() / n_obs
        manual_std = np.sqrt(((returns_only[sample_ticker] - manual_mean) ** 2).sum() / n_obs)
        logger.info(f"  [OK] VERIFICATION for {sample_ticker}:")
        logger.info(f"    Sum of all {n_obs} returns: {returns_only[sample_ticker].sum():.7f}")
        logger.info(f"    Manual Mean:  {returns_only[sample_ticker].sum():.7f} / {n_obs} = {manual_mean:.7f}%")
        logger.info(f"    Computed Mean:                 {means[sample_ticker]:.7f}%")
        logger.info(f"    Match: {np.isclose(manual_mean, means[sample_ticker])}")
        logger.info(f"    Manual Std:   √(Σ(r-μ)² / N) = {manual_std:.7f}%")
        logger.info(f"    Computed Std:                  {stds[sample_ticker]:.7f}%")
        logger.info(f"    Match: {np.isclose(manual_std, stds[sample_ticker])}")
        logger.info("")

        log_dataframe(logger, stats_df.T, "STATISTICS TABLE")

        # Compute covariance matrix (population) in DECIMAL form
        # Returns are in percentage form (e.g., 3.47 = 3.47%), so divide by 100 first
        # Covariance of decimals = Covariance of percentages / 10000 (since cov scales with square of factor)
        returns_decimal = returns_only / 100  # Convert to decimal form (0.0347 = 3.47%)
        cov_matrix = returns_decimal.cov() * (n_obs - 1) / n_obs

        # =====================================================================
        # LOG COVARIANCE MATRIX WITH EXCEL INSTRUCTIONS AND VERIFICATION
        # =====================================================================
        logger.info("")
        logger.info("=" * 100)
        logger.info(f"  COVARIANCE MATRIX ({len(tickers)}x{len(tickers)})")
        logger.info("=" * 100)
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  MATHEMATICAL FORMULA:                                                          │")
        logger.info("  │                                                                                 │")
        logger.info("  │    Cov(X,Y) = (1/N) × Σᵢ [(Xᵢ - μₓ)(Yᵢ - μᵧ)]    (Population covariance)       │")
        logger.info("  │                                                                                 │")
        logger.info("  │  MATRIX FORM:                                                                   │")
        logger.info("  │    Σ = (1/N) × (R - μ)ᵀ × (R - μ)                                               │")
        logger.info("  │                                                                                 │")
        logger.info("  │  VARIABLE DEFINITIONS:                                                          │")
        logger.info(f"  │    N      = Number of observations = {n_obs}                                     │")
        logger.info("  │    Xᵢ, Yᵢ = Return of asset X or Y at time i                                    │")
        logger.info("  │    μₓ, μᵧ = Mean return of asset X or Y (mu_x, mu_y)                            │")
        logger.info("  │    Σ      = Covariance matrix (capital Sigma)                                   │")
        logger.info("  │    R      = Matrix of returns (N rows × k assets)                               │")
        logger.info("  │    μ      = Row vector of mean returns                                          │")
        logger.info("  │    (R-μ)  = Demeaned returns matrix (subtract mean from each column)            │")
        logger.info("  │    ᵀ      = Transpose operator (rows become columns)                            │")
        logger.info("  │    Cov(X,X) = Var(X) = σₓ²  (diagonal elements = variance)                      │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")
        logger.info("  ╔═════════════════════════════════════════════════════════════════════════════════╗")
        logger.info("  ║             EXCEL METHOD 1: Data Analysis ToolPak - Covariance                  ║")
        logger.info("  ╚═════════════════════════════════════════════════════════════════════════════════╝")
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  HOW TO USE:                                                                    │")
        logger.info("  │    1. Data tab → Data Analysis → Covariance                                     │")
        logger.info(f"  │    2. Input Range: B3:AF{n_obs + 2} (all returns data)                          │")
        logger.info("  │    3. [x] Labels in first row (if you include row 2 with tickers)                 │")
        logger.info("  │    4. Output Range: Select destination cell for top-left of matrix              │")
        logger.info("  │    5. Click OK                                                                  │")
        logger.info("  │                                                                                 │")
        logger.info("  │  OUTPUT:                                                                        │")
        logger.info(f"  │    • {len(tickers)}×{len(tickers)} covariance matrix                                             │")
        logger.info("  │    • Diagonal elements = Variance of each asset                                 │")
        logger.info("  │    • Off-diagonal = Covariance between pairs of assets                          │")
        logger.info("  │    • Lower triangle filled (upper triangle blank - it's symmetric)              │")
        logger.info("  │                                                                                 │")
        logger.info("  │  WARNING:  IMPORTANT: Data Analysis uses SAMPLE covariance (divides by N-1)           │")
        logger.info(f"  │     To convert to POPULATION covariance: multiply each cell by (N-1)/N         │")
        logger.info(f"  │     = ({n_obs-1}/{n_obs}) = {(n_obs-1)/n_obs:.7f}                                              │")
        logger.info("  │                                                                                 │")
        logger.info("  │  CONVERSION FORMULA:                                                            │")
        logger.info(f"  │     Population_Cov = Sample_Cov × {(n_obs-1)/n_obs:.7f}                                  │")
        logger.info("  │     In Excel: =DataAnalysisResult × (COUNT(returns)-1)/COUNT(returns)           │")
        logger.info("  │                                                                                 │")
        logger.info("  │  CELL VALUES IN INPUT RANGE (returns %):                                        │")
        # Preview values for covariance Data Analysis section
        cov_t1, cov_t2, cov_t3 = tickers[0], tickers[1], tickers[2]
        logger.info(f"  │    B3={returns_only[cov_t1].iloc[0]:>8.4f} ({cov_t1})  C3={returns_only[cov_t2].iloc[0]:>8.4f} ({cov_t2})  D3={returns_only[cov_t3].iloc[0]:>8.4f} ({cov_t3})        │")
        logger.info(f"  │    B4={returns_only[cov_t1].iloc[1]:>8.4f}       C4={returns_only[cov_t2].iloc[1]:>8.4f}       D4={returns_only[cov_t3].iloc[1]:>8.4f}             │")
        logger.info(f"  │    ...                                                                         │")
        logger.info(f"  │    B{n_obs+2}={returns_only[cov_t1].iloc[-1]:>8.4f}      C{n_obs+2}={returns_only[cov_t2].iloc[-1]:>8.4f}      D{n_obs+2}={returns_only[cov_t3].iloc[-1]:>8.4f}            │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")
        # Get sample values for examples
        t1, t2, t3 = tickers[0], tickers[1], tickers[2]  # e.g., SPY, AAPL, AMGN
        r1 = returns_only[t1].values
        r2 = returns_only[t2].values
        m1, m2, m3 = means[t1], means[t2], means[tickers[2]]
        first_r1, first_r2 = r1[0], r2[0]  # First returns (B3, C3)
        first_demeaned_r1 = first_r1 - m1  # B103 value
        first_demeaned_r2 = first_r2 - m2  # C103 value

        logger.info("")
        logger.info("  ╔═════════════════════════════════════════════════════════════════════════════════════════════════════╗")
        logger.info("  ║                    EXCEL METHOD 2: MMULT (Matrix Multiplication)                                    ║")
        logger.info("  ║                        Recommended - gives population covariance directly                           ║")
        logger.info("  ╚═════════════════════════════════════════════════════════════════════════════════════════════════════╝")
        logger.info("")

        # Get demeaned values for ALL tickers (first 5 for display)
        d1 = [r1[i] - m1 for i in range(len(r1))]  # Demeaned SPY
        d2 = [r2[i] - m2 for i in range(len(r2))]  # Demeaned AAPL
        d3 = [returns_only[t3].values[i] - means[t3] for i in range(len(r1))]  # Demeaned AMGN
        t4, t5 = tickers[3], tickers[4]  # AXP, BA
        d4 = [returns_only[t4].values[i] - means[t4] for i in range(len(r1))]
        d5 = [returns_only[t5].values[i] - means[t5] for i in range(len(r1))]

        logger.info("  ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
        logger.info("  ┃  STEP 1: Create Demeaned Returns Matrix (R - μ)                                                     ┃")
        logger.info("  ┃          Subtract the mean from each return to center the data around zero                          ┃")
        logger.info("  ┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛")
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  EXCEL FORMULAS FOR STEP 1:                                                                         │")
        logger.info("  │                                                                                                     │")
        logger.info("  │  Location: Create demeaned returns starting at Row 103 (below the original data)                    │")
        logger.info("  │                                                                                                     │")
        logger.info(f"  │  Cell B103: =B3-B$64       (where B$64 = {m1:.4f} is the mean of {t1})                             │")
        logger.info(f"  │  Cell C103: =C3-C$64       (where C$64 = {m2:.4f} is the mean of {t2})                            │")
        logger.info(f"  │  Cell D103: =D3-D$64       (where D$64 = {means[t3]:.4f} is the mean of {t3})                            │")
        logger.info("  │                                                                                                     │")
        logger.info("  │  Then drag the formula:                                                                             │")
        logger.info("  │    • RIGHT across all columns (B103 → AF103)                                                        │")
        logger.info(f"  │    • DOWN for all {n_obs} rows (Row 103 → Row {102 + n_obs})                                                     │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")
        logger.info("  ORIGINAL RETURNS (Rows 3-74):                              DEMEANED RETURNS (Rows 103-174):")
        logger.info(f"  ┌──────────────────────────────────────────────────────┐   ┌──────────────────────────────────────────────────────┐")
        logger.info(f"  │  Row      {t1:>6}   {t2:>6}   {t3:>6}   {t4:>6}   {t5:>6}  │   │  Row      {t1:>6}   {t2:>6}   {t3:>6}   {t4:>6}   {t5:>6}  │")
        logger.info(f"  ├──────────────────────────────────────────────────────┤   ├──────────────────────────────────────────────────────┤")
        for i in range(min(8, n_obs)):
            r_row = i + 3
            d_row = i + 103
            rv1, rv2, rv3, rv4, rv5 = r1[i], r2[i], returns_only[t3].values[i], returns_only[t4].values[i], returns_only[t5].values[i]
            logger.info(f"  │  {r_row:>3}   {rv1:>8.2f} {rv2:>8.2f} {rv3:>8.2f} {rv4:>8.2f} {rv5:>8.2f}  │   │  {d_row:>3}   {d1[i]:>8.2f} {d2[i]:>8.2f} {d3[i]:>8.2f} {d4[i]:>8.2f} {d5[i]:>8.2f}  │")
        logger.info(f"  │  ...       ...      ...      ...      ...      ...  │   │  ...       ...      ...      ...      ...      ...  │")
        logger.info(f"  │  {n_obs+2:>3}   {r1[-1]:>8.2f} {r2[-1]:>8.2f} {returns_only[t3].values[-1]:>8.2f} {returns_only[t4].values[-1]:>8.2f} {returns_only[t5].values[-1]:>8.2f}  │   │  {n_obs+102:>3}   {d1[-1]:>8.2f} {d2[-1]:>8.2f} {d3[-1]:>8.2f} {d4[-1]:>8.2f} {d5[-1]:>8.2f}  │")
        logger.info(f"  └──────────────────────────────────────────────────────┘   └──────────────────────────────────────────────────────┘")
        logger.info(f"            SIZE: {n_obs} rows × {len(tickers)} columns                            SIZE: {n_obs} rows × {len(tickers)} columns")
        logger.info("")
        logger.info("  CALCULATION EXAMPLE:")
        logger.info(f"    B103 = B3 - B$64 = {r1[0]:.4f} - {m1:.4f} = {d1[0]:.4f}")
        logger.info(f"    C103 = C3 - C$64 = {r2[0]:.4f} - {m2:.4f} = {d2[0]:.4f}")
        logger.info(f"    D103 = D3 - D$64 = {returns_only[t3].values[0]:.4f} - {means[t3]:.4f} = {d3[0]:.4f}")
        logger.info("")

        # Create full demeaned matrix DataFrame and log it
        demeaned_matrix = returns_only.copy()
        for ticker in tickers:
            demeaned_matrix[ticker] = returns_only[ticker] - means[ticker]
        log_dataframe(logger, demeaned_matrix, "FULL DEMEANED MATRIX (R - μ) - ALL VALUES",
                      decimals=4, start_row=103, start_col=1)

        logger.info("  ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
        logger.info("  ┃  STEP 2: Transpose the Demeaned Matrix                                                              ┃")
        logger.info("  ┃          Convert rows to columns: (R-μ)ᵀ                                                            ┃")
        logger.info("  ┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛")
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  EXCEL FORMULA FOR TRANSPOSE:                                                                       │")
        logger.info("  │                                                                                                     │")
        logger.info(f"  │  =TRANSPOSE(B103:AF{102 + n_obs})                                                                           │")
        logger.info("  │                                                                                                     │")
        logger.info("  │  This converts:                                                                                     │")
        logger.info(f"  │    • Original: {n_obs} rows × {len(tickers)} columns  →  Transposed: {len(tickers)} rows × {n_obs} columns                         │")
        logger.info("  │    • Each ticker's returns become a ROW instead of a column                                         │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")
        logger.info("  DEMEANED MATRIX (R-μ):                              TRANSPOSED MATRIX (R-μ)ᵀ:")
        logger.info(f"  {n_obs} rows × {len(tickers)} cols                                       {len(tickers)} rows × {n_obs} cols")
        logger.info("")
        logger.info(f"  ┌────────────────────────────────────────┐         ┌─────────────────────────────────────────────────────────────────────────┐")
        logger.info(f"  │        {t1:>6}  {t2:>6}  {t3:>6}  ...  │         │  Ticker   Col1     Col2     Col3     Col4    ...   Col{n_obs}  │")
        logger.info(f"  ├────────────────────────────────────────┤         ├─────────────────────────────────────────────────────────────────────────┤")
        logger.info(f"  │ Row1  {d1[0]:>7.2f} {d2[0]:>7.2f} {d3[0]:>7.2f}  ...  │         │  {t1:>6}  {d1[0]:>7.2f}  {d1[1]:>7.2f}  {d1[2]:>7.2f}  {d1[3]:>7.2f}   ...  {d1[-1]:>7.2f}  │")
        logger.info(f"  │ Row2  {d1[1]:>7.2f} {d2[1]:>7.2f} {d3[1]:>7.2f}  ...  │   ─►    │  {t2:>6}  {d2[0]:>7.2f}  {d2[1]:>7.2f}  {d2[2]:>7.2f}  {d2[3]:>7.2f}   ...  {d2[-1]:>7.2f}  │")
        logger.info(f"  │ Row3  {d1[2]:>7.2f} {d2[2]:>7.2f} {d3[2]:>7.2f}  ...  │         │  {t3:>6}  {d3[0]:>7.2f}  {d3[1]:>7.2f}  {d3[2]:>7.2f}  {d3[3]:>7.2f}   ...  {d3[-1]:>7.2f}  │")
        logger.info(f"  │ Row4  {d1[3]:>7.2f} {d2[3]:>7.2f} {d3[3]:>7.2f}  ...  │         │  {t4:>6}  {d4[0]:>7.2f}  {d4[1]:>7.2f}  {d4[2]:>7.2f}  {d4[3]:>7.2f}   ...  {d4[-1]:>7.2f}  │")
        logger.info(f"  │  ...     ...     ...     ...   ...  │         │  {t5:>6}  {d5[0]:>7.2f}  {d5[1]:>7.2f}  {d5[2]:>7.2f}  {d5[3]:>7.2f}   ...  {d5[-1]:>7.2f}  │")
        logger.info(f"  │ Row{n_obs}  {d1[-1]:>7.2f} {d2[-1]:>7.2f} {d3[-1]:>7.2f}  ...  │         │    ...     ...      ...      ...      ...    ...      ...   │")
        logger.info(f"  └────────────────────────────────────────┘         └─────────────────────────────────────────────────────────────────────────┘")
        logger.info("")

        # Log full transposed demeaned matrix
        transposed_demeaned = demeaned_matrix.T
        log_dataframe(logger, transposed_demeaned, "FULL TRANSPOSED DEMEANED MATRIX (R - μ)ᵀ - ALL VALUES",
                      decimals=4, start_row=1, start_col=1, show_excel_refs=False)

        logger.info("  ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
        logger.info("  ┃  STEP 3: Matrix Multiplication (R-μ)ᵀ × (R-μ)                                                       ┃")
        logger.info("  ┃          Multiply transposed matrix by original demeaned matrix                                     ┃")
        logger.info("  ┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛")
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  EXCEL FORMULA FOR MATRIX MULTIPLICATION (DECIMAL COVARIANCE):                                      │")
        logger.info("  │                                                                                                     │")
        logger.info(f"  │  =MMULT(TRANSPOSE(B103:AF{102+n_obs}), B103:AF{102+n_obs}) / {n_obs} / 10000                                        │")
        logger.info("  │                                                                                                     │")
        logger.info("  │  WARNING:  IMPORTANT: Returns are in PERCENTAGE form (e.g., 3.47 = 3.47%)                                 │")
        logger.info("  │     Covariance should be in DECIMAL form for portfolio optimization!                                │")
        logger.info("  │     Divide by 10000 (= 100²) to convert from %² to decimal²                                         │")
        logger.info("  │                                                                                                     │")
        logger.info("  │  Breaking it down:                                                                                  │")
        logger.info(f"  │    • TRANSPOSE(B103:AF{102+n_obs}) = the transposed demeaned matrix ({len(tickers)}×{n_obs})                        │")
        logger.info(f"  │    • B103:AF{102+n_obs} = the original demeaned matrix ({n_obs}×{len(tickers)})                                      │")
        logger.info(f"  │    • MMULT multiplies them: ({len(tickers)}×{n_obs}) × ({n_obs}×{len(tickers)}) = ({len(tickers)}×{len(tickers)})                                      │")
        logger.info(f"  │    • Divide by {n_obs} for population covariance (in %²)                                                   │")
        logger.info("  │    • Divide by 10000 to convert to decimal covariance                                               │")
        logger.info("  │                                                                                                     │")
        logger.info("  │  ALTERNATIVE (if you want returns in decimal form first):                                           │")
        logger.info(f"  │  =MMULT(TRANSPOSE(B103:AF{102+n_obs}/100), B103:AF{102+n_obs}/100) / {n_obs}                                        │")
        logger.info("  │                                                                                                     │")
        logger.info("  │  Select a {0}×{0} range, type the formula, press Ctrl+Shift+Enter (or Enter in Excel 365)          │".format(len(tickers)))
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")
        logger.info("  MATRIX MULTIPLICATION VISUALIZATION:")
        logger.info("")
        logger.info(f"       (R-μ)ᵀ                              ×              (R-μ)                      =        Σ × N")
        logger.info(f"    ({len(tickers)} × {n_obs})                                        ({n_obs} × {len(tickers)})                          ({len(tickers)} × {len(tickers)})")
        logger.info("")
        logger.info(f"  ┌─────────────────────────────────┐       ┌─────────────────────────┐       ┌─────────────────────────┐")
        logger.info(f"  │ {t1}: {d1[0]:>5.1f} {d1[1]:>5.1f} {d1[2]:>5.1f} ... {d1[-1]:>5.1f} │       │      {t1:>5} {t2:>5} {t3:>5} ... │       │      {t1:>5} {t2:>5} {t3:>5} ... │")
        logger.info(f"  │ {t2}: {d2[0]:>5.1f} {d2[1]:>5.1f} {d2[2]:>5.1f} ... {d2[-1]:>5.1f} │   ×   │ R1:  {d1[0]:>5.1f} {d2[0]:>5.1f} {d3[0]:>5.1f} ... │   =   │ {t1}: {sum(d1[i]*d1[i] for i in range(len(d1))):>5.0f} {sum(d1[i]*d2[i] for i in range(len(d1))):>5.0f} {sum(d1[i]*d3[i] for i in range(len(d1))):>5.0f} ... │")
        logger.info(f"  │ {t3}: {d3[0]:>5.1f} {d3[1]:>5.1f} {d3[2]:>5.1f} ... {d3[-1]:>5.1f} │       │ R2:  {d1[1]:>5.1f} {d2[1]:>5.1f} {d3[1]:>5.1f} ... │       │ {t2}: {sum(d1[i]*d2[i] for i in range(len(d1))):>5.0f} {sum(d2[i]*d2[i] for i in range(len(d2))):>5.0f} {sum(d2[i]*d3[i] for i in range(len(d2))):>5.0f} ... │")
        logger.info(f"  │ ...  ...  ...  ...  ...   ...   │       │ R3:  {d1[2]:>5.1f} {d2[2]:>5.1f} {d3[2]:>5.1f} ... │       │ {t3}: {sum(d1[i]*d3[i] for i in range(len(d1))):>5.0f} {sum(d2[i]*d3[i] for i in range(len(d2))):>5.0f} {sum(d3[i]*d3[i] for i in range(len(d3))):>5.0f} ... │")
        logger.info(f"  │                                 │       │ ...   ...   ...   ...  ... │       │ ...   ...   ...   ...  ... │")
        logger.info(f"  └─────────────────────────────────┘       └─────────────────────────┘       └─────────────────────────┘")
        logger.info("")

        # Compute actual dot products
        sum_d1_d1 = sum(d1[i] * d1[i] for i in range(len(d1)))
        sum_d1_d2 = sum(d1[i] * d2[i] for i in range(len(d1)))
        sum_d1_d3 = sum(d1[i] * d3[i] for i in range(len(d1)))
        sum_d2_d2 = sum(d2[i] * d2[i] for i in range(len(d2)))
        sum_d2_d3 = sum(d2[i] * d3[i] for i in range(len(d2)))
        sum_d3_d3 = sum(d3[i] * d3[i] for i in range(len(d3)))

        logger.info("  ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
        logger.info("  ┃  STEP 4: Computing Each Element (Dot Product / N)                                                   ┃")
        logger.info("  ┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛")
        logger.info("")
        logger.info(f"  Each element Cov(X,Y) = (1/N) × Σᵢ[(Xᵢ - μₓ)(Yᵢ - μᵧ)]  where N = {n_obs}")
        logger.info("")
        logger.info(f"  ═══════════════════════════════════════════════════════════════════════════════════════════════════")
        logger.info(f"  Cov({t1},{t1}) = Var({t1}) = (1/{n_obs}) × [{t1} row] • [{t1} column]")
        logger.info(f"  ═══════════════════════════════════════════════════════════════════════════════════════════════════")
        logger.info(f"")
        logger.info(f"    = (1/{n_obs}) × [ ({d1[0]:.2f})×({d1[0]:.2f}) + ({d1[1]:.2f})×({d1[1]:.2f}) + ({d1[2]:.2f})×({d1[2]:.2f}) + ({d1[3]:.2f})×({d1[3]:.2f}) + ({d1[4]:.2f})×({d1[4]:.2f}) + ({d1[5]:.2f})×({d1[5]:.2f}) + ... ]")
        logger.info(f"")
        logger.info(f"    = (1/{n_obs}) × [ {d1[0]*d1[0]:>8.2f} + {d1[1]*d1[1]:>8.2f} + {d1[2]*d1[2]:>8.2f} + {d1[3]*d1[3]:>8.2f} + {d1[4]*d1[4]:>8.2f} + {d1[5]*d1[5]:>8.2f} + ... + {d1[-1]*d1[-1]:>8.2f} ]")
        logger.info(f"")
        logger.info(f"    = (1/{n_obs}) × {sum_d1_d1:>12.4f}")
        logger.info(f"")
        logger.info(f"    = {sum_d1_d1/n_obs:>12.7f}  ← This equals σ²({t1}) = ({stds[t1]:.7f})² = {stds[t1]**2:.7f} [OK]")
        logger.info("")
        logger.info(f"  ═══════════════════════════════════════════════════════════════════════════════════════════════════")
        logger.info(f"  Cov({t1},{t2}) = (1/{n_obs}) × [{t1} row] • [{t2} column]")
        logger.info(f"  ═══════════════════════════════════════════════════════════════════════════════════════════════════")
        logger.info(f"")
        logger.info(f"    = (1/{n_obs}) × [ ({d1[0]:.2f})×({d2[0]:.2f}) + ({d1[1]:.2f})×({d2[1]:.2f}) + ({d1[2]:.2f})×({d2[2]:.2f}) + ({d1[3]:.2f})×({d2[3]:.2f}) + ({d1[4]:.2f})×({d2[4]:.2f}) + ({d1[5]:.2f})×({d2[5]:.2f}) + ... ]")
        logger.info(f"")
        logger.info(f"    = (1/{n_obs}) × [ {d1[0]*d2[0]:>8.2f} + {d1[1]*d2[1]:>8.2f} + {d1[2]*d2[2]:>8.2f} + {d1[3]*d2[3]:>8.2f} + {d1[4]*d2[4]:>8.2f} + {d1[5]*d2[5]:>8.2f} + ... + {d1[-1]*d2[-1]:>8.2f} ]")
        logger.info(f"")
        logger.info(f"    = (1/{n_obs}) × {sum_d1_d2:>12.4f}")
        logger.info(f"")
        logger.info(f"    = {sum_d1_d2/n_obs:>12.7f}")
        logger.info("")
        logger.info(f"  ═══════════════════════════════════════════════════════════════════════════════════════════════════")
        logger.info(f"  Cov({t2},{t2}) = Var({t2}) = (1/{n_obs}) × [{t2} row] • [{t2} column]")
        logger.info(f"  ═══════════════════════════════════════════════════════════════════════════════════════════════════")
        logger.info(f"")
        logger.info(f"    = (1/{n_obs}) × {sum_d2_d2:>12.4f}")
        logger.info(f"")
        logger.info(f"    = {sum_d2_d2/n_obs:>12.7f}  ← This equals σ²({t2}) = ({stds[t2]:.7f})² = {stds[t2]**2:.7f} [OK]")
        logger.info("")

        logger.info("  ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
        logger.info("  ┃  STEP 5: Final Covariance Matrix (Σ)                                                                ┃")
        logger.info("  ┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛")
        logger.info("")

        # Compute and log the raw MMULT result (sum of products, before dividing by N)
        # This is (R-μ)ᵀ × (R-μ) in percentage form
        mmult_raw = demeaned_matrix.T.dot(demeaned_matrix)
        log_dataframe(logger, mmult_raw, "RAW MMULT RESULT: (R-μ)ᵀ × (R-μ) - Sum of Products (in %²)",
                      decimals=4, start_row=1, start_col=1, show_excel_refs=False)

        # Compute and log the covariance in percentage form (divide by N)
        cov_pct_form = mmult_raw / n_obs
        log_dataframe(logger, cov_pct_form, "COVARIANCE MATRIX (% form): MMULT / N - Population Covariance (in %²)",
                      decimals=6, start_row=1, start_col=1, show_excel_refs=False)

        cov_spy_spy = cov_matrix.loc[t1, t1]
        cov_spy_aapl = cov_matrix.loc[t1, t2]
        cov_spy_amgn = cov_matrix.loc[t1, t3]
        cov_spy_axp = cov_matrix.loc[t1, t4]
        cov_spy_ba = cov_matrix.loc[t1, t5]
        cov_aapl_aapl = cov_matrix.loc[t2, t2]
        cov_aapl_amgn = cov_matrix.loc[t2, t3]
        cov_aapl_axp = cov_matrix.loc[t2, t4]
        cov_aapl_ba = cov_matrix.loc[t2, t5]
        cov_amgn_amgn = cov_matrix.loc[t3, t3]
        cov_amgn_axp = cov_matrix.loc[t3, t4]
        cov_amgn_ba = cov_matrix.loc[t3, t5]
        cov_axp_axp = cov_matrix.loc[t4, t4]
        cov_axp_ba = cov_matrix.loc[t4, t5]
        cov_ba_ba = cov_matrix.loc[t5, t5]

        logger.info("  The result is a symmetric matrix where:")
        logger.info("    • Diagonal elements = Variance of each asset (σ²)")
        logger.info("    • Off-diagonal elements = Covariance between pairs of assets")
        logger.info("    • Matrix is symmetric: Cov(X,Y) = Cov(Y,X)")
        logger.info("")
        logger.info(f"  ┌────────────────────────────────────────────────────────────────────────────────────────────────────┐")
        logger.info(f"  │                    {t1:>8}      {t2:>8}      {t3:>8}      {t4:>8}      {t5:>8}     ...    │")
        logger.info(f"  ├────────────────────────────────────────────────────────────────────────────────────────────────────┤")
        logger.info(f"  │  {t1:>6}      {cov_spy_spy:>10.4f}  {cov_spy_aapl:>10.4f}  {cov_spy_amgn:>10.4f}  {cov_spy_axp:>10.4f}  {cov_spy_ba:>10.4f}     ...    │")
        logger.info(f"  │  {t2:>6}      {cov_spy_aapl:>10.4f}  {cov_aapl_aapl:>10.4f}  {cov_aapl_amgn:>10.4f}  {cov_aapl_axp:>10.4f}  {cov_aapl_ba:>10.4f}     ...    │")
        logger.info(f"  │  {t3:>6}      {cov_spy_amgn:>10.4f}  {cov_aapl_amgn:>10.4f}  {cov_amgn_amgn:>10.4f}  {cov_amgn_axp:>10.4f}  {cov_amgn_ba:>10.4f}     ...    │")
        logger.info(f"  │  {t4:>6}      {cov_spy_axp:>10.4f}  {cov_aapl_axp:>10.4f}  {cov_amgn_axp:>10.4f}  {cov_axp_axp:>10.4f}  {cov_axp_ba:>10.4f}     ...    │")
        logger.info(f"  │  {t5:>6}      {cov_spy_ba:>10.4f}  {cov_aapl_ba:>10.4f}  {cov_amgn_ba:>10.4f}  {cov_axp_ba:>10.4f}  {cov_ba_ba:>10.4f}     ...    │")
        logger.info(f"  │    ...           ...          ...          ...          ...          ...      ...    │")
        logger.info(f"  └────────────────────────────────────────────────────────────────────────────────────────────────────┘")
        logger.info(f"                                        SIZE: {len(tickers)} × {len(tickers)}")
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  COMPLETE EXCEL FORMULA FOR DECIMAL COVARIANCE (copy this exactly):                                 │")
        logger.info("  │                                                                                                     │")
        logger.info(f"  │  =MMULT(TRANSPOSE(B103:AF{102+n_obs}), B103:AF{102+n_obs}) / {n_obs} / 10000                                        │")
        logger.info("  │                                                                                                     │")
        logger.info("  │  WARNING:  The /10000 converts from percentage² to decimal² (since 100² = 10000)                          │")
        logger.info("  │                                                                                                     │")
        logger.info(f"  │  1. Select a {len(tickers)}×{len(tickers)} range (e.g., B180:AF210) for the output                                         │")
        logger.info("  │  2. Type the formula above                                                                          │")
        logger.info("  │  3. Press Ctrl+Shift+Enter (array formula) or just Enter in Excel 365                               │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")
        logger.info("  ╔═════════════════════════════════════════════════════════════════════════════════╗")
        logger.info("  ║                           VERIFICATION                                          ║")
        logger.info("  ╚═════════════════════════════════════════════════════════════════════════════════╝")
        logger.info("")

        # Verify covariance calculation manually (convert to decimal form)
        # Returns are in % form, so demeaned products are in %² form
        # Divide by 10000 to convert to decimal form
        manual_cov_pct = np.sum((r1 - m1) * (r2 - m2)) / n_obs  # In %² form
        manual_cov = manual_cov_pct / 10000  # Convert to decimal form
        computed_cov = cov_matrix.loc[t1, t2]

        # Calculate decimal demeaned values for display
        first_r1_dec = first_r1 / 100  # Convert to decimal (0.0347 = 3.47%)
        first_r2_dec = first_r2 / 100
        m1_dec = m1 / 100
        m2_dec = m2 / 100
        first_demeaned_r1_dec = first_r1_dec - m1_dec
        first_demeaned_r2_dec = first_r2_dec - m2_dec

        logger.info(f"  [OK] VERIFICATION - Covariance({t1}, {t2}):")
        logger.info(f"    Formula: Σ[(r_{t1} - μ_{t1})(r_{t2} - μ_{t2})] / N  (using decimal returns)")
        logger.info(f"    ")
        logger.info(f"    ┌───────────────────────────────────────────────────────────────────────┐")
        logger.info(f"    │  Step-by-step calculation for first observation (in DECIMAL form):   │")
        logger.info(f"    ├───────────────────────────────────────────────────────────────────────┤")
        logger.info(f"    │  {t1} return (B3):    {first_r1_dec:>12.7f} (= {first_r1:.4f}% / 100)          │")
        logger.info(f"    │  {t1} mean (B64):     {m1_dec:>12.7f} (= {m1:.4f}% / 100)           │")
        logger.info(f"    │  {t1} demeaned:       {first_r1_dec:.7f} - {m1_dec:.7f} = {first_demeaned_r1_dec:>12.7f}  │")
        logger.info(f"    │                                                                       │")
        logger.info(f"    │  {t2} return (C3):   {first_r2_dec:>12.7f} (= {first_r2:.4f}% / 100)          │")
        logger.info(f"    │  {t2} mean (C64):    {m2_dec:>12.7f} (= {m2:.4f}% / 100)           │")
        logger.info(f"    │  {t2} demeaned:      {first_r2_dec:.7f} - {m2_dec:.7f} = {first_demeaned_r2_dec:>12.7f}  │")
        logger.info(f"    │                                                                       │")
        logger.info(f"    │  Product: {first_demeaned_r1_dec:.7f} × {first_demeaned_r2_dec:.7f} = {first_demeaned_r1_dec * first_demeaned_r2_dec:>12.9f}  │")
        logger.info(f"    └───────────────────────────────────────────────────────────────────────┘")
        logger.info(f"    ")
        logger.info(f"    Sum of all {n_obs} products (decimal): {np.sum((r1/100 - m1/100) * (r2/100 - m2/100)):.9f}")
        logger.info(f"    Divide by N: {np.sum((r1/100 - m1/100) * (r2/100 - m2/100)):.9f} / {n_obs} = {manual_cov:.9f}")
        logger.info(f"    Matrix value:                     {computed_cov:.9f}")
        logger.info(f"    Match: {np.isclose(manual_cov, computed_cov)}")
        logger.info("")

        # Verify variance (diagonal) = std^2 (in decimal form)
        # Std is in % form (5.19 = 5.19%), convert to decimal (0.0519) then square
        std_decimal = stds[t1] / 100  # Convert % to decimal
        manual_var = std_decimal ** 2  # Variance in decimal form
        matrix_var = cov_matrix.loc[t1, t1]
        logger.info(f"  [OK] VERIFICATION - Variance({t1}) = Std² (in decimal form):")
        logger.info(f"    Std({t1}) = {stds[t1]:.7f}% = {std_decimal:.9f} (decimal)")
        logger.info(f"    Std({t1})² = {std_decimal:.9f}² = {manual_var:.9f}")
        logger.info(f"    Cov({t1},{t1}) from matrix:        {matrix_var:.9f}")
        logger.info(f"    Match: {np.isclose(manual_var, matrix_var)}")
        logger.info("")

        # =====================================================================
        # EXCEL TABLE EXPLANATION
        # =====================================================================
        logger.info("")
        logger.info("  ╔═════════════════════════════════════════════════════════════════════════════════════════════════════╗")
        logger.info("  ║                    ALTERNATIVE: Using Excel Tables for LN Returns                                   ║")
        logger.info("  ╚═════════════════════════════════════════════════════════════════════════════════════════════════════╝")
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  WHAT IS AN EXCEL TABLE?                                                                            │")
        logger.info("  │                                                                                                     │")
        logger.info("  │  An Excel Table is a structured data range with:                                                    │")
        logger.info("  │    • Named columns that can be referenced by name instead of cell addresses                        │")
        logger.info("  │    • Automatic expansion when you add new data                                                      │")
        logger.info("  │    • Structured references that are more readable than A1-style references                         │")
        logger.info("  │    • Built-in filtering and sorting capabilities                                                    │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")
        logger.info("  ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
        logger.info("  ┃  HOW TO CREATE A TABLE FROM LN RETURNS DATA                                                         ┃")
        logger.info("  ┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛")
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  STEP-BY-STEP:                                                                                      │")
        logger.info("  │                                                                                                     │")
        logger.info(f"  │    1. Select your LN Returns data INCLUDING headers: A2:AF{2+n_obs}                                      │")
        logger.info("  │       (Row 2 has ticker names: Date, SPY, AAPL, AMGN, ... WMT)                                      │")
        logger.info(f"  │       (Rows 3-{2+n_obs} have the date and return values)                                                    │")
        logger.info("  │                                                                                                     │")
        logger.info("  │    2. Press Ctrl+T (or go to Insert tab → Table)                                                    │")
        logger.info("  │                                                                                                     │")
        logger.info("  │    3. Check [x] 'My table has headers'                                                                │")
        logger.info("  │                                                                                                     │")
        logger.info("  │    4. Click OK                                                                                      │")
        logger.info("  │                                                                                                     │")
        logger.info("  │    5. RENAME THE TABLE (important!):                                                                │")
        logger.info("  │       • Click anywhere in the table                                                                 │")
        logger.info("  │       • Go to 'Table Design' tab (appears when table is selected)                                   │")
        logger.info("  │       • In the 'Table Name' box (top left), change 'Table1' to 'LNReturns'                          │")
        logger.info("  │       • Press Enter                                                                                 │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")
        logger.info("  ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
        logger.info("  ┃  STRUCTURED REFERENCE SYNTAX                                                                        ┃")
        logger.info("  ┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛")
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  BASIC SYNTAX:                                                                                      │")
        logger.info("  │                                                                                                     │")
        logger.info("  │    TableName[ColumnName]     → References entire column (excluding header)                          │")
        logger.info("  │    TableName[[#Headers]]     → References only the header row                                       │")
        logger.info("  │    TableName[[#All]]         → References entire table including headers                            │")
        logger.info("  │    TableName[[#Data]]        → References only the data (no headers)                                │")
        logger.info("  │    TableName[@ColumnName]    → References the current row's value in that column                    │")
        logger.info("  │                                                                                                     │")
        logger.info("  │  EXAMPLES with our 'LNReturns' table:                                                               │")
        logger.info("  │                                                                                                     │")
        logger.info("  │    LNReturns[SPY]            → All SPY returns (B3:B74 equivalent)                                  │")
        logger.info("  │    LNReturns[AAPL]           → All AAPL returns (C3:C74 equivalent)                                 │")
        logger.info("  │    LNReturns[Date]           → All dates (A3:A74 equivalent)                                        │")
        logger.info("  │    LNReturns[@SPY]           → Current row's SPY value (use in calculated columns)                  │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")
        logger.info("  ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
        logger.info("  ┃  FORMULAS COMPARISON: Cell References vs Table References                                           ┃")
        logger.info("  ┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛")
        logger.info("")
        logger.info("  ┌──────────────────────────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  CALCULATING MEAN:                                                                                   │")
        logger.info("  │                                                                                                      │")
        logger.info("  │    Cell Reference:     =AVERAGE(B3:B74)                                                              │")
        logger.info("  │    Table Reference:    =AVERAGE(LNReturns[SPY])                                                      │")
        logger.info("  │                                                                                                      │")
        logger.info("  │  CALCULATING STANDARD DEVIATION:                                                                     │")
        logger.info("  │                                                                                                      │")
        logger.info("  │    Cell Reference:     =STDEV.P(B3:B74)                                                              │")
        logger.info("  │    Table Reference:    =STDEV.P(LNReturns[SPY])                                                      │")
        logger.info("  │                                                                                                      │")
        logger.info("  │  CALCULATING COVARIANCE:                                                                             │")
        logger.info("  │                                                                                                      │")
        logger.info("  │    Cell Reference:     =COVARIANCE.P(B3:B74, C3:C74)                                                 │")
        logger.info("  │    Table Reference:    =COVARIANCE.P(LNReturns[SPY], LNReturns[AAPL])                                │")
        logger.info("  │                                                                                                      │")
        logger.info("  │  CALCULATING CORRELATION:                                                                            │")
        logger.info("  │                                                                                                      │")
        logger.info("  │    Cell Reference:     =CORREL(B3:B74, C3:C74)                                                       │")
        logger.info("  │    Table Reference:    =CORREL(LNReturns[SPY], LNReturns[AAPL])                                      │")
        logger.info("  └──────────────────────────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")
        logger.info("  ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
        logger.info("  ┃  DEMEANING WITH TABLE REFERENCES                                                                    ┃")
        logger.info("  ┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛")
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  OPTION 1: Add a calculated column to the table                                                     │")
        logger.info("  │                                                                                                     │")
        logger.info("  │    1. Click on the cell to the right of the last data column in the table                           │")
        logger.info("  │    2. Type a header name: 'SPY_Demeaned'                                                             │")
        logger.info("  │    3. In the first data row, enter:                                                                  │")
        logger.info("  │                                                                                                     │")
        logger.info("  │       =[@SPY] - AVERAGE(LNReturns[SPY])                                                              │")
        logger.info("  │                                                                                                     │")
        logger.info("  │    4. The formula auto-fills for all rows!                                                           │")
        logger.info("  │                                                                                                     │")
        logger.info("  │  OPTION 2: Create a separate demeaned table (recommended for matrix math)                           │")
        logger.info("  │                                                                                                     │")
        logger.info("  │    1. In a new location, create headers matching your tickers                                        │")
        logger.info("  │    2. In the first cell under SPY header:                                                            │")
        logger.info("  │                                                                                                     │")
        logger.info("  │       =INDEX(LNReturns[SPY], ROW()-[first_data_row]+1) - AVERAGE(LNReturns[SPY])                     │")
        logger.info("  │                                                                                                     │")
        logger.info("  │    3. Or simply reference the original cell minus mean:                                              │")
        logger.info("  │                                                                                                     │")
        logger.info("  │       =LNReturns[@SPY] - AVERAGE(LNReturns[SPY])     (if in same row context)                        │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")
        logger.info("  ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
        logger.info("  ┃  MATRIX MULTIPLICATION WITH TABLE REFERENCES                                                        ┃")
        logger.info("  ┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛")
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  WARNING:  LIMITATION: MMULT does not work directly with structured table references                      │")
        logger.info("  │                                                                                                     │")
        logger.info("  │  For MMULT, you still need to use cell ranges. However, you can:                                    │")
        logger.info("  │                                                                                                     │")
        logger.info("  │    1. Use INDIRECT to convert table references to ranges:                                           │")
        logger.info("  │                                                                                                     │")
        logger.info("  │       =MMULT(TRANSPOSE(INDIRECT(\"DemeanedTable\")), INDIRECT(\"DemeanedTable\")) / ROWS(LNReturns)    │")
        logger.info("  │                                                                                                     │")
        logger.info("  │    2. Or simply reference the table's underlying range:                                              │")
        logger.info("  │                                                                                                     │")
        logger.info(f"  │       =MMULT(TRANSPOSE(B103:AF{102+n_obs}), B103:AF{102+n_obs}) / ROWS(LNReturns[SPY])                              │")
        logger.info("  │                                                                                                     │")
        logger.info("  │       Using ROWS(LNReturns[SPY]) instead of 72 makes it dynamic!                                    │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")
        logger.info("  ┏━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┓")
        logger.info("  ┃  ADVANTAGES OF USING EXCEL TABLES                                                                   ┃")
        logger.info("  ┗━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━┛")
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  [OK] READABILITY: =AVERAGE(LNReturns[SPY]) is clearer than =AVERAGE(B3:B74)                           │")
        logger.info("  │                                                                                                     │")
        logger.info("  │  [OK] AUTOMATIC EXPANSION: Add new data rows and formulas update automatically                         │")
        logger.info("  │                                                                                                     │")
        logger.info("  │  [OK] LESS ERROR-PRONE: No need to remember row numbers or update ranges                               │")
        logger.info("  │                                                                                                     │")
        logger.info("  │  [OK] SELF-DOCUMENTING: Formula clearly shows which column is being used                               │")
        logger.info("  │                                                                                                     │")
        logger.info("  │  [OK] EASIER AUDITING: Table references make formulas easier to understand                             │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  QUICK REFERENCE - ALL FORMULAS WITH TABLE 'LNReturns':                                             │")
        logger.info("  │                                                                                                     │")
        logger.info("  │    Mean of SPY:           =AVERAGE(LNReturns[SPY])                                                  │")
        logger.info("  │    Std Dev of SPY:        =STDEV.P(LNReturns[SPY])                                                  │")
        logger.info("  │    Variance of SPY:       =VAR.P(LNReturns[SPY])                                                    │")
        logger.info("  │    Cov(SPY, AAPL):        =COVARIANCE.P(LNReturns[SPY], LNReturns[AAPL])                            │")
        logger.info("  │    Corr(SPY, AAPL):       =CORREL(LNReturns[SPY], LNReturns[AAPL])                                  │")
        logger.info("  │    Count of observations: =ROWS(LNReturns[SPY])                                                     │")
        logger.info("  │    Sum of SPY returns:    =SUM(LNReturns[SPY])                                                      │")
        logger.info("  │    Min SPY return:        =MIN(LNReturns[SPY])                                                      │")
        logger.info("  │    Max SPY return:        =MAX(LNReturns[SPY])                                                      │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")

        # Log covariance matrix with Excel references
        log_dataframe(logger, cov_matrix, "COVARIANCE MATRIX VALUES", start_row=180, start_col=1)

        # Compute correlation matrix
        corr_matrix = returns_only.corr()

        # =====================================================================
        # LOG CORRELATION MATRIX WITH EXCEL INSTRUCTIONS AND VERIFICATION
        # =====================================================================
        logger.info("")
        logger.info("=" * 100)
        logger.info(f"  CORRELATION MATRIX ({len(tickers)}x{len(tickers)})")
        logger.info("=" * 100)
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  MATHEMATICAL FORMULA:                                                          │")
        logger.info("  │                                                                                 │")
        logger.info("  │    ρₓᵧ = Cov(X,Y) / (σₓ × σᵧ)                                                   │")
        logger.info("  │                                                                                 │")
        logger.info("  │    Or equivalently:                                                             │")
        logger.info("  │    ρₓᵧ = Σᵢ[(Xᵢ - μₓ)(Yᵢ - μᵧ)] / √[Σᵢ(Xᵢ - μₓ)² × Σᵢ(Yᵢ - μᵧ)²]             │")
        logger.info("  │                                                                                 │")
        logger.info("  │  VARIABLE DEFINITIONS:                                                          │")
        logger.info("  │    ρₓᵧ     = Correlation coefficient between X and Y (Greek letter 'rho')       │")
        logger.info("  │    Cov(X,Y) = Covariance between X and Y                                        │")
        logger.info("  │    σₓ, σᵧ  = Standard deviation of X and Y (sigma_x, sigma_y)                   │")
        logger.info("  │    Xᵢ, Yᵢ  = Return of asset X or Y at time i                                   │")
        logger.info("  │    μₓ, μᵧ  = Mean return of asset X or Y                                        │")
        logger.info("  │    Σᵢ      = Summation over all observations i                                  │")
        logger.info("  │    √       = Square root                                                        │")
        logger.info("  │                                                                                 │")
        logger.info("  │  PROPERTIES:                                                                    │")
        logger.info("  │    • ρₓₓ = 1.0 (diagonal = perfect correlation with itself)                     │")
        logger.info("  │    • -1 ≤ ρ ≤ +1 (bounded between -1 and 1)                                     │")
        logger.info("  │    • ρₓᵧ = ρᵧₓ (symmetric matrix)                                               │")
        logger.info("  │    • ρ = +1: perfect positive correlation                                       │")
        logger.info("  │    • ρ = -1: perfect negative correlation                                       │")
        logger.info("  │    • ρ = 0: no linear correlation                                               │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  EXCEL FORMULAS:                                                                │")
        logger.info("  │                                                                                 │")
        logger.info("  │  Method 1: Direct correlation function                                          │")
        logger.info("  │    =CORREL(ticker1_returns, ticker2_returns)                                    │")
        logger.info("  │                                                                                 │")
        logger.info("  │  Method 2: Using Covariance Matrix                                              │")
        logger.info("  │    =Cov(i,j) / SQRT(Cov(i,i) × Cov(j,j))                                        │")
        logger.info("  │    Or: =Cov(i,j) / (StdDev_i × StdDev_j)                                        │")
        logger.info("  │                                                                                 │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")
        logger.info("  ╔═════════════════════════════════════════════════════════════════════════════════╗")
        logger.info("  ║              EXCEL METHOD 3: Data Analysis ToolPak - Correlation                ║")
        logger.info("  ╚═════════════════════════════════════════════════════════════════════════════════╝")
        logger.info("")
        logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
        logger.info("  │  HOW TO USE:                                                                    │")
        logger.info("  │    1. Data tab → Data Analysis → Correlation                                    │")
        logger.info(f"  │    2. Input Range: B3:AF{n_obs + 2} (all returns data)                          │")
        logger.info("  │    3. [x] Labels in first row (if you include row 2 with tickers)                 │")
        logger.info("  │    4. Output Range: Select destination cell for top-left of matrix              │")
        logger.info("  │    5. Click OK                                                                  │")
        logger.info("  │                                                                                 │")
        logger.info("  │  OUTPUT:                                                                        │")
        logger.info(f"  │    • {len(tickers)}×{len(tickers)} correlation matrix                                               │")
        logger.info("  │    • Diagonal elements = 1.0 (asset correlated with itself)                     │")
        logger.info("  │    • Lower triangle filled (upper triangle blank - it's symmetric)              │")
        logger.info("  │                                                                                 │")
        logger.info("  │  [OK] NOTE: Unlike covariance, correlation doesn't need N vs N-1 adjustment        │")
        logger.info("  │    The N or N-1 cancels out in the formula: Cov(X,Y) / (σₓ × σᵧ)               │")
        logger.info("  │                                                                                 │")
        logger.info("  │  CELL VALUES IN INPUT RANGE (returns %):                                        │")
        logger.info(f"  │    B3={returns_only[t1].iloc[0]:>8.4f} ({t1})  C3={returns_only[t2].iloc[0]:>8.4f} ({t2})  D3={returns_only[t3].iloc[0]:>8.4f} ({t3})        │")
        logger.info(f"  │    B4={returns_only[t1].iloc[1]:>8.4f}       C4={returns_only[t2].iloc[1]:>8.4f}       D4={returns_only[t3].iloc[1]:>8.4f}             │")
        logger.info(f"  │    ...                                                                         │")
        logger.info(f"  │    B{n_obs+2}={returns_only[t1].iloc[-1]:>8.4f}      C{n_obs+2}={returns_only[t2].iloc[-1]:>8.4f}      D{n_obs+2}={returns_only[t3].iloc[-1]:>8.4f}            │")
        logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
        logger.info("")

        # Verify correlation calculation
        manual_corr = manual_cov / (stds[t1] * stds[t2])
        computed_corr = corr_matrix.loc[t1, t2]

        logger.info(f"  [OK] VERIFICATION - Correlation({t1}, {t2}):")
        logger.info(f"    Formula: Cov({t1},{t2}) / (σ_{t1} × σ_{t2})")
        logger.info(f"    = {manual_cov:.7f} / ({stds[t1]:.7f} × {stds[t2]:.7f})")
        logger.info(f"    = {manual_cov:.7f} / {stds[t1] * stds[t2]:.7f}")
        logger.info(f"    Manual calculation: {manual_corr:.7f}")
        logger.info(f"    Matrix value:       {computed_corr:.7f}")
        logger.info(f"    Match: {np.isclose(manual_corr, computed_corr)}")
        logger.info("")

        # Verify diagonal = 1
        diag_check = all(np.isclose(corr_matrix.loc[t, t], 1.0) for t in tickers)
        logger.info(f"  [OK] VERIFICATION - All diagonal elements = 1.0: {diag_check}")
        logger.info("")

        log_dataframe(logger, corr_matrix, "CORRELATION MATRIX VALUES")

        # Create the LN_Returns sheet with all data
        create_ln_returns_sheet(wb, ln_returns, date_col, means, stds, cov_matrix)
        logger.info(f"  Created LN_Returns sheet")

        # Create a Summary sheet with prices and all calculations
        create_summary_sheet(wb, prices_df, ln_returns, date_col, tickers, means, stds, cov_matrix)
        logger.info(f"  Created Summary sheet")

        # Auto-fit columns for all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for column_cells in ws.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                # Add a little extra width for padding
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 to avoid very wide columns
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save to processed folder
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_name = f"{excel_file.stem}_processed_{timestamp}.xlsx"
        output_path = PROCESSED_DIR / output_name

        wb.save(output_path)
        logger.info(f"  Saved: {output_path}")

        # Store output path and data for end of log
        global PROCESSED_FILE_PATH, STORED_COV_MATRIX, STORED_MEANS
        PROCESSED_FILE_PATH = output_path
        STORED_COV_MATRIX = cov_matrix
        STORED_MEANS = means

        # Log detailed Excel instructions
        log_excel_instructions(logger, len(ln_returns), len(tickers), tickers,
                              means, stds, cov_matrix, prices_df, ln_returns, date_col)

    except Exception as e:
        logger.error(f"  Error processing {excel_file.name}: {e}")
        import traceback
        traceback.print_exc()


# =============================================================================
# TASK SOLUTIONS
# =============================================================================

def get_task_solution(task_num: str, task_text: str) -> List[str]:
    """
    Provide solution steps for each homework task.

    Args:
        task_num: Task number (e.g., "1", "2", "10")
        task_text: Full task text

    Returns:
        List of solution lines
    """
    solutions = {
        "1": [
            "Use the provided W4E1PrepData.xlsx file.",
            "The Prices sheet contains 5 years of monthly data (Jun 2019 - Jun 2024).",
            "30 DJ stocks + SPY = 31 total columns of price data.",
            "Save file as: Lastname_Firstname_W4E1Prep.xlsx"
        ],
        "2": [
            "Review tabs: Prices, DJReturns, DJFrontier, SP Hist",
            "Do NOT delete or move existing content.",
            "Enter answers in designated spaces only."
        ],
        "3": [
            "In DJReturns sheet, cell B3 enter: =LN(Prices!B3/Prices!B4)*100",
            "Drag across to column AF (all 31 tickers)",
            "Drag down to row 74 (72 returns)",
            "",
            "Row 64 - Mean: =AVERAGE(B3:B74)  [drag across]",
            "Row 65 - StDev: =STDEV.P(B3:B74)  [drag across]",
            "",
            "Rows 69-99: Data Analysis → Covariance (paste result here)",
            "",
            "Rows 103-133: Matrix algebra covariance:",
            "Step 1: Create demeaned returns (B103:AF174): =B3-B$64",
            "Step 2: Select 31x31 area, enter array formula:",
            "        =MMULT(TRANSPOSE(B103:AF174),B103:AF174)/72"
        ],
        "4": [
            "In DJFrontier sheet:",
            "Column A (under 'Means'): Transpose mean returns from DJReturns",
            "  Use: =TRANSPOSE(DJReturns!B64:AF64) or paste special → transpose",
            "",
            "Column C: Trial weights (start with equal: =1/30)",
            "",
            "Portfolio formulas in column C:",
            "PFMean:  =SUMPRODUCT($C$10:$C$40, $A$10:$A$40)",
            "PFVar:   =MMULT(MMULT(TRANSPOSE($C$10:$C$40), CovMatrix), $C$10:$C$40)",
            "PFStD:   =SQRT(PFVar cell)"
        ],
        "5": [
            "EFFICIENT PORTFOLIO @ 4% STD DEV (save in column F):",
            "Solver: Max → PFMean",
            "        Change → weights",
            "        Constraints: SUM(weights)=1, PFStD=4",
            "",
            "EFFICIENT PORTFOLIO @ 7% STD DEV (save in column I):",
            "Solver: Max → PFMean",
            "        Change → weights",
            "        Constraints: SUM(weights)=1, PFStD=7",
            "",
            "Column L: Risky-Risky combination",
            "  =lambda*F10 + (1-lambda)*I10  [for each weight]",
            "",
            "Data Table (K10:M60):",
            "  K: Lambda values (-0.5 to 2.0)",
            "  L: =SQRT(lam^2*Var1 + (1-lam)^2*Var2 + 2*lam*(1-lam)*Cov12)",
            "  M: =lam*Mean1 + (1-lam)*Mean2"
        ],
        "6": [
            "Create Scatter Chart with Smooth Lines:",
            "1. Select data table columns (Std Dev, Mean)",
            "2. Insert → Scatter → Smooth Lines",
            "3. Add title: 'Efficient Frontier'",
            "4. X-axis: 'Standard Deviation (%)'",
            "5. Y-axis: 'Expected Return (%)'",
            "",
            "Add 30 DJ stocks as scatter points:",
            "1. Right-click → Select Data → Add Series",
            "2. X: individual stock std devs",
            "3. Y: individual stock means",
            "4. Add data labels from cells (ticker names)"
        ],
        "7": [
            "Sharpe Ratio in C7: =(PFMean - $RF$) / PFStD",
            "Where RF = 0.03% (0.0003 or 0.03 if using %)",
            "",
            "TANGENT PORTFOLIO (save in column starting at X1):",
            "Solver: Max → Sharpe Ratio",
            "        Change → weights",
            "        Constraints: SUM(weights)=1",
            "",
            "Column AB: Risky-Riskless combination",
            "  Weight in tangent: w (from 0 to 2.5)",
            "  Return: =w*TangentReturn + (1-w)*RF",
            "  Std Dev: =w*TangentStd",
            "",
            "Data Table (AA8+): populate with w, StdDev, Return"
        ],
        "8": [
            "Add CML to existing plot:",
            "1. Right-click chart → Select Data → Add Series",
            "2. X values: CML Std Dev column",
            "3. Y values: CML Return column",
            "4. Format as dashed line (no markers)",
            "",
            "Add Tangent Portfolio marker:",
            "1. Add series with single point (TangentStd, TangentReturn)",
            "2. Format as green triangle marker",
            "3. Add label 'Tangent'",
            "",
            "Rescale axes if needed to show full CML"
        ],
        "9": [
            "Add SP500 (SPY) to plot:",
            "1. Calculate SPY mean and std dev from returns",
            "2. Add as single point series",
            "3. Format with distinct marker",
            "4. Add label 'SP500' or 'SPY'"
        ],
        "10": [
            "Column AF: Equal-weighted portfolio",
            "Weight for each stock: =1/30 (≈ 0.0333)",
            "",
            "Calculate EW portfolio stats:",
            "  Mean: =SUMPRODUCT(EW_weights, means)",
            "  Var:  =MMULT(MMULT(TRANSPOSE(EW), Cov), EW)",
            "  Std:  =SQRT(Var)",
            "",
            "Add to plot as single point with label 'Equal-Weighted'"
        ],
        "11": [
            "PENSION FUND (no short selling):",
            "",
            "@ 5% Std Dev (save in column AI):",
            "Solver: Max → PFMean",
            "        Change → weights",
            "        Constraints: SUM(weights)=1, PFStD=5, weights>=0",
            "  WARNING: Add 'weights >= 0' explicitly, NOT the checkbox!",
            "",
            "@ 6% Std Dev (save in column AL):",
            "Solver: Max → PFMean",
            "        Change → weights",
            "        Constraints: SUM(weights)=1, PFStD=6, weights>=0",
            "",
            "Add both pension portfolios to plot with labels"
        ],
        "12": [
            "DISAVOWEL INVESTOR (excludes AAPL, AMGN, AXP, IBM, INTC):",
            "",
            "Set weights for AAPL, AMGN, AXP, IBM, INTC = 0",
            "",
            "@ 7% Std Dev:",
            "Solver: Max → PFMean",
            "        Change → remaining weights (25 stocks)",
            "        Constraints: SUM(weights)=1, PFStD=7, excluded=0",
            "",
            "Compare to normal investor's efficient portfolio @ 7%:",
            "  If Disavowel Return < Normal Return → Portfolio is inefficient",
            "  The difference proves the cost of excluding stocks"
        ],
        "13": [
            "To prove Equal-Weighted is inefficient:",
            "",
            "1. Calculate EW portfolio std dev (σ_EW)",
            "2. Find efficient portfolio at same std dev (σ_EW)",
            "3. Compare returns:",
            "   • EW Return < Efficient Return → EW is inefficient",
            "",
            "OR",
            "",
            "1. Calculate EW portfolio return (μ_EW)",
            "2. Find efficient portfolio at same return (μ_EW)",
            "3. Compare std devs:",
            "   • EW Std > Efficient Std → EW is inefficient"
        ],
        "14": [
            "SP500 Histogram in 'SP Hist' sheet:",
            "",
            "1. Create bins: -15%, -14%, ..., -1%, 0%, 1%, ..., 15%",
            "   (or appropriate range based on data)",
            "",
            "2. Use FREQUENCY array function:",
            "   =FREQUENCY(SPY_returns, bins)",
            "   Select output range, enter formula, Ctrl+Shift+Enter",
            "",
            "3. Create bar chart from frequency results"
        ],
        "15": [
            "Final formatting checklist:",
            "[ ] All graphs have titles",
            "[ ] All axes are labeled",
            "[ ] All data points are labeled",
            "[ ] Legend is clear and positioned well",
            "[ ] Important results highlighted with color",
            "[ ] Comments added to explain key formulas",
            "[ ] Consistent number formatting",
            "[ ] Borders and shading for organization"
        ]
    }

    return solutions.get(task_num, ["Refer to the Excel Formula Guide above for general instructions."])


# =============================================================================
# EXCEL FORMULA INSTRUCTIONS
# =============================================================================

def log_excel_instructions(logger: logging.Logger, n_returns: int, n_tickers: int, tickers: List[str],
                           means: pd.Series = None, stds: pd.Series = None, cov_matrix: pd.DataFrame = None,
                           prices_df: pd.DataFrame = None, ln_returns: pd.DataFrame = None, date_col: str = None):
    """
    Log detailed Excel formula instructions for portfolio analysis.
    """
    logger.info("")
    logger.info("╔" + "═" * 98 + "╗")
    logger.info("║" + " " * 30 + "EXCEL FORMULA GUIDE" + " " * 49 + "║")
    logger.info("║" + " " * 25 + "Step-by-Step Instructions" + " " * 48 + "║")
    logger.info("╚" + "═" * 98 + "╝")

    # =========================================================================
    # STEP 1: LN Returns
    # =========================================================================
    logger.info("")
    logger.info("┌" + "─" * 98 + "┐")
    logger.info("│  STEP 1: COMPUTING LOG RETURNS (LN Returns)" + " " * 53 + "│")
    logger.info("└" + "─" * 98 + "┘")
    logger.info("")
    logger.info("  Log returns measure continuous compounding and are additive over time.")
    logger.info("")
    logger.info("  ┌─────────────────────────────────────────────────────────────────┐")
    logger.info("  │  FORMULA: =LN(current_price / previous_price) * 100             │")
    logger.info("  └─────────────────────────────────────────────────────────────────┘")
    logger.info("")
    logger.info("  SETUP IN EXCEL:")
    logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
    logger.info("  │  1. Prices are in the 'Prices' sheet:                                           │")
    logger.info("  │     • Dates in column A (A3:A74 for 72 months)                                  │")
    logger.info("  │     • Tickers in row 2 (B2, C2, D2, ...)                                        │")
    logger.info("  │     • Prices starting at B3                                                     │")
    logger.info("  │                                                                                 │")
    logger.info("  │  2. In 'DJReturns' sheet, cell B3:                                              │")
    logger.info("  │     ┌─────────────────────────────────────┐                                     │")
    logger.info("  │     │  =LN(Prices!B3/Prices!B4)*100       │                                     │")
    logger.info("  │     └─────────────────────────────────────┘                                     │")
    logger.info("  │                                                                                 │")
    logger.info("  │  3. Drag across all ticker columns (B3 → AF3)                                   │")
    logger.info("  │  4. Drag down for all rows (row 3 → row 74)                                     │")
    logger.info("  │                                                                                 │")
    logger.info("  │  NOTE: Prices sorted newest→oldest, so B3/B4.                                   │")
    logger.info("  │        If oldest→newest, use: =LN(B4/B3)*100                                    │")
    logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
    logger.info("")

    # =========================================================================
    # STEP 2: Mean
    # =========================================================================
    logger.info("┌" + "─" * 98 + "┐")
    logger.info("│  STEP 2: COMPUTING MEAN (AVERAGE) RETURNS" + " " * 55 + "│")
    logger.info("└" + "─" * 98 + "┘")
    logger.info("")
    logger.info("  ┌─────────────────────────────────────────────────────────────────┐")
    logger.info("  │  FORMULA: =AVERAGE(range)                                       │")
    logger.info("  └─────────────────────────────────────────────────────────────────┘")
    logger.info("")
    logger.info("  EXCEL SETUP:")
    logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
    logger.info(f"  │  Row {n_returns + 5}: Label 'Mean' in column A                                            │")
    logger.info(f"  │  Cell B{n_returns + 5}:                                                                    │")
    logger.info("  │     ┌─────────────────────────────────────┐                                     │")
    logger.info(f"  │     │  =AVERAGE(B3:B{n_returns + 2})                   │                                     │")
    logger.info("  │     └─────────────────────────────────────┘                                     │")
    logger.info("  │  Drag across all ticker columns                                                 │")
    logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
    logger.info("")

    # =========================================================================
    # STEP 3: Std Dev
    # =========================================================================
    logger.info("┌" + "─" * 98 + "┐")
    logger.info("│  STEP 3: COMPUTING STANDARD DEVIATION" + " " * 59 + "│")
    logger.info("└" + "─" * 98 + "┘")
    logger.info("")
    logger.info("  • STDEV.P → Population (divides by N) ← USE THIS")
    logger.info("  • STDEV.S → Sample (divides by N-1)")
    logger.info("")
    logger.info("  ┌─────────────────────────────────────────────────────────────────┐")
    logger.info("  │  FORMULA: =STDEV.P(range)                                       │")
    logger.info("  └─────────────────────────────────────────────────────────────────┘")
    logger.info("")
    logger.info("  EXCEL SETUP:")
    logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
    logger.info(f"  │  Row {n_returns + 6}: Label 'Std Dev' in column A                                         │")
    logger.info(f"  │  Cell B{n_returns + 6}:                                                                    │")
    logger.info("  │     ┌─────────────────────────────────────┐                                     │")
    logger.info(f"  │     │  =STDEV.P(B3:B{n_returns + 2})                   │                                     │")
    logger.info("  │     └─────────────────────────────────────┘                                     │")
    logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
    logger.info("")

    # =========================================================================
    # STEP 4: Covariance Matrix with MMULT
    # =========================================================================
    logger.info("┌" + "─" * 98 + "┐")
    logger.info("│  STEP 4: COMPUTING COVARIANCE MATRIX USING MMULT" + " " * 48 + "│")
    logger.info("└" + "─" * 98 + "┘")
    logger.info("")
    logger.info("  Mathematical Formula:")
    logger.info("  ┌─────────────────────────────────────────────────────────────────┐")
    logger.info("  │  Cov = (1/N) × (R - μ)ᵀ × (R - μ)                               │")
    logger.info("  │  Where R = returns matrix, μ = mean vector                      │")
    logger.info("  └─────────────────────────────────────────────────────────────────┘")
    logger.info("")
    logger.info("  ╔═════════════════════════════════════════════════════════════════════════════════╗")
    logger.info("  ║  METHOD 1: MMULT (Recommended)                                                  ║")
    logger.info("  ╠═════════════════════════════════════════════════════════════════════════════════╣")
    logger.info("  ║                                                                                 ║")
    logger.info("  ║  Step A: Create demeaned returns matrix                                         ║")
    logger.info("  ║  ┌─────────────────────────────────────────────────────────────┐                ║")
    logger.info(f"  ║  │  Cell formula: =B3-B${n_returns + 5}                                  │                ║")
    logger.info("  ║  │  ($ keeps mean row fixed when dragging)                     │                ║")
    logger.info("  ║  └─────────────────────────────────────────────────────────────┘                ║")
    logger.info("  ║                                                                                 ║")
    logger.info(f"  ║  Step B: Demeaned returns in range B103:AF{103 + n_returns - 1}                             ║")
    logger.info("  ║                                                                                 ║")
    logger.info(f"  ║  Step C: Select a {n_tickers}×{n_tickers} range for covariance matrix                              ║")
    logger.info("  ║                                                                                 ║")
    logger.info("  ║  Step D: Enter ARRAY FORMULA (Ctrl+Shift+Enter in older Excel):                 ║")
    logger.info("  ║  ┌─────────────────────────────────────────────────────────────────────────┐    ║")
    logger.info(f"  ║  │  =MMULT(TRANSPOSE(B103:AF{103 + n_returns - 1}), B103:AF{103 + n_returns - 1})/{n_returns}             │    ║")
    logger.info("  ║  └─────────────────────────────────────────────────────────────────────────┘    ║")
    logger.info("  ║                                                                                 ║")
    logger.info("  ║  In Excel 365: Just press Enter - it auto-spills!                               ║")
    logger.info("  ╚═════════════════════════════════════════════════════════════════════════════════╝")
    logger.info("")
    logger.info("  ╔═════════════════════════════════════════════════════════════════════════════════╗")
    logger.info("  ║  METHOD 2: Data Analysis ToolPak                                                ║")
    logger.info("  ╠═════════════════════════════════════════════════════════════════════════════════╣")
    logger.info("  ║  1. Data tab → Data Analysis                                                    ║")
    logger.info("  ║  2. Select 'Covariance'                                                         ║")
    logger.info(f"  ║  3. Input Range: B3:AF{n_returns + 2}                                                       ║")
    logger.info("  ║  4. Check 'Labels in first row'                                                 ║")
    logger.info("  ║  5. Output Range: Select destination                                            ║")
    logger.info("  ║  6. Click OK                                                                    ║")
    logger.info("  ║                                                                                 ║")
    logger.info("  ║  WARNING:  NOTE: Data Analysis uses N-1 (sample).                                      ║")
    logger.info("  ║     Multiply result by (N-1)/N to get population covariance.                    ║")
    logger.info("  ╚═════════════════════════════════════════════════════════════════════════════════╝")
    logger.info("")

    # =========================================================================
    # STEP 5: Portfolio Variance using MMULT
    # =========================================================================
    logger.info("┌" + "─" * 98 + "┐")
    logger.info("│  STEP 5: PORTFOLIO VARIANCE/STD DEV USING MMULT" + " " * 49 + "│")
    logger.info("└" + "─" * 98 + "┘")
    logger.info("")
    logger.info("  ┌─────────────────────────────────────────────────────────────────┐")
    logger.info("  │  σ²ₚ = wᵀ × Σ × w                                               │")
    logger.info("  │  Where w = weights, Σ = covariance matrix                       │")
    logger.info("  └─────────────────────────────────────────────────────────────────┘")
    logger.info("")
    logger.info("  EXCEL FORMULAS:")
    logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
    logger.info("  │  Assume:                                                                        │")
    logger.info("  │  • Weights in C10:C40 (31 assets)                                               │")
    logger.info("  │  • Covariance matrix in E10:AI40 (31×31)                                        │")
    logger.info("  │                                                                                 │")
    logger.info("  │  Portfolio Variance:                                                            │")
    logger.info("  │  ┌───────────────────────────────────────────────────────────────────────┐      │")
    logger.info("  │  │  =MMULT(MMULT(TRANSPOSE(C10:C40), E10:AI40), C10:C40)                  │      │")
    logger.info("  │  └───────────────────────────────────────────────────────────────────────┘      │")
    logger.info("  │                                                                                 │")
    logger.info("  │  Portfolio Std Dev:                                                             │")
    logger.info("  │  ┌───────────────────────────────────────────────────────────────────────┐      │")
    logger.info("  │  │  =SQRT(MMULT(MMULT(TRANSPOSE(C10:C40), E10:AI40), C10:C40))            │      │")
    logger.info("  │  └───────────────────────────────────────────────────────────────────────┘      │")
    logger.info("  │                                                                                 │")
    logger.info("  │  Portfolio Mean:                                                                │")
    logger.info("  │  ┌───────────────────────────────────────────────────────────────────────┐      │")
    logger.info("  │  │  =SUMPRODUCT(C10:C40, A10:A40)                                         │      │")
    logger.info("  │  └───────────────────────────────────────────────────────────────────────┘      │")
    logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
    logger.info("")

    # =========================================================================
    # STEP 6: Sharpe Ratio
    # =========================================================================
    logger.info("┌" + "─" * 98 + "┐")
    logger.info("│  STEP 6: COMPUTING SHARPE RATIO" + " " * 65 + "│")
    logger.info("└" + "─" * 98 + "┘")
    logger.info("")
    logger.info("  ┌─────────────────────────────────────────────────────────────────┐")
    logger.info("  │  Sharpe = (μₚ - Rᶠ) / σₚ                                        │")
    logger.info("  └─────────────────────────────────────────────────────────────────┘")
    logger.info("")
    logger.info("  EXCEL FORMULA (RF = 0.03% in cell G1):")
    logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
    logger.info("  │  ┌───────────────────────────────────────────────────────────────────────┐      │")
    logger.info("  │  │  =(C5 - $G$1) / C7                                                     │      │")
    logger.info("  │  └───────────────────────────────────────────────────────────────────────┘      │")
    logger.info("  │  Where C5 = Portfolio Mean, C7 = Portfolio Std Dev                              │")
    logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
    logger.info("")

    # =========================================================================
    # STEP 7: Efficient Frontier Graph
    # =========================================================================
    logger.info("┌" + "─" * 98 + "┐")
    logger.info("│  STEP 7: CREATING THE EFFICIENT FRONTIER CURVE" + " " * 50 + "│")
    logger.info("└" + "─" * 98 + "┘")
    logger.info("")
    logger.info("  TWO-FUND SEPARATION THEOREM:")
    logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
    logger.info("  │  Any efficient portfolio = linear combination of two efficient portfolios       │")
    logger.info("  │                                                                                 │")
    logger.info("  │  Portfolio 1: Efficient at 4% std dev                                           │")
    logger.info("  │  Portfolio 2: Efficient at 7% std dev                                           │")
    logger.info("  │                                                                                 │")
    logger.info("  │  Combined Return:  μc = λ·μ₁ + (1-λ)·μ₂                                         │")
    logger.info("  │  Combined Variance: σ²c = λ²σ₁² + (1-λ)²σ₂² + 2λ(1-λ)Cov₁₂                      │")
    logger.info("  │  Combined Std Dev:  σc = √(σ²c)                                                 │")
    logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
    logger.info("")
    logger.info("  EXCEL DATA TABLE SETUP:")
    logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
    logger.info("  │     Column K          Column L              Column M                            │")
    logger.info("  │  ┌───────────┐    ┌────────────────┐    ┌────────────────┐                      │")
    logger.info("  │  │  Lambda   │    │   Std Dev      │    │   Mean Return  │                      │")
    logger.info("  │  ├───────────┤    ├────────────────┤    ├────────────────┤                      │")
    logger.info("  │  │   -0.50   │    │  =SQRT(...)    │    │  =λ*μ1+(1-λ)*μ2│                      │")
    logger.info("  │  │   -0.45   │    │                │    │                │                      │")
    logger.info("  │  │   -0.40   │    │                │    │                │                      │")
    logger.info("  │  │    ...    │    │                │    │                │                      │")
    logger.info("  │  │    2.00   │    │                │    │                │                      │")
    logger.info("  │  └───────────┘    └────────────────┘    └────────────────┘                      │")
    logger.info("  │                                                                                 │")
    logger.info("  │  Cov₁₂ formula:                                                                 │")
    logger.info("  │  ┌───────────────────────────────────────────────────────────────────────┐      │")
    logger.info("  │  │  =MMULT(MMULT(TRANSPOSE(weights1), CovMatrix), weights2)               │      │")
    logger.info("  │  └───────────────────────────────────────────────────────────────────────┘      │")
    logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
    logger.info("")

    # =========================================================================
    # STEP 8: Creating the Scatter Chart
    # =========================================================================
    logger.info("┌" + "─" * 98 + "┐")
    logger.info("│  STEP 8: CREATING THE SCATTER CHART WITH STOCK LABELS" + " " * 43 + "│")
    logger.info("└" + "─" * 98 + "┘")
    logger.info("")
    logger.info("  ╔═════════════════════════════════════════════════════════════════════════════════╗")
    logger.info("  ║  A. CREATE BASE CHART                                                           ║")
    logger.info("  ╠═════════════════════════════════════════════════════════════════════════════════╣")
    logger.info("  ║  1. Select frontier data (Std Dev = X, Return = Y)                              ║")
    logger.info("  ║  2. Insert → Charts → Scatter → Scatter with Smooth Lines                       ║")
    logger.info("  ╚═════════════════════════════════════════════════════════════════════════════════╝")
    logger.info("")
    logger.info("  ╔═════════════════════════════════════════════════════════════════════════════════╗")
    logger.info("  ║  B. ADD INDIVIDUAL STOCKS AS POINTS                                             ║")
    logger.info("  ╠═════════════════════════════════════════════════════════════════════════════════╣")
    logger.info("  ║  1. Right-click chart → Select Data → Add Series                                ║")
    logger.info("  ║  2. Series name: 'Individual Stocks'                                            ║")
    logger.info("  ║  3. X values: =Sheet!$D$10:$D$40 (std devs)                                      ║")
    logger.info("  ║  4. Y values: =Sheet!$A$10:$A$40 (means)                                         ║")
    logger.info("  ║  5. Format: Markers only, Circle, Size 8, No line                               ║")
    logger.info("  ╚═════════════════════════════════════════════════════════════════════════════════╝")
    logger.info("")
    logger.info("  ╔═════════════════════════════════════════════════════════════════════════════════╗")
    logger.info("  ║  C. ADD TICKER LABELS TO STOCK POINTS                                           ║")
    logger.info("  ╠═════════════════════════════════════════════════════════════════════════════════╣")
    logger.info("  ║  1. Click on stock points series                                                ║")
    logger.info("  ║  2. Chart Design → Add Chart Element → Data Labels → More Options               ║")
    logger.info("  ║  3. Check [x] 'Value From Cells'                                                  ║")
    logger.info("  ║  4. Select range with ticker names                                              ║")
    logger.info("  ║  5. Uncheck [ ] 'Y Value'                                                         ║")
    logger.info("  ║  6. Label Position: Right or Above                                              ║")
    logger.info("  ╚═════════════════════════════════════════════════════════════════════════════════╝")
    logger.info("")

    # =========================================================================
    # STEP 9: Capital Market Line
    # =========================================================================
    logger.info("┌" + "─" * 98 + "┐")
    logger.info("│  STEP 9: ADDING THE CAPITAL MARKET LINE (CML / TANGENT LINE)" + " " * 36 + "│")
    logger.info("└" + "─" * 98 + "┘")
    logger.info("")
    logger.info("  The CML connects Risk-Free Rate to Tangent Portfolio:")
    logger.info("")
    logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
    logger.info("  │                                                                                 │")
    logger.info("  │     Return ↑                                                                    │")
    logger.info("  │            │                           ╱ CML (dashed line)                      │")
    logger.info("  │            │                         ╱                                          │")
    logger.info("  │            │                    ▲ ╱  Tangent Portfolio                          │")
    logger.info("  │            │                  ╱                                                 │")
    logger.info("  │            │             ╭──╱──╮  Efficient Frontier (curve)                    │")
    logger.info("  │            │           ╱ ╱     ╲                                                │")
    logger.info("  │            │         ╱ ╱       ╲                                                │")
    logger.info("  │            │       ╱■╱          MVP                                             │")
    logger.info("  │          ★ Rf    ╱                                                              │")
    logger.info("  │            │   ╱                                                                │")
    logger.info("  │            └───────────────────────────────→ Std Dev                            │")
    logger.info("  │                                                                                 │")
    logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
    logger.info("")
    logger.info("  CML FORMULAS:")
    logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
    logger.info("  │  For weight w in tangent portfolio (0 = all RF, 1 = all tangent):               │")
    logger.info("  │                                                                                 │")
    logger.info("  │  Return  = w × μₜₐₙ + (1-w) × Rᶠ                                                │")
    logger.info("  │  Std Dev = w × σₜₐₙ                                                             │")
    logger.info("  │                                                                                 │")
    logger.info("  │  Excel formulas (w values 0 to 2.5):                                            │")
    logger.info("  │  ┌───────────────────────────────────────────────────────────────────────┐      │")
    logger.info("  │  │  Std Dev: =AA10 * $TangentStd$                                         │      │")
    logger.info("  │  │  Return:  =AA10 * $TangentRet$ + (1-AA10) * $RF$                       │      │")
    logger.info("  │  └───────────────────────────────────────────────────────────────────────┘      │")
    logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
    logger.info("")

    # =========================================================================
    # STEP 10: Solver Setup
    # =========================================================================
    logger.info("┌" + "─" * 98 + "┐")
    logger.info("│  STEP 10: USING EXCEL SOLVER FOR OPTIMIZATION" + " " * 51 + "│")
    logger.info("└" + "─" * 98 + "┘")
    logger.info("")
    logger.info("  ╔═══════════════════════════════════════════════════════════════════════════════════════════╗")
    logger.info("  ║                                                                                           ║")
    logger.info("  ║   ┌─────────────────────────────── SOLVER DIALOG ───────────────────────────────┐         ║")
    logger.info("  ║   │                                                                             │         ║")
    logger.info("  ║   │  Set Objective:    [$C$5          ▼]                                        │         ║")
    logger.info("  ║   │                                                                             │         ║")
    logger.info("  ║   │  To:   ○ Max   ○ Min   ○ Value of: [        ]                               │         ║")
    logger.info("  ║   │                                                                             │         ║")
    logger.info("  ║   │  By Changing Variable Cells:                                                │         ║")
    logger.info("  ║   │  [$C$10:$C$40                                              ]                │         ║")
    logger.info("  ║   │                                                                             │         ║")
    logger.info("  ║   │  Subject to the Constraints:                                                │         ║")
    logger.info("  ║   │  ┌─────────────────────────────────────────────────────────┐                │         ║")
    logger.info("  ║   │  │  $C$41 = 1        (sum of weights = 1)                  │ [Add]          │         ║")
    logger.info("  ║   │  │  $C$7 = 0.04      (std dev = 4%, for target)            │ [Change]       │         ║")
    logger.info("  ║   │  │  $C$10:$C$40 >= 0 (no short selling)                    │ [Delete]       │         ║")
    logger.info("  ║   │  └─────────────────────────────────────────────────────────┘                │         ║")
    logger.info("  ║   │                                                                             │         ║")
    logger.info("  ║   │  [ ] Make Unconstrained Variables Non-Negative  ← DON'T USE THIS!             │         ║")
    logger.info("  ║   │                                                                             │         ║")
    logger.info("  ║   │  Select Solving Method:  [GRG Nonlinear     ▼]                              │         ║")
    logger.info("  ║   │                                                                             │         ║")
    logger.info("  ║   │                    [  Solve  ]    [ Close ]                                 │         ║")
    logger.info("  ║   └─────────────────────────────────────────────────────────────────────────────┘         ║")
    logger.info("  ║                                                                                           ║")
    logger.info("  ╚═══════════════════════════════════════════════════════════════════════════════════════════╝")
    logger.info("")
    logger.info("  OPTIMIZATION SCENARIOS:")
    logger.info("")
    logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
    logger.info("  │  1. MINIMUM VARIANCE PORTFOLIO (MVP)                                            │")
    logger.info("  │     • Objective: Portfolio Variance → MIN                                       │")
    logger.info("  │     • Constraint: SUM(weights) = 1                                              │")
    logger.info("  ├─────────────────────────────────────────────────────────────────────────────────┤")
    logger.info("  │  2. EFFICIENT PORTFOLIO AT TARGET STD DEV (e.g., 4%)                            │")
    logger.info("  │     • Objective: Portfolio Return → MAX                                         │")
    logger.info("  │     • Constraints: SUM(weights) = 1, StdDev = 4%                                │")
    logger.info("  ├─────────────────────────────────────────────────────────────────────────────────┤")
    logger.info("  │  3. TANGENT PORTFOLIO (Max Sharpe)                                              │")
    logger.info("  │     • Objective: Sharpe Ratio → MAX                                             │")
    logger.info("  │     • Constraint: SUM(weights) = 1                                              │")
    logger.info("  ├─────────────────────────────────────────────────────────────────────────────────┤")
    logger.info("  │  4. NO SHORT SELLING (Pension Fund)                                             │")
    logger.info("  │     • Add constraint: weights >= 0                                              │")
    logger.info("  │     WARNING:  DON'T use 'Make Non-Negative' checkbox!                                  │")
    logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
    logger.info("")

    # =========================================================================
    # STEP 11: Final Chart Formatting
    # =========================================================================
    logger.info("┌" + "─" * 98 + "┐")
    logger.info("│  STEP 11: FINAL CHART FORMATTING" + " " * 63 + "│")
    logger.info("└" + "─" * 98 + "┘")
    logger.info("")
    logger.info("  ┌─────────────────────────────────────────────────────────────────────────────────┐")
    logger.info("  │  AXIS LABELS:                                                                   │")
    logger.info("  │  • X-axis: 'Standard Deviation (%)'                                             │")
    logger.info("  │  • Y-axis: 'Expected Return (%)'                                                │")
    logger.info("  │                                                                                 │")
    logger.info("  │  CHART TITLE: 'Efficient Frontier & Capital Market Line'                        │")
    logger.info("  │                                                                                 │")
    logger.info("  │  LEGEND: Upper left, include all series                                         │")
    logger.info("  │                                                                                 │")
    logger.info("  │  SPECIAL MARKERS:                                                               │")
    logger.info("  │  • MVP: ■ Red square                                                            │")
    logger.info("  │  • Tangent: ▲ Green triangle                                                    │")
    logger.info("  │  • Risk-free: ★ Gold star at (0, Rf)                                            │")
    logger.info("  │                                                                                 │")
    logger.info("  │  GRIDLINES: Major gridlines for readability                                     │")
    logger.info("  └─────────────────────────────────────────────────────────────────────────────────┘")
    logger.info("")
    logger.info("╔" + "═" * 98 + "╗")
    logger.info("║" + " " * 35 + "END OF EXCEL GUIDE" + " " * 45 + "║")
    logger.info("╚" + "═" * 98 + "╝")
    logger.info("")


# =============================================================================
# MAIN
# =============================================================================

def main():
    """Main entry point."""
    # Set up logging
    logger = setup_logging()

    logger.info("=" * 70)
    logger.info("INPUT DATA PROCESSOR")
    logger.info("=" * 70)
    logger.info(f"Project Root: {PROJECT_ROOT}")
    logger.info(f"Excel Input: {EXCEL_INPUT_DIR}")
    logger.info(f"PDF Input: {PDF_INPUT_DIR}")
    logger.info(f"Processed Output: {PROCESSED_DIR}")
    logger.info(f"Archive: {ARCHIVE_DIR}")

    # Check for input files
    excel_files = list(EXCEL_INPUT_DIR.glob("*.xlsx")) + list(EXCEL_INPUT_DIR.glob("*.xls"))
    pdf_files = list(PDF_INPUT_DIR.glob("*.pdf"))

    logger.info(f"Found {len(excel_files)} Excel file(s)")
    for ef in excel_files:
        logger.info(f"  - {ef.name}")

    logger.info(f"Found {len(pdf_files)} PDF file(s)")
    for pf in pdf_files:
        logger.info(f"  - {pf.name}")

    if not excel_files and not pdf_files:
        logger.warning("No input files found!")
        logger.info(f"  Place Excel files in: {EXCEL_INPUT_DIR}")
        logger.info(f"  Place PDF files in: {PDF_INPUT_DIR}")
        return

    # Process PDFs to TODO
    if pdf_files:
        logger.info("-" * 70)
        logger.info("PROCESSING PDFs -> TODO.md")
        logger.info("-" * 70)

        # Archive old TODO
        archive_old_todo()

        # Create new TODO
        todo_content = process_pdfs_to_todo()
        todo_path = PROJECT_ROOT / "TODO.md"

        with open(todo_path, 'w', encoding='utf-8') as f:
            f.write(todo_content)

        logger.info(f"Created: {todo_path}")

        # Log the TODO items extracted
        logger.info("-" * 70)
        logger.info("TODO ITEMS FROM PDFs:")
        logger.info("-" * 70)
        for pdf_file in pdf_files:
            logger.info(f"From: {pdf_file.name}")
            pdf_text = read_pdf(pdf_file)
            # Extract numbered items (likely homework tasks)
            lines = pdf_text.split('\n')
            for line in lines:
                line = line.strip()
                # Look for numbered items like "1)", "2)", etc.
                if line and len(line) > 2 and line[0].isdigit() and ')' in line[:3]:
                    logger.info(f"  {line[:100]}{'...' if len(line) > 100 else ''}")

    # Process Excel files
    if excel_files:
        logger.info("-" * 70)
        logger.info("PROCESSING EXCEL FILES")
        logger.info("-" * 70)

        for excel_file in excel_files:
            logger.info(f"Processing Excel: {excel_file.name}")
            process_single_excel(excel_file, logger)

    # =========================================================================
    # TASKS AND SOLUTIONS SECTION
    # =========================================================================
    if pdf_files:
        logger.info("")
        logger.info("╔" + "═" * 98 + "╗")
        logger.info("║" + " " * 30 + "TASKS AND SOLUTIONS" + " " * 49 + "║")
        logger.info("╚" + "═" * 98 + "╝")

        for pdf_file in pdf_files:
            pdf_text = read_pdf(pdf_file)
            lines = pdf_text.split('\n')

            tasks = []
            for line in lines:
                line = line.strip()
                if line and len(line) > 2 and line[0].isdigit() and ')' in line[:4]:
                    tasks.append(line)

            logger.info("")
            logger.info(f"FROM: {pdf_file.name}")
            logger.info("=" * 100)

            for task in tasks:
                # Extract task number
                task_num = task.split(')')[0].strip()

                logger.info("")
                logger.info("┌" + "─" * 98 + "┐")
                logger.info(f"│  TASK {task_num}" + " " * (90 - len(f"TASK {task_num}")) + "│")
                logger.info("├" + "─" * 98 + "┤")

                # Wrap long task text
                task_text = task
                while len(task_text) > 95:
                    logger.info(f"│  {task_text[:95]}" + " │")
                    task_text = task_text[95:]
                logger.info(f"│  {task_text}" + " " * (96 - len(task_text)) + "│")
                logger.info("├" + "─" * 98 + "┤")
                logger.info("│  SOLUTION:" + " " * 87 + "│")
                logger.info("│" + " " * 98 + "│")

                # Provide solutions based on task number
                solutions = get_task_solution(task_num, task)
                for sol_line in solutions:
                    if len(sol_line) > 95:
                        while len(sol_line) > 95:
                            logger.info(f"│  {sol_line[:95]}" + " │")
                            sol_line = sol_line[95:]
                        if sol_line:
                            logger.info(f"│  {sol_line}" + " " * (96 - len(sol_line)) + "│")
                    else:
                        logger.info(f"│  {sol_line}" + " " * (96 - len(sol_line)) + "│")

                logger.info("│" + " " * 98 + "│")
                logger.info("└" + "─" * 98 + "┘")

    # =========================================================================
    # PYTHON COMPUTED SOLUTIONS
    # =========================================================================
    if STORED_COV_MATRIX is not None and STORED_MEANS is not None:
        logger.info("")
        logger.info("╔" + "═" * 98 + "╗")
        logger.info("║" + " " * 28 + "PYTHON COMPUTED SOLUTIONS" + " " * 45 + "║")
        logger.info("╠" + "═" * 98 + "╣")
        logger.info("║" + " " * 98 + "║")
        logger.info("║  These are the actual portfolio optimization results computed by Python using scipy.optimize." + " " * 5 + "║")
        logger.info("║  Use these to verify your Excel Solver results.                                               " + " " * 2 + "║")
        logger.info("║" + " " * 98 + "║")

        cov_matrix = STORED_COV_MATRIX
        means_pct = STORED_MEANS  # Means in percentage form
        means_dec = means_pct / 100  # Convert to decimal
        tickers = list(means_pct.index)
        n_assets = len(tickers)
        rf = 0.0003  # Risk-free rate = 0.03% per month (as decimal)

        # Portfolio optimization functions
        def portfolio_return(weights, means):
            return np.sum(weights * means)

        def portfolio_variance(weights, cov_matrix):
            return np.dot(weights.T, np.dot(cov_matrix.values, weights))

        def portfolio_std(weights, cov_matrix):
            return np.sqrt(portfolio_variance(weights, cov_matrix))

        def neg_sharpe_ratio(weights, means, cov_matrix, rf):
            ret = portfolio_return(weights, means)
            std = portfolio_std(weights, cov_matrix)
            return -(ret - rf) / std if std > 0 else 0

        # Constraints: weights sum to 1
        constraints = [{'type': 'eq', 'fun': lambda w: np.sum(w) - 1}]

        # Initial guess: equal weights
        init_weights = np.array([1/n_assets] * n_assets)

        # ---------------------------------------------------------------------
        # 1. Equal-Weighted Portfolio
        # ---------------------------------------------------------------------
        ew_weights = np.array([1/n_assets] * n_assets)
        ew_return = portfolio_return(ew_weights, means_dec.values)
        ew_std = portfolio_std(ew_weights, cov_matrix)
        ew_var = ew_std ** 2

        logger.info("║" + "─" * 98 + "║")
        logger.info("║  1. EQUAL-WEIGHTED PORTFOLIO (1/N for each stock)                                             " + " " * 2 + "║")
        logger.info("║" + "─" * 98 + "║")
        logger.info(f"║     Portfolio Return:   {ew_return*100:>12.6f}% per month                                            " + " " * 2 + "║")
        logger.info(f"║     Portfolio Std Dev:  {ew_std*100:>12.6f}% per month                                            " + " " * 2 + "║")
        logger.info(f"║     Portfolio Variance: {ew_var:>12.9f} (decimal form)                                       " + " " * 2 + "║")
        logger.info("║" + " " * 98 + "║")

        # ---------------------------------------------------------------------
        # 2. Tangent Portfolio (Maximum Sharpe Ratio)
        # ---------------------------------------------------------------------
        try:
            result = minimize(
                neg_sharpe_ratio,
                init_weights,
                args=(means_dec.values, cov_matrix, rf),
                method='SLSQP',
                constraints=constraints
            )
            tan_weights = result.x
            tan_return = portfolio_return(tan_weights, means_dec.values)
            tan_std = portfolio_std(tan_weights, cov_matrix)
            tan_var = tan_std ** 2
            tan_sharpe = (tan_return - rf) / tan_std

            logger.info("║" + "─" * 98 + "║")
            logger.info("║  2. TANGENT PORTFOLIO (Maximum Sharpe Ratio, RF = 0.03%)                                      " + " " * 2 + "║")
            logger.info("║" + "─" * 98 + "║")
            logger.info(f"║     Portfolio Return:   {tan_return*100:>12.6f}% per month                                            " + " " * 2 + "║")
            logger.info(f"║     Portfolio Std Dev:  {tan_std*100:>12.6f}% per month                                            " + " " * 2 + "║")
            logger.info(f"║     Portfolio Variance: {tan_var:>12.9f} (decimal form)                                       " + " " * 2 + "║")
            logger.info(f"║     Sharpe Ratio:       {tan_sharpe:>12.6f}                                                       " + " " * 2 + "║")
            logger.info("║" + " " * 98 + "║")
            logger.info("║     Top 5 Weights:                                                                            " + " " * 2 + "║")
            sorted_idx = np.argsort(tan_weights)[::-1][:5]
            for i in sorted_idx:
                logger.info(f"║       {tickers[i]:>6}: {tan_weights[i]*100:>10.4f}%                                                              " + " " * 2 + "║")
            logger.info("║" + " " * 98 + "║")
        except Exception as e:
            logger.info(f"║     Error computing tangent portfolio: {str(e)[:50]}" + " " * 40 + "║")

        # ---------------------------------------------------------------------
        # 3. Efficient Portfolios at Target Std Devs (Unconstrained)
        # ---------------------------------------------------------------------
        target_stds = [0.04, 0.05, 0.06, 0.07]  # 4%, 5%, 6%, 7% in decimal

        logger.info("║" + "─" * 98 + "║")
        logger.info("║  3. EFFICIENT PORTFOLIOS (Unconstrained - Max Return for Target Std Dev)                       " + " " * 2 + "║")
        logger.info("║" + "─" * 98 + "║")

        for target_std in target_stds:
            try:
                # Minimize negative return (maximize return) subject to std constraint
                std_constraint = {'type': 'eq', 'fun': lambda w, ts=target_std: portfolio_std(w, cov_matrix) - ts}
                result = minimize(
                    lambda w: -portfolio_return(w, means_dec.values),
                    init_weights,
                    method='SLSQP',
                    constraints=[constraints[0], std_constraint]
                )
                eff_weights = result.x
                eff_return = portfolio_return(eff_weights, means_dec.values)
                eff_std = portfolio_std(eff_weights, cov_matrix)
                eff_var = eff_std ** 2

                logger.info(f"║     Target Std = {target_std*100:.0f}%:  Return = {eff_return*100:>10.6f}%,  Actual Std = {eff_std*100:>10.6f}%,  Var = {eff_var:>12.9f}" + " " * 2 + "║")
            except Exception as e:
                logger.info(f"║     Target Std = {target_std*100:.0f}%:  Error - {str(e)[:60]}" + " " * 20 + "║")

        logger.info("║" + " " * 98 + "║")

        # ---------------------------------------------------------------------
        # 4. Pension Fund Portfolios (Long-Only, No Short Selling)
        # ---------------------------------------------------------------------
        logger.info("║" + "─" * 98 + "║")
        logger.info("║  4. PENSION FUND PORTFOLIOS (Long-Only: weights >= 0)                                          " + " " * 2 + "║")
        logger.info("║" + "─" * 98 + "║")

        bounds = [(0, 1) for _ in range(n_assets)]  # No short selling

        for target_std in [0.05, 0.06]:  # 5% and 6%
            try:
                std_constraint = {'type': 'eq', 'fun': lambda w, ts=target_std: portfolio_std(w, cov_matrix) - ts}
                result = minimize(
                    lambda w: -portfolio_return(w, means_dec.values),
                    init_weights,
                    method='SLSQP',
                    bounds=bounds,
                    constraints=[constraints[0], std_constraint]
                )
                pf_weights = result.x
                pf_return = portfolio_return(pf_weights, means_dec.values)
                pf_std = portfolio_std(pf_weights, cov_matrix)
                pf_var = pf_std ** 2

                logger.info(f"║     Target Std = {target_std*100:.0f}%:  Return = {pf_return*100:>10.6f}%,  Actual Std = {pf_std*100:>10.6f}%,  Var = {pf_var:>12.9f}" + " " * 2 + "║")
            except Exception as e:
                logger.info(f"║     Target Std = {target_std*100:.0f}%:  Error - {str(e)[:60]}" + " " * 20 + "║")

        logger.info("║" + " " * 98 + "║")

        # ---------------------------------------------------------------------
        # 5. Disavowel Investor (Excludes AAPL, AMGN, AXP, IBM, INTC)
        # ---------------------------------------------------------------------
        excluded = ['AAPL', 'AMGN', 'AXP', 'IBM', 'INTC']
        included_idx = [i for i, t in enumerate(tickers) if t not in excluded]
        included_tickers = [tickers[i] for i in included_idx]

        logger.info("║" + "─" * 98 + "║")
        logger.info("║  5. DISAVOWEL INVESTOR (Excludes: AAPL, AMGN, AXP, IBM, INTC)                                   " + " " * 2 + "║")
        logger.info("║" + "─" * 98 + "║")

        if len(included_idx) > 0:
            # Create sub-covariance matrix and sub-means
            sub_cov = cov_matrix.iloc[included_idx, included_idx]
            sub_means = means_dec.iloc[included_idx]
            n_sub = len(included_idx)
            sub_init = np.array([1/n_sub] * n_sub)
            sub_constraints = [{'type': 'eq', 'fun': lambda w: np.sum(w) - 1}]

            try:
                target_std = 0.07
                std_constraint = {'type': 'eq', 'fun': lambda w, ts=target_std: np.sqrt(np.dot(w.T, np.dot(sub_cov.values, w))) - ts}
                result = minimize(
                    lambda w: -np.sum(w * sub_means.values),
                    sub_init,
                    method='SLSQP',
                    constraints=[sub_constraints[0], std_constraint]
                )
                dis_weights = result.x
                dis_return = np.sum(dis_weights * sub_means.values)
                dis_std = np.sqrt(np.dot(dis_weights.T, np.dot(sub_cov.values, dis_weights)))
                dis_var = dis_std ** 2

                logger.info(f"║     Target Std = 7%:   Return = {dis_return*100:>10.6f}%,  Actual Std = {dis_std*100:>10.6f}%,  Var = {dis_var:>12.9f}" + " " * 2 + "║")

                # Compare with normal investor at 7%
                std_constraint_normal = {'type': 'eq', 'fun': lambda w: portfolio_std(w, cov_matrix) - 0.07}
                result_normal = minimize(
                    lambda w: -portfolio_return(w, means_dec.values),
                    init_weights,
                    method='SLSQP',
                    constraints=[constraints[0], std_constraint_normal]
                )
                normal_return = portfolio_return(result_normal.x, means_dec.values)
                logger.info(f"║     Normal Investor at 7%: Return = {normal_return*100:>10.6f}%                                      " + " " * 2 + "║")
                logger.info(f"║     Difference: {(dis_return - normal_return)*100:>10.6f}% (Disavowel is {'inefficient' if dis_return < normal_return else 'efficient'})" + " " * 26 + "║")
            except Exception as e:
                logger.info(f"║     Error: {str(e)[:70]}" + " " * 20 + "║")

        logger.info("║" + " " * 98 + "║")
        logger.info("╚" + "═" * 98 + "╝")
        logger.info("")

    # =========================================================================
    # OUTPUT FILE LOCATION
    # =========================================================================
    logger.info("")
    logger.info("╔" + "═" * 98 + "╗")
    logger.info("║" + " " * 30 + "OUTPUT FILE LOCATION" + " " * 48 + "║")
    logger.info("╠" + "═" * 98 + "╣")
    logger.info("║" + " " * 98 + "║")

    if PROCESSED_FILE_PATH:
        file_path_str = str(PROCESSED_FILE_PATH)
        # Split long paths across multiple lines if needed
        if len(file_path_str) > 90:
            logger.info(f"║  PROCESSED EXCEL FILE:" + " " * 74 + "║")
            logger.info(f"║  {file_path_str[:90]}" + " " * (96 - min(90, len(file_path_str))) + "║")
            if len(file_path_str) > 90:
                remaining = file_path_str[90:]
                logger.info(f"║  {remaining}" + " " * (96 - len(remaining)) + "║")
        else:
            logger.info(f"║  PROCESSED EXCEL FILE:" + " " * 74 + "║")
            logger.info(f"║  {file_path_str}" + " " * (96 - len(file_path_str)) + "║")

        logger.info("║" + " " * 98 + "║")
        logger.info("║  CONTENTS OF PROCESSED FILE:" + " " * 68 + "║")
        logger.info("║  ┌────────────────────────────────────────────────────────────────────────────────────────────┐  ║")
        logger.info("║  │  Sheet 1: Summary                                                                          │  ║")
        logger.info("║  │    • Section 1: Price Data (all rows, 7 decimal places)                                    │  ║")
        logger.info("║  │    • Section 2: LN Returns (all rows, 7 decimal places)                                    │  ║")
        logger.info("║  │    • Section 3: Statistics (Mean, Std Dev, Variance)                                       │  ║")
        logger.info("║  │    • Section 4: Covariance Matrix (full NxN)                                               │  ║")
        logger.info("║  │    • Section 5: Correlation Matrix (full NxN)                                              │  ║")
        logger.info("║  │                                                                                            │  ║")
        logger.info("║  │  Sheet 2: LN_Returns (detailed log returns sheet)                                          │  ║")
        logger.info("║  │                                                                                            │  ║")
        logger.info("║  │  Original Sheets: Prices, DJReturns, DJFrontier, SP Hist                                   │  ║")
        logger.info("║  └────────────────────────────────────────────────────────────────────────────────────────────┘  ║")
    else:
        logger.info("║  No processed file was created." + " " * 65 + "║")

    logger.info("║" + " " * 98 + "║")
    logger.info("╚" + "═" * 98 + "╝")
    logger.info("")
    logger.info("╔" + "═" * 98 + "╗")
    logger.info("║" + " " * 35 + "PROCESSING COMPLETE" + " " * 44 + "║")
    logger.info("╚" + "═" * 98 + "╝")


if __name__ == "__main__":
    main()
