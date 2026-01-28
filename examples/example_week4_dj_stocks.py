"""
Week 4 Exam Prep - Complete Portfolio Analysis
===============================================
Analyzes 30 DJ stocks + SPY from W4E1PrepData.xlsx
RF = 0.03% monthly
"""

import numpy as np
import pandas as pd
from scipy.optimize import minimize
from pathlib import Path
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Paths
DATA_FILE = r"F:\iCloudDrive\UoU school\SPRING 2026\Week4\W4E1PrepData.xlsx"
OUTPUT_DIR = r"F:\iCloudDrive\UoU school\SPRING 2026\Week4\solution"

# Parameters
RF_RATE = 0.0003  # 0.03% monthly

# ============================================================================
# LOAD DATA
# ============================================================================
print("=" * 70)
print("WEEK 4 EXAM PREP - PORTFOLIO OPTIMIZATION")
print("=" * 70)

print("\n--- Loading Data ---")
# Read with header in row 1 (0-indexed), skipping the first row
df = pd.read_excel(DATA_FILE, header=0)

# Check if first row contains stock symbols (not numeric)
first_row = df.iloc[0]
if not pd.api.types.is_numeric_dtype(first_row.iloc[1]):
    # First row is actually the header
    new_headers = df.iloc[0].values
    df = df[1:]  # Remove first row
    df.columns = new_headers
    df = df.reset_index(drop=True)

print(f"Loaded {len(df)} rows, {len(df.columns)} columns")

# Remove date column
date_col = df.columns[0]
df_prices = df.drop(columns=[date_col])

# Convert to numeric
df_prices = df_prices.apply(pd.to_numeric, errors='coerce')

# Get asset names
asset_names = list(df_prices.columns)
print(f"Assets: {asset_names}")

# Check if prices or returns (prices are usually > 10)
sample_mean = df_prices.mean().mean()
print(f"Sample mean: {sample_mean:.4f}")

if abs(sample_mean) > 2:
    print("Detected PRICE data - converting to log returns")
    returns_df = np.log(df_prices / df_prices.shift(1)).dropna()
else:
    print("Detected RETURN data")
    returns_df = df_prices.dropna()

print(f"Returns shape: {returns_df.shape}")

# Separate SPY for benchmarking
spy_returns = None
if 'SPY' in returns_df.columns:
    spy_returns = returns_df['SPY'].values
    spy_mean = spy_returns.mean()
    spy_std = spy_returns.std()
    print(f"\nSPY: Mean={spy_mean*100:.4f}%, Std={spy_std*100:.4f}%")
    # Remove SPY for optimization (only use 30 DJ stocks)
    returns_df_opt = returns_df.drop(columns=['SPY'])
else:
    returns_df_opt = returns_df

asset_names_opt = list(returns_df_opt.columns)
n_assets = len(asset_names_opt)
print(f"\nUsing {n_assets} assets for optimization")

# Compute statistics
expected_returns = returns_df_opt.mean().values
cov_matrix = returns_df_opt.cov().values * (len(returns_df_opt) - 1) / len(returns_df_opt)  # Population cov

print("\n--- Asset Statistics ---")
print(f"{'Asset':<8} {'Mean':>12} {'Std Dev':>12}")
print("-" * 34)
for i, name in enumerate(asset_names_opt):
    mean_ret = expected_returns[i] * 100
    std_dev = np.sqrt(cov_matrix[i,i]) * 100
    print(f"{name:<8} {mean_ret:>11.4f}% {std_dev:>11.4f}%")

# ============================================================================
# PORTFOLIO FUNCTIONS
# ============================================================================
def portfolio_return(w):
    return np.dot(w, expected_returns)

def portfolio_variance(w):
    return np.dot(w, np.dot(cov_matrix, w))

def portfolio_std(w):
    return np.sqrt(portfolio_variance(w))

def portfolio_sharpe(w):
    ret = portfolio_return(w)
    std = portfolio_std(w)
    return (ret - RF_RATE) / std if std > 1e-10 else 0

# ============================================================================
# OPTIMIZATION FUNCTIONS
# ============================================================================
def minimum_variance_portfolio(allow_short=True):
    w0 = np.ones(n_assets) / n_assets
    constraints = [{'type': 'eq', 'fun': lambda w: np.sum(w) - 1}]
    bounds = None if allow_short else [(0, 1) for _ in range(n_assets)]

    result = minimize(portfolio_variance, w0, method='SLSQP',
                     bounds=bounds, constraints=constraints,
                     options={'ftol': 1e-12, 'maxiter': 1000})
    return result.x, result.success

def tangent_portfolio(allow_short=True):
    w0 = np.ones(n_assets) / n_assets
    constraints = [{'type': 'eq', 'fun': lambda w: np.sum(w) - 1}]
    # Add reasonable bounds to prevent extreme positions
    if allow_short:
        bounds = [(-2, 3) for _ in range(n_assets)]  # Allow shorting up to 200%, long up to 300%
    else:
        bounds = [(0, 1) for _ in range(n_assets)]

    def neg_sharpe(w):
        std = portfolio_std(w)
        if std < 1e-10:
            return 1e10
        return -(portfolio_return(w) - RF_RATE) / std

    result = minimize(neg_sharpe, w0, method='SLSQP',
                     bounds=bounds, constraints=constraints,
                     options={'ftol': 1e-12, 'maxiter': 1000})
    return result.x, result.success

def optimize_for_target_std(target_std, allow_short=True):
    """Maximize return for a given target standard deviation."""
    w0 = np.ones(n_assets) / n_assets

    bounds = None if allow_short else [(0, 1) for _ in range(n_assets)]

    # Use inequality constraint for std (less than or equal)
    constraints = [
        {'type': 'eq', 'fun': lambda w: np.sum(w) - 1},
    ]

    # For constrained optimization, use inequality
    if not allow_short:
        constraints.append({'type': 'ineq', 'fun': lambda w: target_std - portfolio_std(w)})
    else:
        constraints.append({'type': 'eq', 'fun': lambda w: portfolio_std(w) - target_std})

    result = minimize(lambda w: -portfolio_return(w), w0, method='SLSQP',
                     bounds=bounds, constraints=constraints,
                     options={'ftol': 1e-12, 'maxiter': 1000})

    return result.x, result.success

def two_fund_frontier(w1, w2, n_points=300, lambda_range=(-0.5, 2.0)):
    """Generate efficient frontier using two-fund separation."""
    mu1, sigma1 = portfolio_return(w1), portfolio_std(w1)
    mu2, sigma2 = portfolio_return(w2), portfolio_std(w2)
    cov12 = np.dot(w1, np.dot(cov_matrix, w2))

    lambdas = np.linspace(lambda_range[0], lambda_range[1], n_points)
    returns, stds = [], []

    for lam in lambdas:
        mu_p = lam * mu1 + (1 - lam) * mu2
        var_p = lam**2 * sigma1**2 + (1-lam)**2 * sigma2**2 + 2*lam*(1-lam)*cov12
        if var_p >= 0:
            returns.append(mu_p)
            stds.append(np.sqrt(var_p))

    return np.array(returns), np.array(stds)

# ============================================================================
# RUN ANALYSIS
# ============================================================================
results = {}

# 1. Minimum Variance Portfolio
print("\n" + "=" * 70)
print("MINIMUM VARIANCE PORTFOLIO (MVP)")
print("=" * 70)
mvp_w, mvp_success = minimum_variance_portfolio()
mvp_ret = portfolio_return(mvp_w)
mvp_std = portfolio_std(mvp_w)
mvp_sharpe = portfolio_sharpe(mvp_w)
results['mvp'] = {'weights': mvp_w, 'return': mvp_ret, 'std': mvp_std, 'sharpe': mvp_sharpe}
print(f"Return: {mvp_ret*100:.4f}%")
print(f"Std Dev: {mvp_std*100:.4f}%")
print(f"Sharpe: {mvp_sharpe:.4f}")

# 2. Tangent Portfolio
print("\n" + "=" * 70)
print("TANGENT PORTFOLIO (Maximum Sharpe)")
print("=" * 70)
tan_w, tan_success = tangent_portfolio()
tan_ret = portfolio_return(tan_w)
tan_std = portfolio_std(tan_w)
tan_sharpe = portfolio_sharpe(tan_w)
results['tangent'] = {'weights': tan_w, 'return': tan_ret, 'std': tan_std, 'sharpe': tan_sharpe}
print(f"Return: {tan_ret*100:.4f}%")
print(f"Std Dev: {tan_std*100:.4f}%")
print(f"Sharpe: {tan_sharpe:.4f}")

# Show tangent weights (non-zero only)
print("\nTangent Portfolio Weights (top holdings):")
sorted_idx = np.argsort(np.abs(tan_w))[::-1]
for idx in sorted_idx[:10]:
    if abs(tan_w[idx]) > 0.01:
        print(f"  {asset_names_opt[idx]}: {tan_w[idx]*100:>8.2f}%")

# 3. Efficient Portfolio at 4% Std
print("\n" + "=" * 70)
print("EFFICIENT PORTFOLIO AT 4% STD DEV")
print("=" * 70)
eff4_w, eff4_success = optimize_for_target_std(0.04, allow_short=True)
eff4_ret = portfolio_return(eff4_w)
eff4_std = portfolio_std(eff4_w)
eff4_sharpe = portfolio_sharpe(eff4_w)
results['eff_4pct'] = {'weights': eff4_w, 'return': eff4_ret, 'std': eff4_std, 'sharpe': eff4_sharpe}
print(f"Return: {eff4_ret*100:.4f}%")
print(f"Std Dev: {eff4_std*100:.4f}%")
print(f"Sharpe: {eff4_sharpe:.4f}")
print(f"Optimization success: {eff4_success}")

# 4. Efficient Portfolio at 7% Std
print("\n" + "=" * 70)
print("EFFICIENT PORTFOLIO AT 7% STD DEV")
print("=" * 70)
eff7_w, eff7_success = optimize_for_target_std(0.07, allow_short=True)
eff7_ret = portfolio_return(eff7_w)
eff7_std = portfolio_std(eff7_w)
eff7_sharpe = portfolio_sharpe(eff7_w)
results['eff_7pct'] = {'weights': eff7_w, 'return': eff7_ret, 'std': eff7_std, 'sharpe': eff7_sharpe}
print(f"Return: {eff7_ret*100:.4f}%")
print(f"Std Dev: {eff7_std*100:.4f}%")
print(f"Sharpe: {eff7_sharpe:.4f}")
print(f"Optimization success: {eff7_success}")

# 5. Constrained Portfolios (No Short Selling)
print("\n" + "=" * 70)
print("CONSTRAINED PORTFOLIOS (NO SHORT SELLING)")
print("=" * 70)

# Pension Portfolio at 5%
print("\n--- Pension Portfolio at 5% Std (No Short) ---")
pen5_w, pen5_success = optimize_for_target_std(0.05, allow_short=False)
pen5_ret = portfolio_return(pen5_w)
pen5_std = portfolio_std(pen5_w)
pen5_sharpe = portfolio_sharpe(pen5_w)
results['pension_5pct'] = {'weights': pen5_w, 'return': pen5_ret, 'std': pen5_std, 'sharpe': pen5_sharpe}
print(f"Return: {pen5_ret*100:.4f}%")
print(f"Std Dev: {pen5_std*100:.4f}%")
print(f"Sharpe: {pen5_sharpe:.4f}")

# Pension Portfolio at 6%
print("\n--- Pension Portfolio at 6% Std (No Short) ---")
pen6_w, pen6_success = optimize_for_target_std(0.06, allow_short=False)
pen6_ret = portfolio_return(pen6_w)
pen6_std = portfolio_std(pen6_w)
pen6_sharpe = portfolio_sharpe(pen6_w)
results['pension_6pct'] = {'weights': pen6_w, 'return': pen6_ret, 'std': pen6_std, 'sharpe': pen6_sharpe}
print(f"Return: {pen6_ret*100:.4f}%")
print(f"Std Dev: {pen6_std*100:.4f}%")
print(f"Sharpe: {pen6_sharpe:.4f}")

# 6. Equal-Weighted Portfolio
print("\n" + "=" * 70)
print("EQUAL-WEIGHTED PORTFOLIO")
print("=" * 70)
eq_w = np.ones(n_assets) / n_assets
eq_ret = portfolio_return(eq_w)
eq_std = portfolio_std(eq_w)
eq_sharpe = portfolio_sharpe(eq_w)
results['equal_weight'] = {'weights': eq_w, 'return': eq_ret, 'std': eq_std, 'sharpe': eq_sharpe}
print(f"Return: {eq_ret*100:.4f}%")
print(f"Std Dev: {eq_std*100:.4f}%")
print(f"Sharpe: {eq_sharpe:.4f}")

# 7. SPY Benchmark
if spy_returns is not None:
    print("\n" + "=" * 70)
    print("SPY BENCHMARK")
    print("=" * 70)
    spy_mean = spy_returns.mean()
    spy_std_calc = spy_returns.std() * np.sqrt(len(spy_returns)-1) / np.sqrt(len(spy_returns))  # Pop std
    spy_sharpe_calc = (spy_mean - RF_RATE) / spy_std_calc
    results['spy'] = {'return': spy_mean, 'std': spy_std_calc, 'sharpe': spy_sharpe_calc}
    print(f"Return: {spy_mean*100:.4f}%")
    print(f"Std Dev: {spy_std_calc*100:.4f}%")
    print(f"Sharpe: {spy_sharpe_calc:.4f}")

# 8. DisavowelInvestor Portfolio (exclude AAPL, AMGN, AXP, IBM, INTC)
print("\n" + "=" * 70)
print("DISAVOWELINVESTOR PORTFOLIO (Exclude A-stocks)")
print("=" * 70)
exclude_stocks = ['AAPL', 'AMGN', 'AXP', 'IBM', 'INTC']
disavowel_idx = [i for i, name in enumerate(asset_names_opt) if name not in exclude_stocks]
disavowel_names = [asset_names_opt[i] for i in disavowel_idx]
print(f"Excluding: {exclude_stocks}")
print(f"Remaining: {len(disavowel_names)} stocks")

# Create reduced matrices
exp_ret_disavowel = expected_returns[disavowel_idx]
cov_disavowel = cov_matrix[np.ix_(disavowel_idx, disavowel_idx)]
n_disavowel = len(disavowel_idx)

# Functions for disavowel
def disavowel_return(w):
    return np.dot(w, exp_ret_disavowel)

def disavowel_variance(w):
    return np.dot(w, np.dot(cov_disavowel, w))

def disavowel_std(w):
    return np.sqrt(disavowel_variance(w))

def disavowel_sharpe(w):
    ret = disavowel_return(w)
    std = disavowel_std(w)
    return (ret - RF_RATE) / std if std > 1e-10 else 0

# MVP for disavowel
w0 = np.ones(n_disavowel) / n_disavowel
constraints = [{'type': 'eq', 'fun': lambda w: np.sum(w) - 1}]
result = minimize(disavowel_variance, w0, method='SLSQP',
                 constraints=constraints, options={'ftol': 1e-12})
disavowel_mvp_w = result.x
disavowel_mvp_ret = disavowel_return(disavowel_mvp_w)
disavowel_mvp_std = disavowel_std(disavowel_mvp_w)
print(f"DisavowelInvestor MVP:")
print(f"  Return: {disavowel_mvp_ret*100:.4f}%")
print(f"  Std Dev: {disavowel_mvp_std*100:.4f}%")

# Tangent for disavowel (with bounds to prevent extreme positions)
def neg_sharpe_disavowel(w):
    std = disavowel_std(w)
    if std < 1e-10:
        return 1e10
    return -(disavowel_return(w) - RF_RATE) / std

bounds_disavowel = [(-2, 3) for _ in range(n_disavowel)]
result = minimize(neg_sharpe_disavowel, w0, method='SLSQP',
                 bounds=bounds_disavowel, constraints=constraints, options={'ftol': 1e-12})
disavowel_tan_w = result.x
disavowel_tan_ret = disavowel_return(disavowel_tan_w)
disavowel_tan_std = disavowel_std(disavowel_tan_w)
disavowel_tan_sharpe = disavowel_sharpe(disavowel_tan_w)
print(f"DisavowelInvestor Tangent:")
print(f"  Return: {disavowel_tan_ret*100:.4f}%")
print(f"  Std Dev: {disavowel_tan_std*100:.4f}%")
print(f"  Sharpe: {disavowel_tan_sharpe:.4f}")

results['disavowel'] = {
    'mvp': {'return': disavowel_mvp_ret, 'std': disavowel_mvp_std, 'weights': disavowel_mvp_w},
    'tangent': {'return': disavowel_tan_ret, 'std': disavowel_tan_std, 'sharpe': disavowel_tan_sharpe, 'weights': disavowel_tan_w}
}

# ============================================================================
# CREATE OUTPUT DIRECTORY
# ============================================================================
output_path = Path(OUTPUT_DIR)
output_path.mkdir(parents=True, exist_ok=True)

# ============================================================================
# GENERATE EFFICIENT FRONTIER GRAPH
# ============================================================================
print("\n--- Generating Graph ---")
import matplotlib.pyplot as plt

fig, ax = plt.subplots(figsize=(16, 12))

# Generate efficient frontier using two-fund separation
frontier_ret, frontier_std = two_fund_frontier(mvp_w, tan_w)
ax.plot(frontier_std * 100, frontier_ret * 100, 'b-', linewidth=2.5,
       label='Efficient Frontier (Unconstrained)', zorder=2)

# Capital Market Line
cml_weights = np.linspace(0, 2.5, 100)
cml_returns = cml_weights * tan_ret + (1 - cml_weights) * RF_RATE
cml_stds = cml_weights * tan_std
ax.plot(cml_stds * 100, cml_returns * 100, 'g--', linewidth=2.5,
       label='Capital Market Line', zorder=2)

# Individual assets (small dots, labeled)
colors = plt.cm.tab20(np.linspace(0, 1, n_assets))
for i, name in enumerate(asset_names_opt):
    std = np.sqrt(cov_matrix[i, i]) * 100
    ret = expected_returns[i] * 100
    ax.scatter(std, ret, s=80, c=[colors[i]], edgecolors='black', linewidths=0.5, zorder=3, alpha=0.7)
    ax.annotate(name, (std, ret), textcoords='offset points', xytext=(4, 2), fontsize=7)

# Risk-free rate
ax.scatter(0, RF_RATE * 100, s=200, c='gold', edgecolors='black', linewidths=2, marker='*', zorder=6,
          label=f'Risk-Free ({RF_RATE*100:.2f}%)')

# MVP
ax.scatter(mvp_std * 100, mvp_ret * 100, s=250, c='red', edgecolors='black', linewidths=2,
          marker='s', zorder=6)
ax.annotate('MVP', (mvp_std * 100, mvp_ret * 100), textcoords='offset points', xytext=(10, -5),
           fontsize=12, fontweight='bold', color='red')

# Tangent Portfolio
ax.scatter(tan_std * 100, tan_ret * 100, s=250, c='green', edgecolors='black', linewidths=2,
          marker='^', zorder=6)
ax.annotate('Tangent\n(Max Sharpe)', (tan_std * 100, tan_ret * 100), textcoords='offset points',
           xytext=(10, 5), fontsize=10, fontweight='bold', color='green')

# Efficient portfolios (4% and 7%)
ax.scatter(eff4_std * 100, eff4_ret * 100, s=200, c='purple', edgecolors='black', linewidths=2,
          marker='D', zorder=6)
ax.annotate('Eff(4%)', (eff4_std * 100, eff4_ret * 100), textcoords='offset points',
           xytext=(10, -10), fontsize=10, fontweight='bold', color='purple')

ax.scatter(eff7_std * 100, eff7_ret * 100, s=200, c='orange', edgecolors='black', linewidths=2,
          marker='D', zorder=6)
ax.annotate('Eff(7%)', (eff7_std * 100, eff7_ret * 100), textcoords='offset points',
           xytext=(10, -10), fontsize=10, fontweight='bold', color='orange')

# Constrained portfolios
ax.scatter(pen5_std * 100, pen5_ret * 100, s=180, c='cyan', edgecolors='black', linewidths=2,
          marker='o', zorder=6)
ax.annotate('Pension(5%)\nNo Short', (pen5_std * 100, pen5_ret * 100), textcoords='offset points',
           xytext=(10, 5), fontsize=9, fontweight='bold', color='darkcyan')

ax.scatter(pen6_std * 100, pen6_ret * 100, s=180, c='magenta', edgecolors='black', linewidths=2,
          marker='o', zorder=6)
ax.annotate('Pension(6%)\nNo Short', (pen6_std * 100, pen6_ret * 100), textcoords='offset points',
           xytext=(10, -15), fontsize=9, fontweight='bold', color='darkmagenta')

# Equal-weighted portfolio
ax.scatter(eq_std * 100, eq_ret * 100, s=180, c='brown', edgecolors='black', linewidths=2,
          marker='p', zorder=6)
ax.annotate('Equal\nWeight', (eq_std * 100, eq_ret * 100), textcoords='offset points',
           xytext=(10, 0), fontsize=9, fontweight='bold', color='brown')

# SPY
if spy_returns is not None:
    ax.scatter(spy_std_calc * 100, spy_mean * 100, s=220, c='black', edgecolors='yellow', linewidths=2,
              marker='X', zorder=6)
    ax.annotate('SPY', (spy_std_calc * 100, spy_mean * 100), textcoords='offset points',
               xytext=(10, -5), fontsize=11, fontweight='bold', color='black')

ax.set_xlabel('Standard Deviation (%)', fontsize=14, fontweight='bold')
ax.set_ylabel('Expected Return (%)', fontsize=14, fontweight='bold')
ax.set_title('Week 4 Exam Prep: Efficient Frontier\n30 Dow Jones Stocks | RF = 0.03%',
            fontsize=16, fontweight='bold', pad=20)
ax.grid(True, alpha=0.3, linestyle='--')
ax.legend(loc='upper left', fontsize=10)
ax.set_xlim(0, max(15, tan_std*100 + 2))

plt.tight_layout()
graph_path = output_path / "W4_efficient_frontier.png"
plt.savefig(graph_path, dpi=150, bbox_inches='tight', facecolor='white')
print(f"Graph saved to: {graph_path}")
plt.close()

# ============================================================================
# GENERATE PYTHON SOLUTION FILE
# ============================================================================
print("\n--- Generating Python Solution ---")

python_code = f'''"""
================================================================================
WEEK 4 EXAM PREP - PORTFOLIO OPTIMIZATION SOLUTION
================================================================================
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Data: W4E1PrepData.xlsx (30 Dow Jones stocks + SPY)
Risk-Free Rate: {RF_RATE*100:.2f}% per month
================================================================================
"""

import numpy as np
from scipy.optimize import minimize

# ============================================================================
# INPUT DATA
# ============================================================================

# Asset names
asset_names = {asset_names_opt}

# Expected returns (monthly)
expected_returns = np.array({list(expected_returns)})

# Covariance matrix (population covariance)
cov_matrix = np.array({cov_matrix.tolist()})

# Risk-free rate
rf_rate = {RF_RATE}  # {RF_RATE*100:.2f}% monthly

n_assets = len(asset_names)

# ============================================================================
# PORTFOLIO FUNCTIONS
# ============================================================================

def portfolio_return(w):
    """Portfolio return: mu_p = w' * mu"""
    return np.dot(w, expected_returns)

def portfolio_variance(w):
    """Portfolio variance: var_p = w' * Sigma * w"""
    return np.dot(w, np.dot(cov_matrix, w))

def portfolio_std(w):
    """Portfolio standard deviation"""
    return np.sqrt(portfolio_variance(w))

def portfolio_sharpe(w):
    """Sharpe ratio: (mu_p - RF) / sigma_p"""
    ret = portfolio_return(w)
    std = portfolio_std(w)
    return (ret - rf_rate) / std if std > 1e-10 else 0

# ============================================================================
# OPTIMIZATION FUNCTIONS
# ============================================================================

def minimum_variance_portfolio(allow_short=True):
    """Find the Minimum Variance Portfolio"""
    w0 = np.ones(n_assets) / n_assets
    constraints = [{{'type': 'eq', 'fun': lambda w: np.sum(w) - 1}}]
    bounds = None if allow_short else [(0, 1) for _ in range(n_assets)]

    result = minimize(portfolio_variance, w0, method='SLSQP',
                     bounds=bounds, constraints=constraints,
                     options={{'ftol': 1e-12}})
    return result.x

def tangent_portfolio(allow_short=True):
    """Find the Tangent Portfolio (Maximum Sharpe Ratio)"""
    w0 = np.ones(n_assets) / n_assets
    constraints = [{{'type': 'eq', 'fun': lambda w: np.sum(w) - 1}}]
    bounds = None if allow_short else [(0, 1) for _ in range(n_assets)]

    def neg_sharpe(w):
        std = portfolio_std(w)
        if std < 1e-10:
            return 1e10
        return -(portfolio_return(w) - rf_rate) / std

    result = minimize(neg_sharpe, w0, method='SLSQP',
                     bounds=bounds, constraints=constraints,
                     options={{'ftol': 1e-12}})
    return result.x

def optimize_for_target_std(target_std, allow_short=True):
    """Find efficient portfolio at target standard deviation"""
    w0 = np.ones(n_assets) / n_assets
    bounds = None if allow_short else [(0, 1) for _ in range(n_assets)]

    if allow_short:
        constraints = [
            {{'type': 'eq', 'fun': lambda w: np.sum(w) - 1}},
            {{'type': 'eq', 'fun': lambda w: portfolio_std(w) - target_std}}
        ]
    else:
        constraints = [
            {{'type': 'eq', 'fun': lambda w: np.sum(w) - 1}},
            {{'type': 'ineq', 'fun': lambda w: target_std - portfolio_std(w)}}
        ]

    result = minimize(lambda w: -portfolio_return(w), w0, method='SLSQP',
                     bounds=bounds, constraints=constraints,
                     options={{'ftol': 1e-12}})
    return result.x if result.success else None

# ============================================================================
# RESULTS
# ============================================================================

if __name__ == "__main__":
    print("=" * 70)
    print("WEEK 4 EXAM PREP - PORTFOLIO OPTIMIZATION SOLUTION")
    print("=" * 70)

    # MVP
    print("\\n--- Minimum Variance Portfolio ---")
    mvp_w = minimum_variance_portfolio()
    print(f"Return: {{portfolio_return(mvp_w)*100:.4f}}%")
    print(f"Std Dev: {{portfolio_std(mvp_w)*100:.4f}}%")
    print(f"Sharpe: {{portfolio_sharpe(mvp_w):.4f}}")

    # Tangent
    print("\\n--- Tangent Portfolio ---")
    tan_w = tangent_portfolio()
    print(f"Return: {{portfolio_return(tan_w)*100:.4f}}%")
    print(f"Std Dev: {{portfolio_std(tan_w)*100:.4f}}%")
    print(f"Sharpe: {{portfolio_sharpe(tan_w):.4f}}")

    # Efficient at 4%
    print("\\n--- Efficient Portfolio at 4% Std ---")
    eff4_w = optimize_for_target_std(0.04)
    if eff4_w is not None:
        print(f"Return: {{portfolio_return(eff4_w)*100:.4f}}%")
        print(f"Std Dev: {{portfolio_std(eff4_w)*100:.4f}}%")

    # Efficient at 7%
    print("\\n--- Efficient Portfolio at 7% Std ---")
    eff7_w = optimize_for_target_std(0.07)
    if eff7_w is not None:
        print(f"Return: {{portfolio_return(eff7_w)*100:.4f}}%")
        print(f"Std Dev: {{portfolio_std(eff7_w)*100:.4f}}%")

    # Pension 5% (no short)
    print("\\n--- Pension Portfolio at 5% (No Short) ---")
    pen5_w = optimize_for_target_std(0.05, allow_short=False)
    if pen5_w is not None:
        print(f"Return: {{portfolio_return(pen5_w)*100:.4f}}%")
        print(f"Std Dev: {{portfolio_std(pen5_w)*100:.4f}}%")

    # Pension 6% (no short)
    print("\\n--- Pension Portfolio at 6% (No Short) ---")
    pen6_w = optimize_for_target_std(0.06, allow_short=False)
    if pen6_w is not None:
        print(f"Return: {{portfolio_return(pen6_w)*100:.4f}}%")
        print(f"Std Dev: {{portfolio_std(pen6_w)*100:.4f}}%")

    # Equal-weighted
    print("\\n--- Equal-Weighted Portfolio ---")
    eq_w = np.ones(n_assets) / n_assets
    print(f"Return: {{portfolio_return(eq_w)*100:.4f}}%")
    print(f"Std Dev: {{portfolio_std(eq_w)*100:.4f}}%")
    print(f"Sharpe: {{portfolio_sharpe(eq_w):.4f}}")
'''

py_path = output_path / "W4_solution.py"
with open(py_path, 'w') as f:
    f.write(python_code)
print(f"Python solution saved to: {py_path}")

# ============================================================================
# GENERATE EXCEL FILE
# ============================================================================
print("\n--- Generating Excel Solution ---")
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ==================== SHEET 1: INPUT DATA ====================
    ws1 = wb.active
    ws1.title = "Input Data"

    ws1['A1'] = "WEEK 4 EXAM PREP - INPUT DATA"
    ws1['A1'].font = Font(bold=True, size=14)

    ws1['A3'] = f"Risk-Free Rate: {RF_RATE*100:.2f}% monthly"
    ws1['A4'] = f"Number of Assets: {n_assets}"

    # Asset statistics
    ws1['A6'] = "ASSET STATISTICS"
    ws1['A6'].font = header_font

    headers = ['Asset', 'Mean Return', 'Std Dev', 'Variance']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=7, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for i, name in enumerate(asset_names_opt):
        row = i + 8
        ws1.cell(row=row, column=1, value=name).border = thin_border
        ws1.cell(row=row, column=2, value=expected_returns[i]).number_format = '0.0000%'
        ws1.cell(row=row, column=2).border = thin_border
        ws1.cell(row=row, column=3, value=np.sqrt(cov_matrix[i,i])).number_format = '0.0000%'
        ws1.cell(row=row, column=3).border = thin_border
        ws1.cell(row=row, column=4, value=cov_matrix[i,i]).number_format = '0.000000'
        ws1.cell(row=row, column=4).border = thin_border

    # ==================== SHEET 2: PORTFOLIOS ====================
    ws2 = wb.create_sheet("Portfolio Results")

    ws2['A1'] = "PORTFOLIO OPTIMIZATION RESULTS"
    ws2['A1'].font = Font(bold=True, size=14)

    # Summary table
    headers = ['Portfolio', 'Return', 'Std Dev', 'Sharpe']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    portfolios = [
        ('MVP', mvp_ret, mvp_std, mvp_sharpe),
        ('Tangent (Max Sharpe)', tan_ret, tan_std, tan_sharpe),
        ('Efficient (4%)', eff4_ret, eff4_std, eff4_sharpe),
        ('Efficient (7%)', eff7_ret, eff7_std, eff7_sharpe),
        ('Pension (5%, No Short)', pen5_ret, pen5_std, pen5_sharpe),
        ('Pension (6%, No Short)', pen6_ret, pen6_std, pen6_sharpe),
        ('Equal-Weighted', eq_ret, eq_std, eq_sharpe),
    ]

    if spy_returns is not None:
        portfolios.append(('SPY Benchmark', spy_mean, spy_std_calc, spy_sharpe_calc))

    for i, (name, ret, std, sharpe) in enumerate(portfolios):
        row = i + 4
        ws2.cell(row=row, column=1, value=name).border = thin_border
        ws2.cell(row=row, column=2, value=ret).number_format = '0.0000%'
        ws2.cell(row=row, column=2).border = thin_border
        ws2.cell(row=row, column=3, value=std).number_format = '0.0000%'
        ws2.cell(row=row, column=3).border = thin_border
        ws2.cell(row=row, column=4, value=sharpe).number_format = '0.0000'
        ws2.cell(row=row, column=4).border = thin_border

    # ==================== SHEET 3: WEIGHTS ====================
    ws3 = wb.create_sheet("Portfolio Weights")

    ws3['A1'] = "PORTFOLIO WEIGHTS"
    ws3['A1'].font = Font(bold=True, size=14)

    # Headers
    weight_headers = ['Asset', 'MVP', 'Tangent', 'Eff(4%)', 'Eff(7%)', 'Pension(5%)', 'Pension(6%)', 'Equal']
    for col, header in enumerate(weight_headers, 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for i, name in enumerate(asset_names_opt):
        row = i + 4
        ws3.cell(row=row, column=1, value=name).border = thin_border
        ws3.cell(row=row, column=2, value=mvp_w[i]).number_format = '0.00%'
        ws3.cell(row=row, column=2).border = thin_border
        ws3.cell(row=row, column=3, value=tan_w[i]).number_format = '0.00%'
        ws3.cell(row=row, column=3).border = thin_border
        ws3.cell(row=row, column=4, value=eff4_w[i]).number_format = '0.00%'
        ws3.cell(row=row, column=4).border = thin_border
        ws3.cell(row=row, column=5, value=eff7_w[i]).number_format = '0.00%'
        ws3.cell(row=row, column=5).border = thin_border
        ws3.cell(row=row, column=6, value=pen5_w[i]).number_format = '0.00%'
        ws3.cell(row=row, column=6).border = thin_border
        ws3.cell(row=row, column=7, value=pen6_w[i]).number_format = '0.00%'
        ws3.cell(row=row, column=7).border = thin_border
        ws3.cell(row=row, column=8, value=eq_w[i]).number_format = '0.00%'
        ws3.cell(row=row, column=8).border = thin_border

    # Sum row
    sum_row = n_assets + 4
    ws3.cell(row=sum_row, column=1, value="SUM").font = header_font
    for col in range(2, 9):
        ws3.cell(row=sum_row, column=col, value=f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{sum_row-1})")
        ws3.cell(row=sum_row, column=col).number_format = '0.00%'

    # Adjust column widths
    for ws in wb.worksheets:
        for col in range(1, 15):
            ws.column_dimensions[get_column_letter(col)].width = 15

    excel_path = output_path / "W4_solution.xlsx"
    wb.save(excel_path)
    print(f"Excel solution saved to: {excel_path}")

except ImportError as e:
    print(f"Could not generate Excel file: {e}")
    print("Install with: pip install openpyxl")

# ============================================================================
# FINAL SUMMARY
# ============================================================================
print("\n" + "=" * 70)
print("FINAL RESULTS SUMMARY")
print("=" * 70)

print(f"\nMinimum Variance Portfolio (MVP):")
print(f"  Return: {mvp_ret*100:.4f}%")
print(f"  Std Dev: {mvp_std*100:.4f}%")
print(f"  Sharpe: {mvp_sharpe:.4f}")

print(f"\nTangent Portfolio (Max Sharpe):")
print(f"  Return: {tan_ret*100:.4f}%")
print(f"  Std Dev: {tan_std*100:.4f}%")
print(f"  Sharpe: {tan_sharpe:.4f}")

print(f"\nEfficient Portfolio at 4% Std:")
print(f"  Return: {eff4_ret*100:.4f}%")
print(f"  Std Dev: {eff4_std*100:.4f}%")

print(f"\nEfficient Portfolio at 7% Std:")
print(f"  Return: {eff7_ret*100:.4f}%")
print(f"  Std Dev: {eff7_std*100:.4f}%")

print(f"\nPension Portfolio at 5% (No Short):")
print(f"  Return: {pen5_ret*100:.4f}%")
print(f"  Std Dev: {pen5_std*100:.4f}%")

print(f"\nPension Portfolio at 6% (No Short):")
print(f"  Return: {pen6_ret*100:.4f}%")
print(f"  Std Dev: {pen6_std*100:.4f}%")

print(f"\nEqual-Weighted Portfolio:")
print(f"  Return: {eq_ret*100:.4f}%")
print(f"  Std Dev: {eq_std*100:.4f}%")
print(f"  Sharpe: {eq_sharpe:.4f}")

if spy_returns is not None:
    print(f"\nSPY Benchmark:")
    print(f"  Return: {spy_mean*100:.4f}%")
    print(f"  Std Dev: {spy_std_calc*100:.4f}%")
    print(f"  Sharpe: {spy_sharpe_calc:.4f}")

print(f"\nDisavowelInvestor (excl. AAPL, AMGN, AXP, IBM, INTC):")
print(f"  MVP Std: {disavowel_mvp_std*100:.4f}%")
print(f"  Tangent Sharpe: {disavowel_tan_sharpe:.4f}")

print("\n" + "=" * 70)
print(f"All outputs saved to: {OUTPUT_DIR}")
print("=" * 70)
