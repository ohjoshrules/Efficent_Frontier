"""
================================================================================
HOMEWORK ASSISTANT - Portfolio Analysis Automation
================================================================================

This script:
1. Reads homework instructions from a PDF file
2. Loads data from Excel files in the same folder
3. Generates a Python solution file
4. Creates an Excel file with step-by-step formulas
5. Produces all required outputs (graphs, reports)

Default Location: F:/iCloudDrive/UoU school/SPRING 2026/Week4

Usage:
    python homework_assistant.py

Then follow prompts or press Enter to use defaults.

================================================================================
"""

import os
import sys
import numpy as np
import pandas as pd
from scipy.optimize import minimize
from pathlib import Path
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# Try to import PDF reading library
try:
    import fitz  # PyMuPDF
    PDF_SUPPORT = True
except ImportError:
    try:
        from PyPDF2 import PdfReader
        PDF_SUPPORT = True
        PYMUPDF = False
    except ImportError:
        PDF_SUPPORT = False
        print("Warning: PDF reading not available. Install with: pip install PyMuPDF")

# Excel writing
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import ScatterChart, Reference, Series
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("Warning: Excel writing not available. Install with: pip install openpyxl")

import matplotlib.pyplot as plt


# ============================================================================
# CONFIGURATION
# ============================================================================
DEFAULT_FOLDER = r"F:\iCloudDrive\UoU school\SPRING 2026\Week4"
DEFAULT_RF_RATE = 0.0005  # 0.05% monthly


# ============================================================================
# PDF READER
# ============================================================================
def read_pdf(file_path: str) -> str:
    """
    Read text content from a PDF file.

    Args:
        file_path: Path to PDF file

    Returns:
        Extracted text content
    """
    if not PDF_SUPPORT:
        return "PDF reading not available. Please install PyMuPDF: pip install PyMuPDF"

    try:
        # Try PyMuPDF first (better quality)
        import fitz
        doc = fitz.open(file_path)
        text = ""
        for page in doc:
            text += page.get_text()
        doc.close()
        return text
    except:
        try:
            # Fall back to PyPDF2
            from PyPDF2 import PdfReader
            reader = PdfReader(file_path)
            text = ""
            for page in reader.pages:
                text += page.extract_text() or ""
            return text
        except Exception as e:
            return f"Error reading PDF: {e}"


def parse_homework_instructions(pdf_text: str) -> dict:
    """
    Parse homework instructions from PDF text.

    Looks for:
    - Stock symbols
    - Target standard deviations
    - Risk-free rate
    - Required calculations

    Args:
        pdf_text: Extracted text from PDF

    Returns:
        Dictionary with parsed instructions
    """
    instructions = {
        'stocks': [],
        'target_stds': [],
        'rf_rate': DEFAULT_RF_RATE,
        'questions': [],
        'raw_text': pdf_text
    }

    # Common stock symbols to look for
    common_symbols = ['AAPL', 'AXP', 'BA', 'CAT', 'CSCO', 'CVX', 'DIS', 'DD',
                      'GS', 'HD', 'IBM', 'INTC', 'JNJ', 'JPM', 'KO', 'MCD',
                      'MMM', 'MRK', 'MSFT', 'NKE', 'PFE', 'PG', 'TRV', 'UNH',
                      'UTX', 'V', 'VZ', 'WBA', 'WMT', 'XOM', 'SPY', 'GOOGL',
                      'AMZN', 'META', 'NVDA', 'TSLA']

    # Find mentioned stocks
    text_upper = pdf_text.upper()
    for symbol in common_symbols:
        if symbol in text_upper:
            instructions['stocks'].append(symbol)

    # Look for target standard deviations (e.g., "4%", "7%", "st dev of 4%")
    import re
    std_patterns = [
        r'st\.?\s*dev\.?\s*(?:of\s+)?(\d+(?:\.\d+)?)\s*%',
        r'standard\s+deviation\s*(?:of\s+)?(\d+(?:\.\d+)?)\s*%',
        r'stdev\s*=?\s*(\d+(?:\.\d+)?)\s*%',
        r'(\d+(?:\.\d+)?)\s*%\s*(?:st\.?\s*dev|standard\s+deviation)',
    ]

    for pattern in std_patterns:
        matches = re.findall(pattern, pdf_text.lower())
        for match in matches:
            try:
                std_val = float(match) / 100
                if 0.01 <= std_val <= 0.20 and std_val not in instructions['target_stds']:
                    instructions['target_stds'].append(std_val)
            except:
                pass

    # Look for risk-free rate
    rf_patterns = [
        r'rf\s*=?\s*(\d+(?:\.\d+)?)\s*%',
        r'risk.?free\s*(?:rate)?\s*(?:of\s+)?(\d+(?:\.\d+)?)\s*%',
    ]

    for pattern in rf_patterns:
        matches = re.findall(pattern, pdf_text.lower())
        if matches:
            try:
                instructions['rf_rate'] = float(matches[0]) / 100
            except:
                pass

    # Sort target stds
    instructions['target_stds'] = sorted(instructions['target_stds'])

    return instructions


# ============================================================================
# DATA LOADER
# ============================================================================
def find_data_files(folder: str) -> dict:
    """
    Find Excel and CSV files in the folder.

    Args:
        folder: Path to search

    Returns:
        Dictionary of found files by type
    """
    files = {
        'excel': [],
        'csv': [],
        'pdf': []
    }

    folder_path = Path(folder)
    if not folder_path.exists():
        print(f"Warning: Folder does not exist: {folder}")
        return files

    for file in folder_path.iterdir():
        if file.suffix.lower() in ['.xlsx', '.xls']:
            files['excel'].append(str(file))
        elif file.suffix.lower() == '.csv':
            files['csv'].append(str(file))
        elif file.suffix.lower() == '.pdf':
            files['pdf'].append(str(file))

    return files


def load_return_data(file_path: str, sheet_name: str = None) -> tuple:
    """
    Load return data from Excel or CSV file.

    Auto-detects if data is prices or returns and converts accordingly.

    Args:
        file_path: Path to data file
        sheet_name: Sheet name for Excel files

    Returns:
        Tuple of (expected_returns, cov_matrix, asset_names, raw_returns_df)
    """
    path = Path(file_path)

    # Load data
    if path.suffix.lower() in ['.xlsx', '.xls']:
        if sheet_name:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            xl = pd.ExcelFile(file_path)
            print(f"  Available sheets: {xl.sheet_names}")
            # Try common sheet names
            for name in ['Returns', 'Data', 'Sheet1', 'HW2', xl.sheet_names[0]]:
                if name in xl.sheet_names:
                    df = pd.read_excel(file_path, sheet_name=name)
                    print(f"  Using sheet: {name}")
                    break
    else:
        df = pd.read_csv(file_path)

    # Find and remove date column
    date_col = None
    for col in df.columns:
        if 'date' in str(col).lower() or df[col].dtype == 'datetime64[ns]':
            date_col = col
            break

    if date_col:
        df = df.drop(columns=[date_col])

    # Keep only numeric columns
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    df = df[numeric_cols]

    # Remove columns that look like formulas or calculations
    cols_to_keep = []
    for col in df.columns:
        col_str = str(col).lower()
        if not any(x in col_str for x in ['portfolio', 'return', 'formula', 'weight']):
            cols_to_keep.append(col)
    df = df[cols_to_keep]

    # Clean data
    df = df.dropna()

    # Detect if prices or returns
    sample_mean = df.mean().mean()
    if abs(sample_mean) > 2:  # Likely prices
        print("  Detected: PRICE data - converting to log returns")
        returns_df = np.log(df / df.shift(1)).dropna()
    else:
        print("  Detected: RETURN data")
        returns_df = df

    # Compute statistics
    asset_names = list(returns_df.columns)
    expected_returns = returns_df.mean().values
    cov_matrix = returns_df.cov().values * (len(returns_df) - 1) / len(returns_df)  # Population cov

    return expected_returns, cov_matrix, asset_names, returns_df


# ============================================================================
# PORTFOLIO OPTIMIZER
# ============================================================================
class PortfolioOptimizer:
    """Portfolio optimization using Modern Portfolio Theory."""

    def __init__(self, expected_returns, cov_matrix, asset_names, rf_rate=0.0005):
        self.expected_returns = np.array(expected_returns)
        self.cov_matrix = np.array(cov_matrix)
        self.asset_names = asset_names
        self.rf_rate = rf_rate
        self.n_assets = len(expected_returns)

    def portfolio_return(self, w):
        return np.dot(w, self.expected_returns)

    def portfolio_variance(self, w):
        return np.dot(w, np.dot(self.cov_matrix, w))

    def portfolio_std(self, w):
        return np.sqrt(self.portfolio_variance(w))

    def portfolio_sharpe(self, w):
        ret = self.portfolio_return(w)
        std = self.portfolio_std(w)
        return (ret - self.rf_rate) / std if std > 1e-10 else 0

    def minimum_variance_portfolio(self, allow_short=True):
        w0 = np.ones(self.n_assets) / self.n_assets
        constraints = [{'type': 'eq', 'fun': lambda w: np.sum(w) - 1}]
        bounds = None if allow_short else [(0, 1) for _ in range(self.n_assets)]

        result = minimize(self.portfolio_variance, w0, method='SLSQP',
                         bounds=bounds, constraints=constraints, options={'ftol': 1e-12})

        w = result.x
        return w, {
            'return': self.portfolio_return(w),
            'std': self.portfolio_std(w),
            'variance': self.portfolio_variance(w),
            'sharpe': self.portfolio_sharpe(w)
        }

    def tangent_portfolio(self, allow_short=True):
        w0 = np.ones(self.n_assets) / self.n_assets
        constraints = [{'type': 'eq', 'fun': lambda w: np.sum(w) - 1}]
        bounds = None if allow_short else [(0, 1) for _ in range(self.n_assets)]

        def neg_sharpe(w):
            std = self.portfolio_std(w)
            if std < 1e-10:
                return 1e10
            return -(self.portfolio_return(w) - self.rf_rate) / std

        result = minimize(neg_sharpe, w0, method='SLSQP',
                         bounds=bounds, constraints=constraints, options={'ftol': 1e-12})

        w = result.x
        return w, {
            'return': self.portfolio_return(w),
            'std': self.portfolio_std(w),
            'variance': self.portfolio_variance(w),
            'sharpe': self.portfolio_sharpe(w)
        }

    def optimize_for_target_std(self, target_std, allow_short=True):
        w0 = np.ones(self.n_assets) / self.n_assets
        constraints = [
            {'type': 'eq', 'fun': lambda w: np.sum(w) - 1},
            {'type': 'eq', 'fun': lambda w: self.portfolio_std(w) - target_std}
        ]
        bounds = None if allow_short else [(0, 1) for _ in range(self.n_assets)]

        result = minimize(lambda w: -self.portfolio_return(w), w0, method='SLSQP',
                         bounds=bounds, constraints=constraints, options={'ftol': 1e-12})

        if not result.success:
            return None, None

        w = result.x
        return w, {
            'return': self.portfolio_return(w),
            'std': self.portfolio_std(w),
            'variance': self.portfolio_variance(w),
            'sharpe': self.portfolio_sharpe(w)
        }

    def two_fund_frontier(self, w1, w2, n_points=200, lambda_range=(-1, 2.5)):
        mu1, sigma1 = self.portfolio_return(w1), self.portfolio_std(w1)
        mu2, sigma2 = self.portfolio_return(w2), self.portfolio_std(w2)
        cov12 = np.dot(w1, np.dot(self.cov_matrix, w2))

        lambdas = np.linspace(lambda_range[0], lambda_range[1], n_points)
        returns, stds = [], []

        for lam in lambdas:
            mu_p = lam * mu1 + (1 - lam) * mu2
            var_p = lam**2 * sigma1**2 + (1-lam)**2 * sigma2**2 + 2*lam*(1-lam)*cov12
            if var_p >= 0:
                returns.append(mu_p)
                stds.append(np.sqrt(var_p))

        return np.array(returns), np.array(stds)

    def capital_market_line(self, tan_w, n_points=100, max_leverage=2.5):
        tan_ret = self.portfolio_return(tan_w)
        tan_std = self.portfolio_std(tan_w)

        weights = np.linspace(0, max_leverage, n_points)
        returns = weights * tan_ret + (1 - weights) * self.rf_rate
        stds = weights * tan_std

        return returns, stds

    def superportfolio(self, w1, w2, lambda_val):
        w_super = lambda_val * w1 + (1 - lambda_val) * w2

        mu1, sigma1 = self.portfolio_return(w1), self.portfolio_std(w1)
        mu2, sigma2 = self.portfolio_return(w2), self.portfolio_std(w2)
        cov12 = np.dot(w1, np.dot(self.cov_matrix, w2))

        super_ret = lambda_val * mu1 + (1 - lambda_val) * mu2
        super_var = lambda_val**2 * sigma1**2 + (1-lambda_val)**2 * sigma2**2 + 2*lambda_val*(1-lambda_val)*cov12

        return w_super, {
            'return': super_ret,
            'std': np.sqrt(super_var),
            'variance': super_var,
            'cov12': cov12
        }

    def cml_portfolio(self, tan_w, weight_tangent):
        tan_ret = self.portfolio_return(tan_w)
        tan_std = self.portfolio_std(tan_w)

        return {
            'return': weight_tangent * tan_ret + (1 - weight_tangent) * self.rf_rate,
            'std': weight_tangent * tan_std
        }


# ============================================================================
# EXCEL GENERATOR
# ============================================================================
def generate_excel_solution(optimizer, results, output_path, instructions):
    """
    Generate Excel file with step-by-step solution and formulas.

    Args:
        optimizer: PortfolioOptimizer instance
        results: Dictionary with all computed results
        output_path: Path to save Excel file
        instructions: Parsed homework instructions
    """
    if not EXCEL_SUPPORT:
        print("Excel generation not available. Install openpyxl: pip install openpyxl")
        return

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    number_format = '0.0000%'
    decimal_format = '0.000000'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ==================== SHEET 1: INPUT DATA ====================
    ws1 = wb.active
    ws1.title = "Input Data"

    # Asset Statistics
    ws1['A1'] = "ASSET STATISTICS"
    ws1['A1'].font = header_font

    headers = ['Asset', 'Mean Return', 'Std Dev', 'Variance']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=2, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for i, name in enumerate(optimizer.asset_names):
        row = i + 3
        ws1.cell(row=row, column=1, value=name).border = thin_border
        ws1.cell(row=row, column=2, value=optimizer.expected_returns[i]).number_format = number_format
        ws1.cell(row=row, column=2).border = thin_border
        ws1.cell(row=row, column=3, value=np.sqrt(optimizer.cov_matrix[i,i])).number_format = number_format
        ws1.cell(row=row, column=3).border = thin_border
        ws1.cell(row=row, column=4, value=optimizer.cov_matrix[i,i]).number_format = decimal_format
        ws1.cell(row=row, column=4).border = thin_border

    # Covariance Matrix
    cov_start_row = len(optimizer.asset_names) + 5
    ws1.cell(row=cov_start_row, column=1, value="COVARIANCE MATRIX").font = header_font

    # Headers
    ws1.cell(row=cov_start_row+1, column=1, value="").border = thin_border
    for j, name in enumerate(optimizer.asset_names):
        cell = ws1.cell(row=cov_start_row+1, column=j+2, value=name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    # Matrix values
    for i, name in enumerate(optimizer.asset_names):
        cell = ws1.cell(row=cov_start_row+2+i, column=1, value=name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

        for j in range(len(optimizer.asset_names)):
            cell = ws1.cell(row=cov_start_row+2+i, column=j+2, value=optimizer.cov_matrix[i,j])
            cell.number_format = decimal_format
            cell.border = thin_border

    # Risk-free rate
    rf_row = cov_start_row + len(optimizer.asset_names) + 4
    ws1.cell(row=rf_row, column=1, value="Risk-Free Rate:").font = header_font
    ws1.cell(row=rf_row, column=2, value=optimizer.rf_rate).number_format = number_format

    # ==================== SHEET 2: MVP ====================
    ws2 = wb.create_sheet("MVP")

    ws2['A1'] = "MINIMUM VARIANCE PORTFOLIO"
    ws2['A1'].font = Font(bold=True, size=14)

    ws2['A3'] = "Optimization Problem:"
    ws2['A4'] = "Minimize: w'Σw (portfolio variance)"
    ws2['A5'] = "Subject to: Σw = 1 (weights sum to 100%)"

    mvp_w = results['mvp']['weights']
    mvp_stats = results['mvp']['stats']

    # Weights
    ws2['A7'] = "WEIGHTS"
    ws2['A7'].font = header_font

    headers = ['Asset', 'Weight', 'Weight %']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=8, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for i, name in enumerate(optimizer.asset_names):
        row = i + 9
        ws2.cell(row=row, column=1, value=name).border = thin_border
        ws2.cell(row=row, column=2, value=mvp_w[i]).number_format = decimal_format
        ws2.cell(row=row, column=2).border = thin_border
        ws2.cell(row=row, column=3, value=mvp_w[i]).number_format = number_format
        ws2.cell(row=row, column=3).border = thin_border

    sum_row = 9 + len(optimizer.asset_names)
    ws2.cell(row=sum_row, column=1, value="SUM").font = header_font
    ws2.cell(row=sum_row, column=2, value=f"=SUM(B9:B{sum_row-1})")
    ws2.cell(row=sum_row, column=3, value=f"=SUM(C9:C{sum_row-1})").number_format = number_format

    # Stats
    stats_row = sum_row + 2
    ws2.cell(row=stats_row, column=1, value="PORTFOLIO STATISTICS").font = header_font
    ws2.cell(row=stats_row+1, column=1, value="Expected Return:")
    ws2.cell(row=stats_row+1, column=2, value=mvp_stats['return']).number_format = number_format
    ws2.cell(row=stats_row+2, column=1, value="Standard Deviation:")
    ws2.cell(row=stats_row+2, column=2, value=mvp_stats['std']).number_format = number_format
    ws2.cell(row=stats_row+3, column=1, value="Variance:")
    ws2.cell(row=stats_row+3, column=2, value=mvp_stats['variance']).number_format = decimal_format
    ws2.cell(row=stats_row+4, column=1, value="Sharpe Ratio:")
    ws2.cell(row=stats_row+4, column=2, value=mvp_stats['sharpe']).number_format = '0.0000'

    # Formulas explanation
    formula_row = stats_row + 7
    ws2.cell(row=formula_row, column=1, value="FORMULAS").font = header_font
    ws2.cell(row=formula_row+1, column=1, value="Return = Σ(w_i × μ_i) = SUMPRODUCT(weights, returns)")
    ws2.cell(row=formula_row+2, column=1, value="Variance = w'Σw = MMULT(MMULT(TRANSPOSE(w), Σ), w)")
    ws2.cell(row=formula_row+3, column=1, value="Std Dev = SQRT(Variance)")
    ws2.cell(row=formula_row+4, column=1, value="Sharpe = (Return - RF) / Std Dev")

    # ==================== SHEET 3: TANGENT ====================
    ws3 = wb.create_sheet("Tangent")

    ws3['A1'] = "TANGENT PORTFOLIO (Maximum Sharpe Ratio)"
    ws3['A1'].font = Font(bold=True, size=14)

    ws3['A3'] = "Optimization Problem:"
    ws3['A4'] = "Maximize: (μ_p - RF) / σ_p (Sharpe Ratio)"
    ws3['A5'] = "Subject to: Σw = 1"

    tan_w = results['tangent']['weights']
    tan_stats = results['tangent']['stats']

    # Weights
    ws3['A7'] = "WEIGHTS"
    ws3['A7'].font = header_font

    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=8, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    for i, name in enumerate(optimizer.asset_names):
        row = i + 9
        ws3.cell(row=row, column=1, value=name).border = thin_border
        ws3.cell(row=row, column=2, value=tan_w[i]).number_format = decimal_format
        ws3.cell(row=row, column=2).border = thin_border
        ws3.cell(row=row, column=3, value=tan_w[i]).number_format = number_format
        ws3.cell(row=row, column=3).border = thin_border

    sum_row = 9 + len(optimizer.asset_names)
    ws3.cell(row=sum_row, column=1, value="SUM").font = header_font
    ws3.cell(row=sum_row, column=2, value=f"=SUM(B9:B{sum_row-1})")

    # Stats
    stats_row = sum_row + 2
    ws3.cell(row=stats_row, column=1, value="PORTFOLIO STATISTICS").font = header_font
    ws3.cell(row=stats_row+1, column=1, value="Expected Return:")
    ws3.cell(row=stats_row+1, column=2, value=tan_stats['return']).number_format = number_format
    ws3.cell(row=stats_row+2, column=1, value="Standard Deviation:")
    ws3.cell(row=stats_row+2, column=2, value=tan_stats['std']).number_format = number_format
    ws3.cell(row=stats_row+3, column=1, value="Sharpe Ratio:")
    ws3.cell(row=stats_row+3, column=2, value=tan_stats['sharpe']).number_format = '0.0000'

    # ==================== SHEET 4: EFFICIENT PORTFOLIOS ====================
    ws4 = wb.create_sheet("Efficient Portfolios")

    ws4['A1'] = "EFFICIENT PORTFOLIOS AT TARGET STANDARD DEVIATIONS"
    ws4['A1'].font = Font(bold=True, size=14)

    current_row = 3
    for target_std, eff_data in results.get('efficient', {}).items():
        ws4.cell(row=current_row, column=1, value=f"Target Std Dev = {target_std*100:.0f}%").font = header_font
        current_row += 1

        eff_w = eff_data['weights']
        eff_stats = eff_data['stats']

        # Weights header
        for col, header in enumerate(['Asset', 'Weight', 'Weight %'], 1):
            cell = ws4.cell(row=current_row, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
        current_row += 1

        for i, name in enumerate(optimizer.asset_names):
            ws4.cell(row=current_row, column=1, value=name).border = thin_border
            ws4.cell(row=current_row, column=2, value=eff_w[i]).number_format = decimal_format
            ws4.cell(row=current_row, column=2).border = thin_border
            ws4.cell(row=current_row, column=3, value=eff_w[i]).number_format = number_format
            ws4.cell(row=current_row, column=3).border = thin_border
            current_row += 1

        # Stats
        current_row += 1
        ws4.cell(row=current_row, column=1, value="Return:")
        ws4.cell(row=current_row, column=2, value=eff_stats['return']).number_format = number_format
        current_row += 1
        ws4.cell(row=current_row, column=1, value="Std Dev:")
        ws4.cell(row=current_row, column=2, value=eff_stats['std']).number_format = number_format
        current_row += 3

    # ==================== SHEET 5: TWO-FUND SEPARATION ====================
    ws5 = wb.create_sheet("Two-Fund Separation")

    ws5['A1'] = "TWO-FUND SEPARATION THEOREM"
    ws5['A1'].font = Font(bold=True, size=14)

    ws5['A3'] = "Theory: Any efficient portfolio = λ × P1 + (1-λ) × P2"
    ws5['A5'] = "FORMULAS:"
    ws5['A6'] = "Return: μ_combined = λ × μ_1 + (1-λ) × μ_2"
    ws5['A7'] = "Variance: σ²_combined = λ²σ₁² + (1-λ)²σ₂² + 2λ(1-λ)Cov(P1,P2)"
    ws5['A8'] = "where Cov(P1,P2) = w1' × Σ × w2"

    if 'superportfolio' in results:
        super_data = results['superportfolio']

        ws5['A10'] = f"SUPERPORTFOLIO: {super_data['lambda']*100:.0f}% in P1 + {(1-super_data['lambda'])*100:.0f}% in P2"
        ws5['A10'].font = header_font

        ws5['A12'] = "Calculation:"
        ws5['A13'] = f"λ = {super_data['lambda']}"
        ws5['A14'] = f"Cov(P1,P2) = {super_data['stats']['cov12']:.8f}"
        ws5['A15'] = f"Variance = {super_data['lambda']}² × σ₁² + {1-super_data['lambda']}² × σ₂² + 2×{super_data['lambda']}×{1-super_data['lambda']}×Cov"
        ws5['A16'] = f"Std Dev = {super_data['stats']['std']*100:.4f}%"

    # ==================== SHEET 6: CML ====================
    ws6 = wb.create_sheet("Capital Market Line")

    ws6['A1'] = "CAPITAL MARKET LINE (CML)"
    ws6['A1'].font = Font(bold=True, size=14)

    ws6['A3'] = "Theory: CML combines risk-free asset with tangent portfolio"
    ws6['A5'] = "FORMULAS:"
    ws6['A6'] = "Return: μ_CML = w_t × μ_tangent + (1 - w_t) × RF"
    ws6['A7'] = "Std Dev: σ_CML = w_t × σ_tangent (RF has zero risk)"

    if 'cml_portfolio' in results:
        cml_data = results['cml_portfolio']

        ws6['A9'] = f"CML PORTFOLIO: {(1-cml_data['weight_tangent'])*100:.0f}% RF + {cml_data['weight_tangent']*100:.0f}% Tangent"
        ws6['A9'].font = header_font

        ws6['A11'] = "Calculation:"
        ws6['A12'] = f"Return = {cml_data['weight_tangent']} × {tan_stats['return']*100:.4f}% + {1-cml_data['weight_tangent']} × {optimizer.rf_rate*100:.4f}%"
        ws6['A13'] = f"Return = {cml_data['return']*100:.4f}%"
        ws6['A14'] = f"Std Dev = {cml_data['weight_tangent']} × {tan_stats['std']*100:.4f}%"
        ws6['A15'] = f"Std Dev = {cml_data['std']*100:.4f}%"

    # ==================== SHEET 7: FINAL ANSWERS ====================
    ws7 = wb.create_sheet("Final Answers")

    ws7['A1'] = "FINAL ANSWERS SUMMARY"
    ws7['A1'].font = Font(bold=True, size=14)

    ws7['A3'] = "Question"
    ws7['B3'] = "Answer"
    ws7['C3'] = "Derivation"
    for col in range(1, 4):
        cell = ws7.cell(row=3, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    row = 4

    # Answer 1: Mean at higher std
    if results.get('efficient'):
        stds = sorted(results['efficient'].keys())
        if len(stds) >= 2:
            higher_std = stds[-1]
            ws7.cell(row=row, column=1, value=f"Mean of Eff({higher_std*100:.0f}%)").border = thin_border
            ws7.cell(row=row, column=2, value=results['efficient'][higher_std]['stats']['return']).number_format = number_format
            ws7.cell(row=row, column=2).border = thin_border
            ws7.cell(row=row, column=3, value="Optimization for target std").border = thin_border
            row += 1

    # Answer 2: MVP Std Dev
    ws7.cell(row=row, column=1, value="MVP Std Dev").border = thin_border
    ws7.cell(row=row, column=2, value=mvp_stats['std']).number_format = number_format
    ws7.cell(row=row, column=2).border = thin_border
    ws7.cell(row=row, column=3, value="Min variance optimization").border = thin_border
    row += 1

    # Answer 3: Superportfolio Std
    if 'superportfolio' in results:
        ws7.cell(row=row, column=1, value="Superportfolio Std").border = thin_border
        ws7.cell(row=row, column=2, value=results['superportfolio']['stats']['std']).number_format = number_format
        ws7.cell(row=row, column=2).border = thin_border
        ws7.cell(row=row, column=3, value="Two-fund separation formula").border = thin_border
        row += 1

    # Answer 4: Tangent Sharpe
    ws7.cell(row=row, column=1, value="Tangent Sharpe Ratio").border = thin_border
    ws7.cell(row=row, column=2, value=tan_stats['sharpe']).number_format = '0.0000'
    ws7.cell(row=row, column=2).border = thin_border
    ws7.cell(row=row, column=3, value="Max Sharpe optimization").border = thin_border
    row += 1

    # Answer 5: CML Portfolio Return
    if 'cml_portfolio' in results:
        ws7.cell(row=row, column=1, value="CML Portfolio Return").border = thin_border
        ws7.cell(row=row, column=2, value=results['cml_portfolio']['return']).number_format = number_format
        ws7.cell(row=row, column=2).border = thin_border
        ws7.cell(row=row, column=3, value="CML linear combination").border = thin_border

    # Adjust column widths
    for ws in wb.worksheets:
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 18

    # Save
    wb.save(output_path)
    print(f"Excel solution saved to: {output_path}")


# ============================================================================
# PYTHON SOLUTION GENERATOR
# ============================================================================
def generate_python_solution(optimizer, results, output_path, instructions):
    """
    Generate Python solution file.

    Args:
        optimizer: PortfolioOptimizer instance
        results: Dictionary with all computed results
        output_path: Path to save Python file
        instructions: Parsed homework instructions
    """

    # Build the Python code
    code = f'''"""
================================================================================
PORTFOLIO OPTIMIZATION SOLUTION
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
================================================================================

This script solves the portfolio optimization homework.

Assets: {optimizer.asset_names}
Risk-Free Rate: {optimizer.rf_rate*100:.4f}% per period
Target Std Devs: {[f"{s*100:.0f}%" for s in instructions.get('target_stds', [0.04, 0.07])]}

================================================================================
"""

import numpy as np
from scipy.optimize import minimize

# ============================================================================
# INPUT DATA
# ============================================================================

# Asset names
asset_names = {optimizer.asset_names}

# Expected returns (mean returns)
expected_returns = np.array({list(optimizer.expected_returns)})

# Covariance matrix
cov_matrix = np.array({optimizer.cov_matrix.tolist()})

# Risk-free rate
rf_rate = {optimizer.rf_rate}  # {optimizer.rf_rate*100:.4f}% per period

n_assets = len(asset_names)

# ============================================================================
# PORTFOLIO FUNCTIONS
# ============================================================================

def portfolio_return(w):
    """Calculate portfolio return: μ_p = w' × μ"""
    return np.dot(w, expected_returns)

def portfolio_variance(w):
    """Calculate portfolio variance: σ²_p = w' × Σ × w"""
    return np.dot(w, np.dot(cov_matrix, w))

def portfolio_std(w):
    """Calculate portfolio standard deviation"""
    return np.sqrt(portfolio_variance(w))

def portfolio_sharpe(w):
    """Calculate Sharpe ratio: (μ_p - RF) / σ_p"""
    ret = portfolio_return(w)
    std = portfolio_std(w)
    return (ret - rf_rate) / std if std > 1e-10 else 0

# ============================================================================
# OPTIMIZATION FUNCTIONS
# ============================================================================

def minimum_variance_portfolio():
    """Find the Minimum Variance Portfolio (MVP)"""
    w0 = np.ones(n_assets) / n_assets
    constraints = [{{'type': 'eq', 'fun': lambda w: np.sum(w) - 1}}]
    result = minimize(portfolio_variance, w0, method='SLSQP',
                     constraints=constraints, options={{'ftol': 1e-12}})
    return result.x

def tangent_portfolio():
    """Find the Tangent Portfolio (Maximum Sharpe Ratio)"""
    w0 = np.ones(n_assets) / n_assets
    constraints = [{{'type': 'eq', 'fun': lambda w: np.sum(w) - 1}}]
    def neg_sharpe(w):
        std = portfolio_std(w)
        return -(portfolio_return(w) - rf_rate) / std if std > 1e-10 else 1e10
    result = minimize(neg_sharpe, w0, method='SLSQP',
                     constraints=constraints, options={{'ftol': 1e-12}})
    return result.x

def optimize_for_target_std(target_std):
    """Find efficient portfolio at target standard deviation"""
    w0 = np.ones(n_assets) / n_assets
    constraints = [
        {{'type': 'eq', 'fun': lambda w: np.sum(w) - 1}},
        {{'type': 'eq', 'fun': lambda w: portfolio_std(w) - target_std}}
    ]
    result = minimize(lambda w: -portfolio_return(w), w0, method='SLSQP',
                     constraints=constraints, options={{'ftol': 1e-12}})
    return result.x if result.success else None

# ============================================================================
# RUN ANALYSIS
# ============================================================================

if __name__ == "__main__":
    print("=" * 70)
    print("PORTFOLIO OPTIMIZATION SOLUTION")
    print("=" * 70)

    # Print asset statistics
    print("\\n--- Asset Statistics ---")
    print(f"{{'Asset':<10}} {{'Mean':>12}} {{'Std Dev':>12}}")
    print("-" * 36)
    for i, name in enumerate(asset_names):
        mean = expected_returns[i]
        std = np.sqrt(cov_matrix[i,i])
        print(f"{{name:<10}} {{mean*100:>11.4f}}% {{std*100:>11.4f}}%")

    # MVP
    print("\\n--- Minimum Variance Portfolio ---")
    mvp_w = minimum_variance_portfolio()
    print("Weights:")
    for i, name in enumerate(asset_names):
        print(f"  {{name}}: {{mvp_w[i]*100:>8.2f}}%")
    print(f"\\nReturn: {{portfolio_return(mvp_w)*100:.4f}}%")
    print(f"Std Dev: {{portfolio_std(mvp_w)*100:.4f}}%")
    print(f"Sharpe: {{portfolio_sharpe(mvp_w):.4f}}")

    # Tangent
    print("\\n--- Tangent Portfolio ---")
    tan_w = tangent_portfolio()
    print("Weights:")
    for i, name in enumerate(asset_names):
        print(f"  {{name}}: {{tan_w[i]*100:>8.2f}}%")
    print(f"\\nReturn: {{portfolio_return(tan_w)*100:.4f}}%")
    print(f"Std Dev: {{portfolio_std(tan_w)*100:.4f}}%")
    print(f"Sharpe: {{portfolio_sharpe(tan_w):.4f}}")

    # Efficient portfolios at target stds
    target_stds = {instructions.get('target_stds', [0.04, 0.07])}
    eff_portfolios = {{}}

    for target in target_stds:
        print(f"\\n--- Efficient Portfolio at {{target*100:.0f}}% Std ---")
        w = optimize_for_target_std(target)
        if w is not None:
            eff_portfolios[target] = w
            print("Weights:")
            for i, name in enumerate(asset_names):
                print(f"  {{name}}: {{w[i]*100:>8.2f}}%")
            print(f"\\nReturn: {{portfolio_return(w)*100:.4f}}%")
            print(f"Std Dev: {{portfolio_std(w)*100:.4f}}%")

    # Superportfolio (30/70)
    if len(eff_portfolios) >= 2:
        stds = sorted(eff_portfolios.keys())
        w1, w2 = eff_portfolios[stds[0]], eff_portfolios[stds[1]]
        lambda_val = 0.30

        print(f"\\n--- Superportfolio ({{lambda_val*100:.0f}}%/{{(1-lambda_val)*100:.0f}}%) ---")

        # Two-fund separation formula
        mu1, sigma1 = portfolio_return(w1), portfolio_std(w1)
        mu2, sigma2 = portfolio_return(w2), portfolio_std(w2)
        cov12 = np.dot(w1, np.dot(cov_matrix, w2))

        super_var = lambda_val**2 * sigma1**2 + (1-lambda_val)**2 * sigma2**2 + 2*lambda_val*(1-lambda_val)*cov12
        super_std = np.sqrt(super_var)

        print(f"Cov(P1,P2) = {{cov12:.8f}}")
        print(f"Std Dev = {{super_std*100:.4f}}%")

    # CML Portfolio (30% RF + 70% Tangent)
    print("\\n--- CML Portfolio (30% RF + 70% Tangent) ---")
    w_tan = 0.70
    cml_ret = w_tan * portfolio_return(tan_w) + (1 - w_tan) * rf_rate
    cml_std = w_tan * portfolio_std(tan_w)
    print(f"Return: {{cml_ret*100:.4f}}%")
    print(f"Std Dev: {{cml_std*100:.4f}}%")

    # Final Answers
    print("\\n" + "=" * 70)
    print("FINAL ANSWERS")
    print("=" * 70)

    if len(eff_portfolios) >= 2:
        higher_std = sorted(eff_portfolios.keys())[-1]
        print(f"1. Mean of Eff({{higher_std*100:.0f}}%): {{portfolio_return(eff_portfolios[higher_std])*100:.4f}}%")

    print(f"2. MVP Std Dev: {{portfolio_std(mvp_w)*100:.4f}}%")

    if len(eff_portfolios) >= 2:
        print(f"3. Superportfolio Std: {{super_std*100:.4f}}%")

    print(f"4. Tangent Sharpe: {{portfolio_sharpe(tan_w):.4f}}")
    print(f"5. CML Portfolio Return: {{cml_ret*100:.4f}}%")
'''

    with open(output_path, 'w') as f:
        f.write(code)

    print(f"Python solution saved to: {output_path}")


# ============================================================================
# GRAPH GENERATOR
# ============================================================================
def generate_graph(optimizer, results, output_path):
    """Generate efficient frontier graph."""

    fig, ax = plt.subplots(figsize=(14, 10))

    mvp_w = results['mvp']['weights']
    tan_w = results['tangent']['weights']
    mvp_stats = results['mvp']['stats']
    tan_stats = results['tangent']['stats']

    # Efficient frontier
    if results.get('efficient') and len(results['efficient']) >= 2:
        stds = sorted(results['efficient'].keys())
        w1 = results['efficient'][stds[0]]['weights']
        w2 = results['efficient'][stds[1]]['weights']
        frontier_ret, frontier_std = optimizer.two_fund_frontier(w1, w2)
        ax.plot(frontier_std * 100, frontier_ret * 100, 'b-', linewidth=2.5,
               label='Efficient Frontier', zorder=2)

    # CML
    cml_ret, cml_std = optimizer.capital_market_line(tan_w)
    ax.plot(cml_std * 100, cml_ret * 100, 'g--', linewidth=2.5,
           label='Capital Market Line', zorder=2)

    # Individual assets
    colors = plt.cm.Set2(np.linspace(0, 1, len(optimizer.asset_names)))
    for i, name in enumerate(optimizer.asset_names):
        std = np.sqrt(optimizer.cov_matrix[i, i]) * 100
        ret = optimizer.expected_returns[i] * 100
        ax.scatter(std, ret, s=150, c=[colors[i]], edgecolors='black',
                  linewidths=1.5, zorder=5)
        ax.annotate(name, (std, ret), textcoords='offset points',
                   xytext=(8, 5), fontsize=10, fontweight='bold')

    # Risk-free rate
    ax.scatter(0, optimizer.rf_rate * 100, s=200, c='gold', edgecolors='black',
              linewidths=2, marker='*', zorder=6,
              label=f'Risk-Free ({optimizer.rf_rate*100:.2f}%)')

    # MVP
    ax.scatter(mvp_stats['std'] * 100, mvp_stats['return'] * 100,
              s=250, c='red', edgecolors='black', linewidths=2,
              marker='s', zorder=6)
    ax.annotate('MVP', (mvp_stats['std'] * 100, mvp_stats['return'] * 100),
               textcoords='offset points', xytext=(10, -5),
               fontsize=12, fontweight='bold', color='red')

    # Tangent
    ax.scatter(tan_stats['std'] * 100, tan_stats['return'] * 100,
              s=250, c='green', edgecolors='black', linewidths=2,
              marker='^', zorder=6)
    ax.annotate('Tangent\n(Max Sharpe)', (tan_stats['std'] * 100, tan_stats['return'] * 100),
               textcoords='offset points', xytext=(10, 5),
               fontsize=11, fontweight='bold', color='green')

    # Efficient portfolios
    colors_eff = ['purple', 'orange', 'cyan', 'magenta']
    for idx, (target_std, eff) in enumerate(results.get('efficient', {}).items()):
        ax.scatter(eff['stats']['std'] * 100, eff['stats']['return'] * 100,
                  s=200, c=colors_eff[idx % len(colors_eff)], edgecolors='black',
                  linewidths=2, marker='D', zorder=6)
        ax.annotate(f"Eff({target_std*100:.0f}%)",
                   (eff['stats']['std'] * 100, eff['stats']['return'] * 100),
                   textcoords='offset points', xytext=(10, -10),
                   fontsize=10, fontweight='bold', color=colors_eff[idx % len(colors_eff)])

    ax.set_xlabel('Standard Deviation (%)', fontsize=14, fontweight='bold')
    ax.set_ylabel('Expected Return (%)', fontsize=14, fontweight='bold')
    ax.set_title('Efficient Frontier & Capital Market Line\nPortfolio Optimization Solution',
                fontsize=16, fontweight='bold', pad=20)
    ax.grid(True, alpha=0.3, linestyle='--')
    ax.legend(loc='upper left', fontsize=10)

    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches='tight', facecolor='white')
    print(f"Graph saved to: {output_path}")
    plt.close()


# ============================================================================
# MAIN FUNCTION
# ============================================================================
def main():
    """Main entry point."""

    print("\n" + "=" * 70)
    print("   HOMEWORK ASSISTANT - Portfolio Analysis Automation")
    print("=" * 70)

    # Get folder location
    print(f"\nDefault folder: {DEFAULT_FOLDER}")
    folder_input = input("Enter folder path (or press Enter for default): ").strip()
    folder = folder_input if folder_input else DEFAULT_FOLDER
    folder = folder.strip('"').strip("'")

    print(f"\nUsing folder: {folder}")

    # Find files
    print("\nSearching for files...")
    files = find_data_files(folder)

    print(f"  Found {len(files['pdf'])} PDF files")
    print(f"  Found {len(files['excel'])} Excel files")
    print(f"  Found {len(files['csv'])} CSV files")

    # Read PDF instructions
    instructions = {'target_stds': [0.04, 0.07], 'rf_rate': DEFAULT_RF_RATE, 'stocks': []}

    if files['pdf']:
        print("\n--- Reading PDF Instructions ---")
        for pdf_file in files['pdf']:
            print(f"  Reading: {Path(pdf_file).name}")
            pdf_text = read_pdf(pdf_file)
            parsed = parse_homework_instructions(pdf_text)

            # Merge parsed instructions
            if parsed['stocks']:
                instructions['stocks'] = parsed['stocks']
            if parsed['target_stds']:
                instructions['target_stds'] = parsed['target_stds']
            if parsed['rf_rate'] != DEFAULT_RF_RATE:
                instructions['rf_rate'] = parsed['rf_rate']

            # Show extracted info
            print(f"    Stocks mentioned: {parsed['stocks']}")
            print(f"    Target stds: {[f'{s*100:.0f}%' for s in parsed['target_stds']]}")
            print(f"    RF rate: {parsed['rf_rate']*100:.4f}%")

    # Ask for RF rate confirmation
    print(f"\nRisk-free rate detected: {instructions['rf_rate']*100:.4f}%")
    rf_input = input("Enter new RF rate (as %, e.g., 0.05) or press Enter to accept: ").strip()
    if rf_input:
        try:
            instructions['rf_rate'] = float(rf_input) / 100
        except:
            print("Invalid input, using detected rate.")

    # Ask for target stds confirmation
    print(f"Target standard deviations: {[f'{s*100:.0f}%' for s in instructions['target_stds']]}")
    std_input = input("Enter new targets (comma-separated, e.g., 4,7) or press Enter to accept: ").strip()
    if std_input:
        try:
            instructions['target_stds'] = sorted([float(s.strip())/100 for s in std_input.split(',')])
        except:
            print("Invalid input, using detected targets.")

    # Load data
    print("\n--- Loading Data ---")
    data_file = None

    if files['excel']:
        print("Excel files found:")
        for i, f in enumerate(files['excel']):
            print(f"  {i+1}. {Path(f).name}")

        choice = input("Select file number (or press Enter for first): ").strip()
        idx = int(choice) - 1 if choice.isdigit() else 0
        data_file = files['excel'][min(idx, len(files['excel'])-1)]
    elif files['csv']:
        data_file = files['csv'][0]

    if not data_file:
        print("No data files found. Please add Excel or CSV data to the folder.")
        return

    print(f"Loading: {Path(data_file).name}")

    # Ask for sheet name if Excel
    sheet_name = None
    if data_file.endswith(('.xlsx', '.xls')):
        xl = pd.ExcelFile(data_file)
        print(f"Available sheets: {xl.sheet_names}")
        sheet_input = input("Enter sheet name (or press Enter for auto-detect): ").strip()
        sheet_name = sheet_input if sheet_input else None

    # Load the data
    expected_returns, cov_matrix, asset_names, returns_df = load_return_data(data_file, sheet_name)

    print(f"Loaded {len(asset_names)} assets: {asset_names}")

    # Filter to specific stocks if mentioned in instructions
    if instructions['stocks']:
        matching = [s for s in instructions['stocks'] if s in asset_names]
        if matching and len(matching) >= 2:
            print(f"Filtering to mentioned stocks: {matching}")
            # Would need to implement filtering here

    # Create optimizer
    optimizer = PortfolioOptimizer(expected_returns, cov_matrix, asset_names, instructions['rf_rate'])

    # Run analysis
    print("\n--- Running Optimization ---")

    results = {}

    # MVP
    print("Computing MVP...")
    mvp_w, mvp_stats = optimizer.minimum_variance_portfolio()
    results['mvp'] = {'weights': mvp_w, 'stats': mvp_stats}

    # Tangent
    print("Computing Tangent Portfolio...")
    tan_w, tan_stats = optimizer.tangent_portfolio()
    results['tangent'] = {'weights': tan_w, 'stats': tan_stats}

    # Efficient portfolios at target stds
    results['efficient'] = {}
    for target_std in instructions['target_stds']:
        print(f"Computing Efficient Portfolio at {target_std*100:.0f}%...")
        eff_w, eff_stats = optimizer.optimize_for_target_std(target_std)
        if eff_w is not None:
            results['efficient'][target_std] = {'weights': eff_w, 'stats': eff_stats}

    # Superportfolio
    if len(results['efficient']) >= 2:
        stds = sorted(results['efficient'].keys())
        w1 = results['efficient'][stds[0]]['weights']
        w2 = results['efficient'][stds[1]]['weights']
        super_w, super_stats = optimizer.superportfolio(w1, w2, 0.30)
        results['superportfolio'] = {
            'weights': super_w,
            'stats': super_stats,
            'lambda': 0.30,
            'stds': stds
        }

    # CML Portfolio
    cml_port = optimizer.cml_portfolio(tan_w, 0.70)
    cml_port['weight_tangent'] = 0.70
    results['cml_portfolio'] = cml_port

    # Generate outputs
    print("\n--- Generating Outputs ---")

    output_dir = Path(folder) / "solution"
    output_dir.mkdir(exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # Python solution
    py_path = output_dir / f"solution_{timestamp}.py"
    generate_python_solution(optimizer, results, str(py_path), instructions)

    # Excel solution
    excel_path = output_dir / f"solution_{timestamp}.xlsx"
    generate_excel_solution(optimizer, results, str(excel_path), instructions)

    # Graph
    graph_path = output_dir / f"efficient_frontier_{timestamp}.png"
    generate_graph(optimizer, results, str(graph_path))

    # Print final answers
    print("\n" + "=" * 70)
    print("FINAL ANSWERS")
    print("=" * 70)

    if results.get('efficient'):
        stds = sorted(results['efficient'].keys())
        if len(stds) >= 2:
            higher_std = stds[-1]
            print(f"1. Mean of Eff({higher_std*100:.0f}%): {results['efficient'][higher_std]['stats']['return']*100:.4f}%")

    print(f"2. MVP Std Dev: {mvp_stats['std']*100:.4f}%")

    if 'superportfolio' in results:
        print(f"3. Superportfolio Std: {results['superportfolio']['stats']['std']*100:.4f}%")

    print(f"4. Tangent Sharpe: {tan_stats['sharpe']:.4f}")
    print(f"5. CML Portfolio Return: {results['cml_portfolio']['return']*100:.4f}%")

    print(f"\nOutputs saved to: {output_dir}")


if __name__ == "__main__":
    main()
